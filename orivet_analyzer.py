#!/usr/bin/env python3
"""
Orivet 遺伝子検査PDF 自動解析ツール
====================================
Orivet (Paw Print Genetics) の Genetic Summary Report PDF を自動解析し、
分かりやすいHTMLレポートとExcelファイルを出力します。

使い方:
    python orivet_analyzer.py *.pdf
    python orivet_analyzer.py folder_with_pdfs/
    python orivet_analyzer.py dog1.pdf dog2.pdf dog3.pdf

出力:
    - orivet_report.html  (インタラクティブHTMLレポート)
    - orivet_report.xlsx  (Excelスプレッドシート)

必要ライブラリ:
    pip install pdfplumber openpyxl
"""

import sys
import os
import re
import glob
import json
from dataclasses import dataclass, field
from typing import Optional
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    print("エラー: pdfplumber が必要です。\n  pip install pdfplumber")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("エラー: openpyxl が必要です。\n  pip install openpyxl")
    sys.exit(1)


# ============================================================
# データ構造
# ============================================================

@dataclass
class TestResult:
    """個別の検査結果"""
    category: str          # カテゴリー (例: Musculoskeletal, Trait)
    test_name: str         # 検査名
    genotype: str          # 遺伝子型 (例: N/N, P/P, P/N)
    result_text: str       # 結果テキスト全文
    status: str            # normal / carrier / positive / trait
    japanese_name: str = "" # 日本語名

@dataclass
class DogProfile:
    """犬1頭分のプロファイル"""
    pet_name: str = ""
    registered_name: str = ""
    breed: str = ""
    sex: str = ""
    dob: str = ""
    colour: str = ""
    microchip: str = ""
    case_number: str = ""
    owner_name: str = ""
    test_date: str = ""
    test_requested: str = ""
    approved_collection: str = ""
    sample_type: str = ""
    health_results: list = field(default_factory=list)
    trait_results: list = field(default_factory=list)
    source_file: str = ""


# ============================================================
# 日本語マッピング
# ============================================================

CATEGORY_JP = {
    "musculoskeletal": "筋骨格系",
    "haemolymphatic": "血液・リンパ系",
    "nervous": "神経系",
    "neurologic": "神経系",
    "metabolic": "代謝系",
    "ophthalmologic": "眼科系",
    "cardiorespiratory": "心肺系",
    "cardiovascular": "心血管系",
    "dermatologic": "皮膚科系",
    "immunological": "免疫系",
    "respiratory": "呼吸器系",
    "urogenital": "泌尿生殖器系",
    "trait": "形質（毛色・外見）",
}

TEST_NAME_JP = {
    "chondrodystrophy with intervertebral disc disease": "軟骨異栄養症+椎間板疾患リスク (CDDY+IVDD)",
    "cddy with ivdd": "軟骨異栄養症+椎間板疾患リスク (CDDY+IVDD)",
    "osteochondrodysplasia": "骨軟骨異形成症",
    "chondrodysplasia": "軟骨異形成症 (CDPA)",
    "cdpa": "軟骨異形成症 (CDPA)",
    "congenital macrothrombocytopenia": "先天性巨大血小板減少症",
    "congenital methemoglobinemia": "先天性メトヘモグロビン血症",
    "von willebrand": "フォンウィルブランド病 I型",
    "degenerative myelopathy": "変性性脊髄症 (DM)",
    "gangliosidosis gm2": "ガングリオシドーシス GM2",
    "progressive rod cone degeneration": "進行性網膜萎縮症 (prcd-PRA)",
    "prcd": "進行性網膜萎縮症 (prcd-PRA)",
    "progressive retinal atrophy": "進行性網膜萎縮症 (PRA)",
    "exercise-induced collapse": "運動誘発性虚脱 (EIC)",
    "neonatal encephalopathy": "新生児脳症",
    "hyperuricosuria": "高尿酸尿症",
    "cystinuria": "シスチン尿症",
    "a locus": "A座位 (アグーチ)",
    "b locus": "B座位 (ブラウン)",
    "d locus": "D座位 (ダイリュート)",
    "dilute": "D座位 (ダイリュート)",
    "e locus": "E座位 (エクステンション)",
    "em locus": "EM座位 (メラニスティックマスク)",
    "mc1r": "EM座位 (メラニスティックマスク)",
    "k locus": "K座位 (ドミナントブラック)",
    "m locus": "M座位 (マール/ダップル)",
    "merle": "M座位 (マール/ダップル)",
    "curly coat": "巻き毛 (Curly Coat)",
    "furnishings": "ファーニシング (RSPO2)",
    "rspo2": "ファーニシング (RSPO2)",
    "pied": "パイド (Pied)",
    "brown tyrp1": "ブラウン TYRP1",
    "improper coat": "インプロパーコート",
    "coat length": "毛の長さ",
    "shedding": "換毛",
}


# ============================================================
# PDF解析エンジン
# ============================================================

def extract_all_text(pdf_path: str) -> str:
    """PDFから全ページのテキストを抽出"""
    texts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
    return "\n\n".join(texts)


def parse_animal_details(text: str) -> dict:
    """動物の基本情報を抽出"""
    info = {}
    patterns = {
        "registered_name": r"Registered\s+Name\s*:?\s*(.+?)(?:\n|$)",
        "pet_name": r"Pet\s+Name\s*:?\s*(.+?)(?:\n|$)",
        "breed": r"Breed\s*:?\s*:?\s*(.+?)(?:\n|$)",
        "microchip": r"Microchip\s+Number\s*:?\s*(\d[\d\s]*\d)",
        "sex": r"Sex\s*:?\s*:?\s*(.+?)(?:\n|$)",
        "dob": r"Date\s+of\s+Birth\s*:?\s*(.+?)(?:\n|$)",
        "colour": r"Colour\s*:?\s*(.+?)(?:\n|$)",
        "case_number": r"Case\s+Number\s*:?\s*(\S+)",
        "owner_name": r"(?:Owner|Name)\s*:?\s*([\w\s]+Kamide|[\w\s]+(?:san|sama))",
        "test_date": r"Date\s+of\s+Test\s*:?\s*(.+?)(?:\n|$)",
        "test_requested": r"Test\s+Requested\s*:?\s*(.+?)(?:\n|$)",
        "approved_collection": r"Approved\s+Collection\s*(?:Method)?\s*:?\s*(Yes|No)",
        "sample_type": r"Sample\s+Type\s*:?\s*(\S+)",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            val = match.group(1).strip()
            # Clean up common artifacts
            val = re.sub(r'\s+', ' ', val).strip()
            if val and val.lower() not in (':', ''):
                info[key] = val

    # Also try Animal Name from cover page
    if "pet_name" not in info:
        m = re.search(r"Animal\s+Name\s*:?\s*(.+?)(?:\n|$)", text, re.IGNORECASE)
        if m:
            info["pet_name"] = m.group(1).strip()

    # Owner name fallback
    if "owner_name" not in info:
        m = re.search(r"Name\s*:\s*(.+?)(?:\n|$)", text)
        if m:
            info["owner_name"] = m.group(1).strip()

    return info


def classify_result(result_text: str) -> str:
    """結果テキストからステータスを分類"""
    text_upper = result_text.upper()
    if "POSITIVE (P/P)" in text_upper or "TWO COPIES" in text_upper:
        return "positive"
    elif "CARRIER (P/N)" in text_upper or "ONE COPY OF THE" in text_upper:
        return "carrier"
    elif "NORMAL (N/N)" in text_upper or "NO VARIANT DETECTED" in text_upper:
        return "normal"
    elif "POSITIVE HETEROZYGOUS" in text_upper:
        return "carrier"
    elif "POSITIVE" in text_upper and "P/P" in text_upper:
        return "positive"
    elif "CARRIER" in text_upper:
        return "carrier"
    elif "NORMAL" in text_upper:
        return "normal"
    else:
        return "trait"


def get_japanese_name(test_name: str) -> str:
    """検査名から日本語名を取得"""
    name_lower = test_name.lower()
    for key, jp_name in TEST_NAME_JP.items():
        if key in name_lower:
            return jp_name
    return ""


def get_category_jp(category: str) -> str:
    """カテゴリーの日本語名を取得"""
    cat_lower = category.lower()
    for key, jp_name in CATEGORY_JP.items():
        if key in cat_lower:
            return jp_name
    return category


def extract_genotype(result_text: str) -> str:
    """結果テキストから遺伝子型を抽出"""
    # P/P, N/N, P/N patterns
    m = re.search(r'\b([PN])/([PN])\b', result_text)
    if m:
        return f"{m.group(1)}/{m.group(2)}"

    # Specific genotype patterns for traits
    patterns = [
        r'(at/at|ay/at|ay/ay|a/a|aw/at)',           # A locus
        r'(Bb|BB|bb)\b',                              # B locus
        r'(D/D|D/d|d/d)\b',                          # D locus
        r'(E/e|e/e|E/E|Em/E|Em/e)\b',               # E locus
        r'(En/En|EM/EM)',                             # EM locus
        r'(K/K|KB/ky|KB/kbr|ky/ky|kbr/ky)',         # K locus
        r'(m/m|M/m|M/M)',                             # M locus
        r'(Cu/Cu|Cu/N|N/N)',                          # Curly
        r'(F/F|F/f|f/f)',                             # Furnishings
        r'(S/S|S/sp|sp/sp)',                          # Pied
        r'(BL/BL|BL/bs|bs/bs)',                       # TYRP1
    ]
    for p in patterns:
        m = re.search(p, result_text, re.IGNORECASE)
        if m:
            return m.group(1)

    # Try broader extraction
    m = re.search(r'^([A-Za-z/\[\]\d\s]{1,30})\s*[-–]', result_text)
    if m:
        return m.group(1).strip()

    return ""


def parse_health_tests(text: str) -> list:
    """健康検査と形質検査の結果を解析"""
    results = []

    # Known categories with their section headers
    category_headers = [
        ("Musculoskeletal", r"Musculoskeletal"),
        ("Haemolymphatic", r"Haemolymphatic"),
        ("Nervous system / Neurologic", r"Nervous\s+system|Neurologic"),
        ("Metabolic", r"Metabolic"),
        ("Ophthalmologic", r"Ophthalmologic"),
        ("Cardiorespiratory", r"Cardiorespiratory"),
        ("Cardiovascular", r"Cardiovascular"),
        ("Dermatologic", r"Dermatologic"),
        ("Immunological", r"Immunological"),
        ("Respiratory", r"Respiratory"),
        ("Urogenital", r"Urogenital"),
    ]

    # Parse trait results (page 3 typically)
    # Pattern: test_name followed by result on same or next line
    trait_patterns = [
        (r"A\s+Locus\s*\(Agouti\)(.*?)(?=\n[A-Z]|\n\n)", "A Locus (Agouti)"),
        (r"B\s+Locus\s*[-–]\s*(?:Bd,?\s*Bs,?\s*Bc\s*\[Various Breeds\])?(.*?)(?=\n[A-Z]|\n\n)", "B Locus (Brown)"),
        (r"Chondrodysplasia\s*\(CDPA\)(.*?)(?=\n[A-Z]|\n\n)", "Chondrodysplasia (CDPA)"),
        (r"Curly\s+Coat/Hair\s+Variant\s+1?(.*?)(?=\n[A-Z]|\n\n)", "Curly Coat/Hair Variant 1"),
        (r"D\s*\(Dilute\)\s+Locus(.*?)(?=\n[A-Z]|\n\n)", "D (Dilute) Locus"),
        (r"E\s+Locus\s*[-–]\s*\(Cream/Red/Yellow\)(.*?)(?=\n[A-Z]|\n\n)", "E Locus (Cream/Red/Yellow)"),
        (r"EM\s*\(MC1R\)\s+Locus\s*[-–]\s*Melanistic\s+Mask(.*?)(?=\n[A-Z]|\n\n)", "EM (MC1R) Locus - Melanistic Mask"),
        (r"Furnishings\s*\(RSPO2\)(.*?)(?=\n[A-Z]|\n\n)", "Furnishings (RSPO2)"),
        (r"K\s+Locus\s*\(Dominant\s+Black\)(.*?)(?=\n[A-Z]|\n\n)", "K Locus (Dominant Black)"),
        (r"M\s+Locus\s*\(Merle/Dapple\)(.*?)(?=\n[A-Z]|\n\n)", "M Locus (Merle/Dapple)"),
        (r"Pied\s*\(BOTH\s+SINE\s+and\s+REPEAT\s+VARIANTS?\)(.*?)(?=\n[A-Z]|\n\n)", "Pied"),
        (r"Brown\s+TYRP1\s*\[.*?\]\s*=\s*Bl\s+(.*?)(?=\n[A-Z]|\n\n)", "Brown TYRP1"),
    ]

    # Strategy: Use line-by-line parsing for more robust extraction
    lines = text.split('\n')
    current_category = "Trait"
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        # Check for category headers
        for cat_name, cat_pattern in category_headers:
            if re.search(cat_pattern, line, re.IGNORECASE):
                current_category = cat_name
                break

        # Check for Trait section
        if "Trait (Associated with Phenotype)" in line:
            current_category = "Trait"

        # Parse NORMAL/CARRIER/POSITIVE results
        if re.search(r'NORMAL\s*\(N/N\)|CARRIER\s*\(P/N\)|POSITIVE\s*\(P/P\)', line, re.IGNORECASE):
            # Try to extract test name and result from this line
            # Often format: "Test Name RESULT_TEXT"
            m = re.match(r'(.+?)\s+((?:NORMAL|CARRIER|POSITIVE)\s*\([PN]/[PN]\).+)', line, re.IGNORECASE)
            if m:
                test_name = m.group(1).strip()
                result_text = m.group(2).strip()
            else:
                # Result might be on next line, test name on prev line
                test_name = lines[i-1].strip() if i > 0 else ""
                result_text = line.strip()

            # Clean test name
            test_name = re.sub(r'^[\s\uf0b7\u2022\u25cf]+', '', test_name)  # Remove bullet points
            test_name = re.sub(r'\s+', ' ', test_name).strip()

            if test_name and len(test_name) > 2:
                status = classify_result(result_text)
                genotype = extract_genotype(result_text)
                jp_name = get_japanese_name(test_name)
                cat_jp = get_category_jp(current_category)

                result = TestResult(
                    category=cat_jp,
                    test_name=test_name,
                    genotype=genotype,
                    result_text=result_text,
                    status=status,
                    japanese_name=jp_name,
                )

                if current_category == "Trait":
                    # Skip, handle separately
                    pass
                else:
                    results.append(result)

        i += 1

    # 重複除去（extract_text + extract_tables で同じ結果が2回取得される場合がある）
    seen = set()
    unique_results = []
    for r in results:
        key = (r.test_name, r.status, r.genotype)
        if key not in seen:
            seen.add(key)
            unique_results.append(r)

    return unique_results


def parse_trait_results_from_text(text: str) -> list:
    """形質（毛色）結果をテキストから解析"""
    results = []

    # Known trait test items and their result patterns
    # We look for lines containing the test name followed by the result
    trait_items = [
        ("A Locus (Agouti)", r"A\s+Locus\s*\(Agouti\)"),
        ("B Locus (Brown)", r"B\s+Locus\s*[-–]?\s*(?:Bd|Bs|Bc|Various)"),
        ("Chondrodysplasia (CDPA)", r"Chondrodysplasia\s*\(CDPA\)"),
        ("Curly Coat/Hair Variant 1", r"Curly\s+Coat"),
        ("D (Dilute) Locus", r"D\s*\(?Dilute\)?\s+Locus"),
        ("E Locus (Cream/Red/Yellow)", r"E\s+Locus\s*[-–]?\s*\(?Cream"),
        ("EM (MC1R) - Melanistic Mask", r"EM\s*\(MC1R\)|Melanistic\s+Mask"),
        ("Furnishings (RSPO2)", r"Furnishings\s*\(RSPO2\)"),
        ("K Locus (Dominant Black)", r"K\s+Locus\s*\(Dominant"),
        ("M Locus (Merle/Dapple)", r"M\s+Locus\s*\(Merle"),
        ("Pied", r"Pied\s*\(BOTH\s+SINE"),
        ("Brown TYRP1", r"Brown\s+TYRP1"),
        ("Improper Coat", r"Improper\s+Coat"),
        ("Coat Length", r"Coat\s+Length"),
    ]

    lines = text.split('\n')

    for test_name, pattern in trait_items:
        for i, line in enumerate(lines):
            if re.search(pattern, line, re.IGNORECASE):
                # Get the result - might be on same line or next line
                result_text = line
                # Also check next 1-2 lines for continuation
                for j in range(1, 3):
                    if i + j < len(lines):
                        next_line = lines[i + j].strip()
                        # If next line looks like a continuation (not a new test)
                        if next_line and not re.search(r'^[A-Z]\s+Locus|^Breed|^Owner|^Microchip|^Pied|^Brown|^Curly|^Furnish|^Chondro', next_line):
                            result_text += " " + next_line
                        else:
                            break

                # Remove the test name prefix to get just the result
                result_clean = re.sub(pattern, '', result_text, flags=re.IGNORECASE).strip()
                result_clean = re.sub(r'^[\s\-–:]+', '', result_clean).strip()

                genotype = extract_genotype(result_text)
                status = classify_result(result_text)
                if status == "normal" and test_name != "Chondrodysplasia (CDPA)":
                    status = "trait"  # Most trait results are informational

                jp_name = get_japanese_name(test_name)

                results.append(TestResult(
                    category="形質（毛色・外見）",
                    test_name=test_name,
                    genotype=genotype,
                    result_text=result_clean if result_clean else result_text,
                    status=status,
                    japanese_name=jp_name,
                ))
                break  # Found this trait, move to next

    return results


def parse_pdf(pdf_path: str) -> Optional[DogProfile]:
    """PDFファイル1つを解析してDogProfileを返す"""
    # Skip DNAP (DNA Profile) files - they contain marker data, not health/trait results
    basename = os.path.basename(pdf_path)
    if "DNAP" in basename.upper() or "DNA PROFILE" in basename.upper():
        # Check if it's actually a DNA Profile (marker data only)
        text = extract_all_text(pdf_path)
        if "ISAG Profile" in text or "DNA Profile" in text:
            if "Health Tests Reported" not in text:
                print(f"  スキップ (DNAプロファイル): {basename}")
                return None

    # Skip guide/instruction PDFs
    if "見方" in basename or "説明" in basename:
        print(f"  スキップ (ガイド): {basename}")
        return None

    print(f"  解析中: {basename}")

    text = extract_all_text(pdf_path)

    # Must contain Genetic Summary Report or Health Tests
    if "Genetic Summary Report" not in text and "Health Tests Reported" not in text:
        print(f"  → Orivet Genetic Summary Report ではありません。スキップします。")
        return None

    # Extract basic info
    info = parse_animal_details(text)
    if not info.get("pet_name") and not info.get("registered_name"):
        print(f"  → 動物情報を検出できませんでした。スキップします。")
        return None

    # Parse health results
    health_results = parse_health_tests(text)

    # Parse trait results
    trait_results = parse_trait_results_from_text(text)

    dog = DogProfile(
        pet_name=info.get("pet_name", ""),
        registered_name=info.get("registered_name", ""),
        breed=info.get("breed", ""),
        sex=info.get("sex", ""),
        dob=info.get("dob", ""),
        colour=info.get("colour", ""),
        microchip=info.get("microchip", ""),
        case_number=info.get("case_number", ""),
        owner_name=info.get("owner_name", ""),
        test_date=info.get("test_date", ""),
        test_requested=info.get("test_requested", ""),
        approved_collection=info.get("approved_collection", ""),
        sample_type=info.get("sample_type", ""),
        health_results=health_results,
        trait_results=trait_results,
        source_file=basename,
    )

    print(f"  → {dog.pet_name or dog.registered_name}: 健康{len(health_results)}項目, 形質{len(trait_results)}項目")
    return dog


# ============================================================
# HTML出力
# ============================================================

def status_badge(status: str, text: str) -> str:
    return f'<span class="status {status}">{text}</span>'


def generate_html(dogs: list, output_path: str):
    """インタラクティブHTMLレポートを生成"""

    # Count totals
    total_normal = sum(len([r for r in d.health_results if r.status == "normal"]) for d in dogs)
    total_carrier = sum(len([r for r in d.health_results if r.status == "carrier"]) for d in dogs)
    total_positive = sum(len([r for r in d.health_results if r.status == "positive"]) for d in dogs)

    # Build dog tabs HTML
    tab_buttons = ""
    tab_contents = ""

    for idx, dog in enumerate(dogs):
        name = dog.pet_name or dog.registered_name or f"犬{idx+1}"
        safe_id = re.sub(r'[^a-zA-Z0-9]', '_', name.lower())

        tab_buttons += f'    <div class="tab" onclick="showTab(\'{safe_id}\')">{name}</div>\n'

        sex_class = "male" if "male" in dog.sex.lower() else "female"
        sex_label = "オス" if "male" in dog.sex.lower() else "メス"

        # Health results table
        health_rows = ""
        for r in dog.health_results:
            display_name = r.japanese_name if r.japanese_name else r.test_name
            badge = status_badge(r.status, r.genotype if r.genotype else r.status.upper())
            health_rows += f"""        <tr>
          <td>{r.category}</td>
          <td>{display_name}<br><small style="color:#6b7280">{r.test_name}</small></td>
          <td>{badge}</td>
          <td style="font-size:0.85em">{r.result_text[:120]}</td>
        </tr>\n"""

        # Trait results table
        trait_rows = ""
        for r in dog.trait_results:
            display_name = r.japanese_name if r.japanese_name else r.test_name
            badge = status_badge("trait", r.genotype if r.genotype else "—")
            trait_rows += f"""        <tr>
          <td>{display_name}<br><small style="color:#6b7280">{r.test_name}</small></td>
          <td>{badge}</td>
          <td style="font-size:0.85em">{r.result_text[:150]}</td>
        </tr>\n"""

        tab_contents += f"""
  <div id="{safe_id}" class="tab-content">
    <div class="dog-card">
      <div class="dog-header">
        <div>
          <div class="dog-name">{name}</div>
          <div class="dog-reg">{dog.registered_name} — {dog.case_number}</div>
        </div>
        <div class="dog-meta">
          <span class="meta-tag {sex_class}">{sex_label} ({dog.sex})</span>
          <span class="meta-tag">{dog.breed}</span>
          <span class="meta-tag">{dog.dob}</span>
          {'<span class="meta-tag">MC: ' + dog.microchip + '</span>' if dog.microchip else ''}
          {'<span class="meta-tag">毛色: ' + dog.colour + '</span>' if dog.colour else ''}
        </div>
      </div>

      <h3 class="section-title">健康検査結果 ({len(dog.health_results)}項目)</h3>
      <table class="results-table">
        <tr><th>カテゴリー</th><th>検査項目</th><th>結果</th><th>詳細</th></tr>
{health_rows}
      </table>

      <h3 class="section-title">毛色・形質検査結果 ({len(dog.trait_results)}項目)</h3>
      <table class="results-table">
        <tr><th>検査項目</th><th>遺伝子型</th><th>詳細</th></tr>
{trait_rows}
      </table>
    </div>
  </div>"""

    # Build comparison table
    compare_health_rows = ""
    if len(dogs) > 1:
        # Collect all unique health test names
        all_tests = {}
        for dog in dogs:
            for r in dog.health_results:
                key = r.test_name
                if key not in all_tests:
                    all_tests[key] = {"jp": r.japanese_name or r.test_name, "name": r.test_name}

        compare_header = "<th>検査項目</th>"
        for dog in dogs:
            compare_header += f"<th>{dog.pet_name or dog.registered_name}</th>"

        for test_key, test_info in all_tests.items():
            row = f"<td>{test_info['jp']}</td>"
            for dog in dogs:
                found = False
                for r in dog.health_results:
                    if r.test_name == test_key:
                        badge = status_badge(r.status, r.genotype if r.genotype else r.status.upper())
                        row += f"<td>{badge}</td>"
                        found = True
                        break
                if not found:
                    row += "<td>—</td>"
            compare_health_rows += f"        <tr>{row}</tr>\n"

    # Alerts section
    alerts_html = ""
    for dog in dogs:
        for r in dog.health_results:
            name = dog.pet_name or dog.registered_name
            display_name = r.japanese_name if r.japanese_name else r.test_name
            if r.status == "positive":
                alerts_html += f"""      <div class="breed-warn danger">
        <div class="warn-title">{display_name} — ポジティブ (P/P): {name}</div>
        <p>変異が2コピー検出されました。発症リスクがあります。獣医師にご相談の上、適切なケアをお願いいたします。繁殖にも十分な注意が必要です。</p>
        <p><small>原文: {r.result_text[:200]}</small></p>
      </div>\n"""
            elif r.status == "carrier":
                alerts_html += f"""      <div class="breed-warn">
        <div class="warn-title">{display_name} — キャリア (P/N): {name}</div>
        <p>変異が1コピー検出されました。自身の発症リスクは低いですが、キャリアまたはポジティブの個体との交配で発症する子犬が生まれる可能性があります。</p>
      </div>\n"""

    if not alerts_html:
        alerts_html = '<div class="breed-warn" style="background:#dcfce7;border-color:#86efac;"><div class="warn-title" style="color:#166534;">全頭クリア</div><p>全ての健康検査項目でノーマル（変異なし）でした。</p></div>'

    # Pre-build overview table rows (can't use backslash in f-strings)
    overview_table_rows = ""
    for d in dogs:
        overview_table_rows += f"<tr><td><strong>{d.pet_name}</strong></td><td>{d.registered_name}</td><td>{d.breed}</td><td>{d.sex}</td><td>{d.dob}</td><td>{d.case_number}</td></tr>\n"

    # Pre-build compare tab
    compare_tab_button = ""
    compare_tab_html = ""
    if len(dogs) > 1:
        compare_tab_button = '<div class="tab" onclick="showTab(\'compare\')">比較表</div>'
        compare_tab_html = f"""<div id="compare" class="tab-content">
    <div class="dog-card">
      <h2 class="section-title">健康検査 比較表</h2>
      <div style="overflow-x:auto;">
      <table class="compare-table">
        <tr>{compare_header}</tr>
{compare_health_rows}
      </table>
      </div>
    </div>
  </div>"""

    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Orivet 遺伝子検査レポート</title>
<style>
:root {{ --pink:#e6007e; --purple:#4a1a7a; --green:#22c55e; --yellow:#eab308; --red:#ef4444; --gray:#6b7280; --bg:#f8f9fa; }}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI','Hiragino Sans','Meiryo',sans-serif; background:var(--bg); color:#1f2937; line-height:1.6; }}
.container {{ max-width:1200px; margin:0 auto; padding:20px; }}
header {{ background:linear-gradient(135deg,var(--purple),var(--pink)); color:white; padding:30px 0; margin-bottom:30px; border-radius:0 0 20px 20px; }}
header .container {{ display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px; }}
header h1 {{ font-size:1.8em; }}
header p {{ opacity:0.9; font-size:0.95em; }}
.badge {{ display:inline-block; background:rgba(255,255,255,0.2); padding:4px 12px; border-radius:20px; font-size:0.85em; }}
.summary-row {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:15px; margin-bottom:30px; }}
.summary-card {{ background:white; border-radius:12px; padding:20px; text-align:center; box-shadow:0 2px 8px rgba(0,0,0,0.06); }}
.summary-card .num {{ font-size:2em; font-weight:700; }}
.summary-card .label {{ font-size:0.85em; color:var(--gray); margin-top:4px; }}
.num.green {{ color:var(--green); }} .num.yellow {{ color:var(--yellow); }} .num.red {{ color:var(--red); }} .num.blue {{ color:#3b82f6; }}
.tabs {{ display:flex; gap:8px; margin-bottom:20px; flex-wrap:wrap; }}
.tab {{ padding:10px 20px; border-radius:10px 10px 0 0; cursor:pointer; font-weight:600; border:2px solid #e5e7eb; border-bottom:none; background:white; transition:all 0.2s; font-size:0.95em; }}
.tab:hover {{ background:#f3e8ff; }}
.tab.active {{ background:var(--purple); color:white; border-color:var(--purple); }}
.tab-content {{ display:none; }} .tab-content.active {{ display:block; }}
.dog-card {{ background:white; border-radius:16px; padding:24px; margin-bottom:24px; box-shadow:0 2px 12px rgba(0,0,0,0.06); }}
.dog-header {{ display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:15px; margin-bottom:20px; padding-bottom:16px; border-bottom:2px solid #f3f4f6; }}
.dog-name {{ font-size:1.4em; font-weight:700; color:var(--purple); }}
.dog-reg {{ font-size:0.85em; color:var(--gray); }}
.dog-meta {{ display:flex; gap:12px; flex-wrap:wrap; }}
.meta-tag {{ background:#f3f4f6; padding:4px 12px; border-radius:20px; font-size:0.82em; white-space:nowrap; }}
.meta-tag.male {{ background:#dbeafe; color:#1e40af; }} .meta-tag.female {{ background:#fce7f3; color:#be185d; }}
.section-title {{ font-size:1.1em; font-weight:700; color:var(--purple); margin:20px 0 12px; padding-left:10px; border-left:4px solid var(--pink); }}
.results-table {{ width:100%; border-collapse:separate; border-spacing:0; font-size:0.88em; }}
.results-table th {{ background:var(--purple); color:white; padding:10px 12px; text-align:left; font-weight:600; }}
.results-table th:first-child {{ border-radius:8px 0 0 0; }} .results-table th:last-child {{ border-radius:0 8px 0 0; }}
.results-table td {{ padding:10px 12px; border-bottom:1px solid #f3f4f6; vertical-align:top; }}
.results-table tr:hover td {{ background:#faf5ff; }}
.compare-table {{ width:100%; border-collapse:separate; border-spacing:0; font-size:0.85em; }}
.compare-table th {{ background:var(--purple); color:white; padding:10px; text-align:center; position:sticky; top:0; }}
.compare-table th:first-child {{ text-align:left; min-width:200px; }}
.compare-table td {{ padding:8px 10px; border-bottom:1px solid #f3f4f6; text-align:center; }}
.compare-table td:first-child {{ text-align:left; font-weight:500; }}
.compare-table tr:hover td {{ background:#faf5ff; }}
.status {{ display:inline-block; padding:3px 10px; border-radius:12px; font-size:0.82em; font-weight:600; white-space:nowrap; }}
.status.normal {{ background:#dcfce7; color:#166534; }}
.status.carrier {{ background:#fef3c7; color:#92400e; }}
.status.positive {{ background:#fee2e2; color:#991b1b; }}
.status.trait {{ background:#e0e7ff; color:#3730a3; }}
.legend {{ display:flex; gap:16px; flex-wrap:wrap; margin:16px 0; padding:12px 16px; background:white; border-radius:10px; font-size:0.85em; }}
.legend-item {{ display:flex; align-items:center; gap:6px; }}
.legend-dot {{ width:14px; height:14px; border-radius:4px; }}
.breed-warn {{ background:#fff7ed; border:1px solid #fed7aa; border-radius:10px; padding:16px; margin:12px 0; }}
.breed-warn.danger {{ background:#fef2f2; border-color:#fecaca; }}
.breed-warn .warn-title {{ font-weight:700; color:#c2410c; margin-bottom:4px; }}
.breed-warn.danger .warn-title {{ color:#dc2626; }}
.print-btn {{ background:var(--purple); color:white; border:none; padding:10px 20px; border-radius:8px; cursor:pointer; font-weight:600; }}
.print-btn:hover {{ background:var(--pink); }}
@media (max-width:768px) {{ header h1 {{ font-size:1.3em; }} .dog-header {{ flex-direction:column; }} .compare-table {{ display:block; overflow-x:auto; }} }}
@media print {{ .tabs,.print-btn,header {{ display:none!important; }} .tab-content {{ display:block!important; page-break-inside:avoid; }} .dog-card {{ break-inside:avoid; }} }}
</style>
</head>
<body>
<header>
  <div class="container">
    <div>
      <h1>Orivet 遺伝子検査レポート</h1>
      <p>検査機関: Orivet Genetics (Paw Print Genetics) &nbsp;|&nbsp; ISO/IEC 17025準拠 &nbsp;|&nbsp; ISAG会員</p>
      <p><span class="badge">生成日: {datetime.now().strftime('%Y年%m月%d日')}</span> <span class="badge">{len(dogs)}頭分</span></p>
    </div>
    <button class="print-btn" onclick="window.print()">印刷</button>
  </div>
</header>
<div class="container">
  <div class="summary-row">
    <div class="summary-card"><div class="num blue">{len(dogs)}</div><div class="label">検査頭数</div></div>
    <div class="summary-card"><div class="num green">{total_normal}</div><div class="label">ノーマル項目</div></div>
    <div class="summary-card"><div class="num yellow">{total_carrier}</div><div class="label">キャリア項目</div></div>
    <div class="summary-card"><div class="num red">{total_positive}</div><div class="label">ポジティブ (要注意)</div></div>
  </div>
  <div class="legend">
    <div class="legend-item"><div class="legend-dot" style="background:#dcfce7;border:1px solid #166534;"></div>ノーマル (N/N)</div>
    <div class="legend-item"><div class="legend-dot" style="background:#fef3c7;border:1px solid #92400e;"></div>キャリア (P/N)</div>
    <div class="legend-item"><div class="legend-dot" style="background:#fee2e2;border:1px solid #991b1b;"></div>ポジティブ (P/P)</div>
    <div class="legend-item"><div class="legend-dot" style="background:#e0e7ff;border:1px solid #3730a3;"></div>形質 (Trait)</div>
  </div>

  <div class="tabs">
    <div class="tab active" onclick="showTab('overview')">全体サマリー</div>
{tab_buttons}
    {compare_tab_button}
  </div>

  <!-- Overview Tab -->
  <div id="overview" class="tab-content active">
    <div class="dog-card">
      <h2 class="section-title">検査対象一覧</h2>
      <table class="results-table">
        <tr><th>ペット名</th><th>登録名</th><th>犬種</th><th>性別</th><th>生年月日</th><th>ケース番号</th></tr>
{overview_table_rows}
      </table>
    </div>
    <div class="dog-card">
      <h2 class="section-title">要注意事項</h2>
{alerts_html}
    </div>
  </div>

  <!-- Individual Dog Tabs -->
{tab_contents}

  <!-- Compare Tab -->
  {compare_tab_html}
</div>
<script>
function showTab(id) {{
  document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(el => el.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  event.target.classList.add('active');
}}
</script>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"\nHTML レポート出力: {output_path}")


# ============================================================
# Excel出力
# ============================================================

def sanitize_for_excel(text: str) -> str:
    """Excelで使えない文字を除去"""
    if not text:
        return text
    import re as _re
    # Remove control characters except tab, newline, carriage return
    return _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)

def generate_excel(dogs: list, output_path: str):
    """Excelスプレッドシートを生成"""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="4A1A7A")
    normal_fill = PatternFill("solid", fgColor="DCFCE7")
    carrier_fill = PatternFill("solid", fgColor="FEF3C7")
    positive_fill = PatternFill("solid", fgColor="FEE2E2")
    trait_fill = PatternFill("solid", fgColor="E0E7FF")
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    def style_cell(ws, row, col, status=None):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if status == "normal":
            cell.fill = normal_fill
        elif status == "carrier":
            cell.fill = carrier_fill
        elif status == "positive":
            cell.fill = positive_fill
        elif status == "trait":
            cell.fill = trait_fill

    # ---- Sheet 1: Overview ----
    ws = wb.active
    ws.title = "概要"
    headers = ["ペット名", "登録名", "犬種", "性別", "生年月日", "マイクロチップ", "ケース番号", "検査日"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for r, dog in enumerate(dogs, 2):
        data = [dog.pet_name, dog.registered_name, dog.breed, dog.sex, dog.dob, dog.microchip, dog.case_number, dog.test_date]
        for c, val in enumerate(data, 1):
            ws.cell(row=r, column=c, value=val)
            style_cell(ws, r, c)

    for col_letter in ['A','B','C','D','E','F','G','H']:
        ws.column_dimensions[col_letter].width = 22

    # ---- Sheet per dog ----
    for dog in dogs:
        name = (dog.pet_name or dog.registered_name or "犬")[:30]
        ws = wb.create_sheet(title=name)

        # Basic info
        ws.cell(row=1, column=1, value="ペット名").font = Font(bold=True, name="Arial")
        ws.cell(row=1, column=2, value=dog.pet_name)
        ws.cell(row=2, column=1, value="登録名").font = Font(bold=True, name="Arial")
        ws.cell(row=2, column=2, value=dog.registered_name)
        ws.cell(row=3, column=1, value="犬種").font = Font(bold=True, name="Arial")
        ws.cell(row=3, column=2, value=dog.breed)
        ws.cell(row=4, column=1, value="性別").font = Font(bold=True, name="Arial")
        ws.cell(row=4, column=2, value=dog.sex)
        ws.cell(row=5, column=1, value="生年月日").font = Font(bold=True, name="Arial")
        ws.cell(row=5, column=2, value=dog.dob)
        ws.cell(row=6, column=1, value="マイクロチップ").font = Font(bold=True, name="Arial")
        ws.cell(row=6, column=2, value=dog.microchip)
        ws.cell(row=7, column=1, value="ケース番号").font = Font(bold=True, name="Arial")
        ws.cell(row=7, column=2, value=dog.case_number)

        # Health results
        row = 9
        ws.cell(row=row, column=1, value="【健康検査結果】").font = Font(bold=True, size=12, color="4A1A7A", name="Arial")
        row = 10
        for c, h in enumerate(["カテゴリー", "検査項目", "検査項目(英語)", "遺伝子型", "ステータス", "詳細"], 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, 6)

        for r_idx, result in enumerate(dog.health_results):
            r = row + 1 + r_idx
            ws.cell(row=r, column=1, value=sanitize_for_excel(result.category))
            ws.cell(row=r, column=2, value=sanitize_for_excel(result.japanese_name or result.test_name))
            ws.cell(row=r, column=3, value=sanitize_for_excel(result.test_name))
            ws.cell(row=r, column=4, value=sanitize_for_excel(result.genotype))
            status_jp = {"normal": "ノーマル", "carrier": "キャリア", "positive": "ポジティブ"}.get(result.status, result.status)
            ws.cell(row=r, column=5, value=status_jp)
            ws.cell(row=r, column=6, value=sanitize_for_excel(result.result_text[:150]))
            for c in range(1, 7):
                style_cell(ws, r, c, result.status)

        # Trait results
        row = row + len(dog.health_results) + 3
        ws.cell(row=row, column=1, value="【毛色・形質検査結果】").font = Font(bold=True, size=12, color="4A1A7A", name="Arial")
        row += 1
        for c, h in enumerate(["検査項目", "検査項目(英語)", "遺伝子型", "詳細"], 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, 4)

        for r_idx, result in enumerate(dog.trait_results):
            r = row + 1 + r_idx
            ws.cell(row=r, column=1, value=sanitize_for_excel(result.japanese_name or result.test_name))
            ws.cell(row=r, column=2, value=sanitize_for_excel(result.test_name))
            ws.cell(row=r, column=3, value=sanitize_for_excel(result.genotype))
            ws.cell(row=r, column=4, value=sanitize_for_excel(result.result_text[:150]))
            for c in range(1, 5):
                style_cell(ws, r, c, "trait")

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 60

    # ---- Comparison Sheet (if multiple dogs) ----
    if len(dogs) > 1:
        ws = wb.create_sheet(title="比較表")
        headers = ["検査項目"] + [d.pet_name or d.registered_name for d in dogs]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        style_header(ws, 1, len(headers))

        all_tests = {}
        for dog in dogs:
            for r in dog.health_results:
                key = r.test_name
                if key not in all_tests:
                    all_tests[key] = r.japanese_name or r.test_name

        row = 2
        for test_key, jp_name in all_tests.items():
            ws.cell(row=row, column=1, value=sanitize_for_excel(jp_name))
            style_cell(ws, row, 1)
            for d_idx, dog in enumerate(dogs):
                col = d_idx + 2
                found = False
                for r in dog.health_results:
                    if r.test_name == test_key:
                        ws.cell(row=row, column=col, value=r.genotype)
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                        style_cell(ws, row, col, r.status)
                        found = True
                        break
                if not found:
                    ws.cell(row=row, column=col, value="—")
                    style_cell(ws, row, col)
            row += 1

        ws.column_dimensions['A'].width = 40
        for i in range(len(dogs)):
            ws.column_dimensions[chr(66 + i)].width = 18

    wb.save(output_path)
    print(f"Excel レポート出力: {output_path}")


# ============================================================
# メインエントリポイント
# ============================================================

def collect_pdf_files(args: list) -> list:
    """引数からPDFファイルリストを収集"""
    pdf_files = []
    for arg in args:
        if os.path.isdir(arg):
            pdf_files.extend(glob.glob(os.path.join(arg, "*.pdf")))
            pdf_files.extend(glob.glob(os.path.join(arg, "*.PDF")))
        elif os.path.isfile(arg) and arg.lower().endswith('.pdf'):
            pdf_files.append(arg)
        elif '*' in arg or '?' in arg:
            pdf_files.extend(glob.glob(arg))
    return sorted(set(pdf_files))


def main():
    print("=" * 60)
    print("  Orivet 遺伝子検査PDF 自動解析ツール")
    print("=" * 60)

    if len(sys.argv) < 2:
        # If no args, look for PDFs in current directory
        pdf_files = glob.glob("*.pdf") + glob.glob("*.PDF")
        if not pdf_files:
            print("\n使い方:")
            print("  python orivet_analyzer.py *.pdf")
            print("  python orivet_analyzer.py folder_with_pdfs/")
            print("  python orivet_analyzer.py dog1.pdf dog2.pdf")
            sys.exit(1)
    else:
        pdf_files = collect_pdf_files(sys.argv[1:])

    if not pdf_files:
        print("\nエラー: PDFファイルが見つかりません。")
        sys.exit(1)

    print(f"\n{len(pdf_files)} 個のPDFファイルを検出しました。\n")

    # Parse all PDFs
    dogs = []
    for pdf_path in pdf_files:
        dog = parse_pdf(pdf_path)
        if dog:
            dogs.append(dog)

    if not dogs:
        print("\nエラー: 解析可能なOrivetレポートが見つかりませんでした。")
        sys.exit(1)

    print(f"\n{len(dogs)} 頭分のデータを解析しました。")

    # Determine output directory (use current working directory)
    output_dir = os.getcwd()
    # Check if --output or -o flag is provided
    for i, arg in enumerate(sys.argv):
        if arg in ('--output', '-o') and i + 1 < len(sys.argv):
            output_dir = sys.argv[i + 1]
            break
    os.makedirs(output_dir, exist_ok=True)
    html_path = os.path.join(output_dir, "orivet_report.html")
    xlsx_path = os.path.join(output_dir, "orivet_report.xlsx")

    # Generate outputs
    generate_html(dogs, html_path)
    generate_excel(dogs, xlsx_path)

    print(f"\n完了! 以下のファイルが生成されました:")
    print(f"  HTML: {html_path}")
    print(f"  Excel: {xlsx_path}")
    print()


if __name__ == "__main__":
    main()
