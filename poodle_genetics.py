#!/usr/bin/env python3
"""
Orivet 遺伝子解析ツール — 全犬種対応
=====================================================
Orivet遺伝子検査PDFの自動解析 + JKC血統書OCR + COI算出を
1本にまとめた統合ツール。全犬種のOrivetレポートに対応。

コマンド:
    # 遺伝子検査PDFだけ解析
    python poodle_genetics.py orivet *.pdf

    # 血統書写真だけ解析
    python poodle_genetics.py pedigree photo.jpg

    # 両方まとめて統合レポート
    python poodle_genetics.py all *.pdf --pedigree seven

    # 既知データでデモ
    python poodle_genetics.py demo

出力:
    - orivet_report.html  (統合HTMLレポート: 遺伝子 + 血統 + COI)
    - orivet_report.xlsx  (Excelスプレッドシート)

必要ライブラリ:
    pip install pdfplumber openpyxl
    # 血統書OCRを使う場合:
    pip install pytesseract Pillow
    # + Tesseract OCR本体
"""

import sys
import os
import re
import glob
import json
from dataclasses import dataclass, field
from typing import Optional
from datetime import datetime

# ============================================================
# ライブラリ読み込み
# ============================================================

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pytesseract
    from PIL import Image
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
except ImportError:
    pass


# ============================================================
# データ構造
# ============================================================

@dataclass
class TestResult:
    """個別の検査結果"""
    category: str
    test_name: str
    genotype: str
    result_text: str
    status: str          # normal / carrier / positive / trait
    japanese_name: str = ""

@dataclass
class DogProfile:
    """犬1頭分の遺伝子検査プロファイル"""
    pet_name: str = ""
    registered_name: str = ""
    breed: str = ""
    sex: str = ""
    dob: str = ""
    colour: str = ""
    microchip: str = ""
    case_number: str = ""
    owner_name: str = ""
    test_date: str = ""
    test_requested: str = ""
    approved_collection: str = ""
    sample_type: str = ""
    health_results: list = field(default_factory=list)
    trait_results: list = field(default_factory=list)
    source_file: str = ""
    # ゲノム多様性指標（Orivet 等の DNA 検査が報告するヘテロ接合率）。
    # 血統ベース COI とは別指標。PDF に記載があれば % 値（0-100）を格納。
    heterozygosity: Optional[float] = None
    # 犬種別の標準域（Typical range）。[low, high] の % 値。個体値の位置比較に使う。
    heterozygosity_range: Optional[list] = None

@dataclass
class Ancestor:
    """血統書上の1頭分"""
    position: str = ""
    name: str = ""
    registration: str = ""
    titles: str = ""
    color: str = ""
    dna_number: str = ""
    microchip: str = ""
    dob: str = ""

@dataclass
class Pedigree:
    """1頭分の完全な血統書データ"""
    dog_name: str = ""
    breed: str = ""
    registration: str = ""
    sex: str = ""
    dob: str = ""
    color: str = ""
    microchip: str = ""
    breeder: str = ""
    owner: str = ""
    sire: Optional[Ancestor] = None
    dam: Optional[Ancestor] = None
    ss: Optional[Ancestor] = None
    sd: Optional[Ancestor] = None
    ds: Optional[Ancestor] = None
    dd: Optional[Ancestor] = None
    sss: Optional[Ancestor] = None
    ssd: Optional[Ancestor] = None
    sds: Optional[Ancestor] = None
    sdd: Optional[Ancestor] = None
    dss: Optional[Ancestor] = None
    dsd: Optional[Ancestor] = None
    dds: Optional[Ancestor] = None
    ddd: Optional[Ancestor] = None
    source_file: str = ""

    def all_ancestors(self):
        """全祖先のリストを返す"""
        return [
            ("父 (Sire)", self.sire),
            ("母 (Dam)", self.dam),
            ("父方祖父 (G.Sire)", self.ss),
            ("父方祖母 (G.Dam)", self.sd),
            ("母方祖父 (G.Sire)", self.ds),
            ("母方祖母 (G.Dam)", self.dd),
            ("父方曾祖父1 (GG.Sire)", self.sss),
            ("父方曾祖母1 (GG.Dam)", self.ssd),
            ("父方曾祖父2 (GG.Sire)", self.sds),
            ("父方曾祖母2 (GG.Dam)", self.sdd),
            ("母方曾祖父1 (GG.Sire)", self.dss),
            ("母方曾祖母1 (GG.Dam)", self.dsd),
            ("母方曾祖父2 (GG.Sire)", self.dds),
            ("母方曾祖母2 (GG.Dam)", self.ddd),
        ]


# ============================================================
# 日本語マッピング
# ============================================================

CATEGORY_JP = {
    "musculoskeletal": "筋骨格系",
    "haemolymphatic": "血液・リンパ系",
    "nervous": "神経系",
    "neurologic": "神経系",
    "metabolic": "代謝系",
    "ophthalmologic": "眼科系",
    "cardiorespiratory": "心肺系",
    "cardiovascular": "心血管系",
    "dermatologic": "皮膚科系",
    "immunological": "免疫系",
    "respiratory": "呼吸器系",
    "urogenital": "泌尿生殖器系",
    "trait": "形質（毛色・外見）",
}

TEST_NAME_JP = {
    "chondrodystrophy with intervertebral disc disease": "軟骨異栄養症+椎間板疾患リスク (CDDY+IVDD)",
    "cddy with ivdd": "軟骨異栄養症+椎間板疾患リスク (CDDY+IVDD)",
    "osteochondrodysplasia": "骨軟骨異形成症",
    "chondrodysplasia": "軟骨異形成症 (CDPA)",
    "cdpa": "軟骨異形成症 (CDPA)",
    "congenital macrothrombocytopenia": "先天性巨大血小板減少症",
    "congenital methemoglobinemia": "先天性メトヘモグロビン血症",
    "von willebrand": "フォンウィルブランド病 I型",
    "degenerative myelopathy": "変性性脊髄症 (DM)",
    "gangliosidosis gm2": "ガングリオシドーシス GM2",
    "progressive rod cone degeneration": "進行性網膜萎縮症 (prcd-PRA)",
    "prcd": "進行性網膜萎縮症 (prcd-PRA)",
    "progressive retinal atrophy": "進行性網膜萎縮症 (PRA)",
    "exercise-induced collapse": "運動誘発性虚脱 (EIC)",
    "neonatal encephalopathy": "新生児脳症",
    "hyperuricosuria": "高尿酸尿症",
    "cystinuria": "シスチン尿症",
    "a locus": "A座位 (アグーチ)",
    "b locus": "B座位 (ブラウン)",
    "d locus": "D座位 (ダイリュート)",
    "dilute": "D座位 (ダイリュート)",
    "e locus": "E座位 (エクステンション)",
    "em locus": "EM座位 (メラニスティックマスク)",
    "mc1r": "EM座位 (メラニスティックマスク)",
    "k locus": "K座位 (ドミナントブラック)",
    "m locus": "M座位 (マール/ダップル)",
    "merle": "M座位 (マール/ダップル)",
    "curly coat": "巻き毛 (Curly Coat)",
    "furnishings": "ファーニシング (RSPO2)",
    "rspo2": "ファーニシング (RSPO2)",
    "pied": "パイド (Pied)",
    "brown tyrp1": "ブラウン TYRP1",
    "improper coat": "インプロパーコート",
    "coat length": "毛の長さ",
    "shedding": "換毛",
}


# ============================================================
# 毛色・形質 遺伝子型の日本語注釈
# ============================================================

def get_trait_annotation(test_name: str, genotype: str) -> str:
    """検査項目と遺伝子型から、分かりやすい日本語の注釈を返す"""
    if not genotype or genotype == "—":
        return ""

    name_lower = test_name.lower()

    # E Locus (エクステンション / クリーム)
    if "e locus" in name_lower:
        return {
            "E/E": "クリーム因子なし。ブラック/ブラウン等の濃い毛色",
            "E/e": "クリーム因子を1つ保有（キャリア）。見た目は濃い色だが、子犬にクリーム/ホワイトが出る可能性あり",
            "e/e": "クリーム因子を2つ保有。クリーム/ホワイト/アプリコット/レッドの毛色になる",
        }.get(genotype, "")

    # K Locus (ドミナントブラック)
    if "k locus" in name_lower:
        mapping = {
            "KB/KB": "ドミナントブラックを2つ保有。全身が単色（ブラックまたはブラウン）",
            "K/K": "ドミナントブラックを2つ保有。全身が単色（ブラックまたはブラウン）",
            "KB/ky": "ドミナントブラック1つ保有。見た目は単色だが、ファントム/タンポイントの子が出る可能性あり",
            "KB/kbr": "ドミナントブラック1つ + ブリンドル1つ保有。見た目は単色",
            "ky/ky": "ドミナントブラックなし。アグーチ座位の模様（ファントム/セーブル等）が発現する",
            "kbr/ky": "ブリンドル因子あり。ブリンドル模様が出る可能性あり",
            "kbr/kbr": "ブリンドル因子を2つ保有。ブリンドル模様が発現する",
        }
        return mapping.get(genotype, "")

    # A Locus (アグーチ)
    if "a locus" in name_lower:
        return {
            "ay/ay": "セーブル（毛先が黒い明るい毛色）。ky/kyの場合に発現する",
            "ay/at": "セーブルだが、ファントム/タンポイントの子が出る可能性あり",
            "ay/aw": "セーブルとワイルドタイプのアグーチ。ky/kyの場合に発現する",
            "at/at": "ファントム/タンポイント（目の上・足先等に明るい斑点）。ky/kyの場合に発現する",
            "aw/at": "ワイルドタイプアグーチだが、ファントム/タンポイントの子が出る可能性あり",
            "aw/aw": "ワイルドタイプアグーチ。ウルフセーブル等の毛色パターン",
            "a/a": "リセッシブブラック。アグーチ模様なし",
        }.get(genotype, "A座位（アグーチ）: 毛色の模様パターンを決定する遺伝子座。K座位がky/kyの場合に模様が発現する")

    # B Locus (ブラウン)
    if "b locus" in name_lower:
        return {
            "BB": "ブラウン因子なし。鼻・肉球はブラック",
            "Bb": "ブラウン因子を1つ保有（キャリア）。見た目はブラックだが、ブラウンの子が出る可能性あり",
            "bb": "ブラウン因子を2つ保有。ブラウン（チョコレート）の毛色。鼻・肉球もブラウン",
        }.get(genotype, "B座位（ブラウン）: ブラウン/チョコレート/レバーの毛色を決定する遺伝子座")

    # D Locus (ダイリュート / 希釈)
    if "d locus" in name_lower or "dilute" in name_lower:
        return {
            "D/D": "希釈因子なし。色素は通常通り発現する",
            "D/d": "希釈因子を1つ保有（キャリア）。見た目は通常色だが、希釈色の子が出る可能性あり",
            "d/d": "希釈因子を2つ保有。ブラック→ブルー、ブラウン→カフェオレに希釈される",
        }.get(genotype, "")

    # M Locus (マール)
    if "m locus" in name_lower or "merle" in name_lower:
        return {
            "m/m": "マール因子なし。単色（ソリッドカラー）",
            "M/m": "マール因子を1つ保有。マール模様（まだら）が出る可能性あり",
            "M/M": "マール因子を2つ保有（ダブルマール）。健康リスク（聴覚・視覚障害）あり",
        }.get(genotype, "")

    # S Locus (パイド / 白斑)
    if "pied" in name_lower:
        return {
            "S/S": "パイド因子なし。白斑模様は出ない",
            "S/sp": "パイド因子を1つ保有（キャリア）。わずかな白斑が出ることがある",
            "sp/sp": "パイド因子を2つ保有。パーティカラー（大きな白斑模様）になる",
        }.get(genotype, "")

    # EM Locus (メラニスティックマスク)
    if "em" in name_lower or "mc1r" in name_lower or "melanistic mask" in name_lower:
        return {
            "En/En": "マスク因子なし。顔にメラニスティックマスクは出ない",
            "EM/EM": "マスク因子を2つ保有。顔周りに黒いマスク模様が出る",
            "EM/En": "マスク因子を1つ保有。マスク模様が出る可能性あり",
        }.get(genotype, "")

    # Furnishings (RSPO2)
    if "furnishings" in name_lower or "rspo2" in name_lower:
        return {
            "F/F": "ファーニシング因子を2つ保有。眉毛・ヒゲ・足の飾り毛あり（プードル・シュナウザー等の特徴的な外見）",
            "F/f": "ファーニシング因子を1つ保有。ファーニシングあり。抜け毛のある子が出る可能性",
            "f/f": "ファーニシング因子なし。ファーニシングなし（スムースフェイス）。抜け毛が多い傾向",
        }.get(genotype, "")

    # Curly Coat
    if "curly coat" in name_lower:
        return {
            "Cu/Cu": "巻き毛因子を2つ保有。しっかりとしたカーリーコート",
            "Cu/N": "巻き毛因子を1つ保有。ウェーブがかった被毛",
            "N/N": "巻き毛因子なし。ストレートな被毛",
        }.get(genotype, "")

    # Brown TYRP1
    if "brown tyrp1" in name_lower:
        return {
            "BL/BL": "TYRP1ブラウン因子なし。ブラウン/レバーにならない",
            "BL/bs": "TYRP1ブラウン因子を1つ保有（キャリア）",
            "bs/bs": "TYRP1ブラウン因子を2つ保有。ブラウン/レバーの毛色に影響する可能性",
        }.get(genotype, "")

    # CDPA (形質として出ることもある)
    if "chondrodysplasia" in name_lower or "cdpa" in name_lower:
        return {
            "N/N": "正常。短足因子(CDPA)なし",
            "P/N": "キャリア。短足因子を1つ保有",
            "P/P": "短足因子を2つ保有。脚が短くなる可能性",
        }.get(genotype, "")

    return ""


# ============================================================
# 健康検査 遺伝子型の日本語注釈
# ============================================================

def get_health_annotation(test_name: str, genotype: str, status: str) -> str:
    """健康検査項目と遺伝子型から、分かりやすい日本語の注釈を返す"""
    if not genotype:
        return ""

    name_lower = test_name.lower()

    # CDDY+IVDD (軟骨異栄養症+椎間板疾患)
    if "chondrodystrophy" in name_lower or "cddy" in name_lower or "ivdd" in name_lower:
        return {
            "N/N": "正常。椎間板疾患(IVDD)のリスクは一般レベル",
            "P/N": "キャリア（保因犬）。CDDY因子を1つ保有。椎間板ヘルニアのリスクがやや高い。繁殖相手の選定に注意が必要",
            "P/P": "発症リスクあり。CDDY因子を2つ保有。椎間板ヘルニアのリスクが高い。体重管理や激しい運動の制限を推奨",
        }.get(genotype, "")

    # 骨軟骨異形成症
    if "osteochondrodysplasia" in name_lower:
        return {
            "N/N": "正常。骨軟骨異形成症の変異なし",
            "P/N": "キャリア（保因犬）。繁殖時にキャリア同士の交配を避けること",
            "P/P": "発症リスクあり。骨・軟骨の異常な発達が起こる可能性。獣医師への相談を推奨",
        }.get(genotype, "")

    # 先天性巨大血小板減少症
    if "macrothrombocytopenia" in name_lower:
        return {
            "N/N": "正常。血小板に関する遺伝的異常なし",
            "P/N": "キャリア（保因犬）。血小板が大きく数が少ない場合がある。臨床的問題は通常なし",
            "P/P": "発症リスクあり。血小板が著しく大きく数が減少。出血傾向に注意",
        }.get(genotype, "")

    # 先天性メトヘモグロビン血症
    if "methemoglobinemia" in name_lower:
        return {
            "N/N": "正常。メトヘモグロビン血症の変異なし",
            "P/N": "キャリア（保因犬）。繁殖時にキャリア同士の交配を避けること",
            "P/P": "発症リスクあり。チアノーゼ（皮膚や粘膜の青紫変色）が見られる可能性。獣医師への相談を推奨",
        }.get(genotype, "")

    # フォンウィルブランド病 I型
    if "von willebrand" in name_lower:
        return {
            "N/N": "正常。出血性疾患のリスクなし",
            "P/N": "キャリア（保因犬）。通常は臨床症状なし。繁殖時にキャリア同士の交配を避けること",
            "P/P": "発症リスクあり。出血が止まりにくい傾向。手術前に獣医師に申告すること",
        }.get(genotype, "")

    # 変性性脊髄症 (DM)
    if "degenerative myelopathy" in name_lower:
        return {
            "N/N": "正常。変性性脊髄症のリスクなし",
            "P/N": "キャリア（保因犬）。通常は発症しない。繁殖時にキャリア同士の交配を避けること",
            "P/P": "発症リスクあり。高齢期に後肢の麻痺が進行する可能性。定期的な神経学的検査を推奨",
        }.get(genotype, "")

    # ガングリオシドーシス GM2
    if "gangliosidosis" in name_lower:
        return {
            "N/N": "正常。ガングリオシドーシスの変異なし",
            "P/N": "キャリア（保因犬）。繁殖時にキャリア同士の交配を避けること",
            "P/P": "発症リスクあり。神経系の進行性疾患。早期の獣医師への相談を推奨",
        }.get(genotype, "")

    # 進行性網膜萎縮症 (prcd-PRA)
    if "progressive rod cone" in name_lower or "prcd" in name_lower or "progressive retinal atrophy" in name_lower:
        return {
            "N/N": "正常。prcd-PRAによる失明リスクなし",
            "P/N": "キャリア（保因犬）。視力への影響なし。繁殖時にキャリア同士の交配を避けること",
            "P/P": "発症リスクあり。進行性の視力低下・失明の可能性。定期的な眼科検査を推奨",
        }.get(genotype, "")

    # 運動誘発性虚脱 (EIC)
    if "exercise-induced collapse" in name_lower or "eic" in name_lower:
        return {
            "N/N": "正常。運動誘発性虚脱のリスクなし",
            "P/N": "キャリア（保因犬）。通常は発症しない。繁殖時にキャリア同士の交配を避けること",
            "P/P": "発症リスクあり。激しい運動後に一過性の虚脱が起こる可能性。運動の強度に注意",
        }.get(genotype, "")

    # 新生児脳症
    if "neonatal encephalopathy" in name_lower:
        return {
            "N/N": "正常。新生児脳症の変異なし",
            "P/N": "キャリア（保因犬）。繁殖時にキャリア同士の交配を避けること",
            "P/P": "発症リスクあり。新生児期に神経症状が出る可能性",
        }.get(genotype, "")

    # 汎用的な注釈（上記に該当しない場合）
    if status == "normal":
        return "正常。この疾患に関連する遺伝子変異は検出されませんでした"
    elif status == "carrier":
        return "キャリア（保因犬）。変異を1つ保有。通常は発症しないが、繁殖相手の選定に注意が必要"
    elif status == "positive":
        return "発症リスクあり。変異を2つ保有。獣医師への相談を推奨"

    return ""


# ============================================================
# PDF解析時の不正データフィルター
# ============================================================

# PDFからの誤抽出を除外するための除外キーワード
HEALTH_TEST_BLACKLIST = [
    "glossary of genetic terms",
    "pass on any disease-causing mutation",
    "unknown then it may produce",
    "affected offspring",
    "genetic terms",
    "copyright",
    "disclaimer",
    "page ",
]


# ============================================================
# 詳細解説ナレッジベース（一般飼い主向け）
# ============================================================
# 「理解できること」を目的に、各疾患・形質について
#   - summary: 1〜2文の概要（何の病気/形質か）
#   - mechanism: 発症メカニズム・特徴
#   - symptoms: 症状（疾患の場合）
#   - inheritance: 遺伝様式
#   - advice: 飼育・繁殖時のアドバイス
#   - references: 参考リンク [{label, url}]
# を提供する。
#
# 参考リンクは URL が常に有効である **Wikipedia 検索URL** と
# **Google 検索URL** を採用。直接記事URLは時間経過で404になる
# リスクがあるため、検索URLでフェイルセーフ。
# 追加リンクが必要な場合は references リストに新エントリを足すだけ。

def _wiki_jp(term: str) -> str:
    """Wikipedia 日本語版の検索URL（常に有効）"""
    from urllib.parse import quote
    return "https://ja.wikipedia.org/wiki/Special:Search/" + quote(term)


def _google_search(query: str) -> str:
    from urllib.parse import quote
    return "https://www.google.com/search?q=" + quote(query)


DISEASE_KB = [
    {
        "match": ["chondrodystrophy", "cddy", "ivdd"],
        "title": "軟骨異栄養症 + 椎間板疾患 (CDDY+IVDD)",
        "severity": "high",
        "summary": "椎間板を構成する軟骨の異常により、椎間板ヘルニアを起こしやすくなる遺伝性疾患です。",
        "mechanism": "FGF4遺伝子のレトロウィルス挿入が原因。椎間板の中心部（髄核）が早期に変性・石灰化し、些細な衝撃で破裂し脊髄を圧迫します。",
        "symptoms": "後肢のふらつき、痛み、歩行困難。重症例では完全麻痺や排尿障害が起こり得ます。",
        "inheritance": "常染色体（不完全）優性。1コピーでもリスクが上がり、2コピーでさらに高まります。",
        "advice": "P/P 同士の交配は避けることを強く推奨。発症リスクのある個体は体重管理と階段の昇降・激しい運動を控えることが重要。",
        "references": [
            {"label": "Wikipedia: 椎間板ヘルニア", "url": _wiki_jp("椎間板ヘルニア")},
            {"label": "詳細を検索 (CDDY)", "url": _google_search("CDDY IVDD 犬 軟骨異栄養症")},
        ],
    },
    {
        "match": ["osteochondrodysplasia"],
        "title": "骨軟骨異形成症 (Osteochondrodysplasia / SLC13A1)",
        "severity": "high",
        "summary": "骨と軟骨の発達異常により、四肢の短縮や関節異常が現れる遺伝性疾患です。",
        "mechanism": "SLC13A1 遺伝子の変異によりミネラル代謝が異常になり、骨格の正常な発達が阻害されます。",
        "symptoms": "四肢の短縮、関節の変形、運動制限。スコティッシュフォールド・ミニチュアプードルなどで報告。",
        "inheritance": "常染色体劣性。両親共にキャリア(P/N)の場合、25% の確率で発症犬(P/P)が生まれます。",
        "advice": "P/N 同士の交配は避けてください。発症犬は獣医師による定期的な整形外科診察を推奨。",
        "references": [
            {"label": "Wikipedia: 骨軟骨異形成症", "url": _wiki_jp("骨軟骨異形成症")},
            {"label": "詳細を検索", "url": _google_search("骨軟骨異形成症 犬 SLC13A1")},
        ],
    },
    {
        "match": ["chondrodysplasia (cdpa)", "cdpa", "chondrodysplasia"],
        "title": "軟骨異形成症 (CDPA / 短足遺伝子)",
        "severity": "low",
        "summary": "短い四肢を生む遺伝子で、ダックスフンド・コーギーなどの「短足犬種」の特徴となる因子です。",
        "mechanism": "FGF4 遺伝子の重複により軟骨形成が変化し、四肢が短くなります。CDDY と異なり病気そのものではなく、犬種特性として定着しています。",
        "symptoms": "通常、症状はなし。短足は犬種スタンダードとして許容されています。",
        "inheritance": "不完全優性。短足犬種では多くが P/P または P/N。",
        "advice": "短足犬種ではこの因子の保有が一般的。CDDY とは別の因子ですが、合わせて検査する場合が多いです。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("CDPA 軟骨異形成症 犬 FGF4")},
        ],
    },
    {
        "match": ["macrothrombocytopenia"],
        "title": "先天性巨大血小板減少症 (Macrothrombocytopenia / β1-tubulin)",
        "severity": "low",
        "summary": "血小板が通常より大きく、数が少ない遺伝性疾患です。多くは無症状ですが手術時に注意が必要。",
        "mechanism": "TUBB1 遺伝子（β1-tubulin）の変異により血小板の形成が異常になります。キャバリア・キング・チャールズ・スパニエルで頻発。",
        "symptoms": "ほとんどの場合無症状。健康診断で血小板数が低く出るが、機能は保たれていることが多い。",
        "inheritance": "常染色体劣性。両親キャリアの場合 25% で発症犬。",
        "advice": "手術前には事前に獣医師に申告してください。誤って「血小板減少症」と診断されないよう注意。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Macrothrombocytopenia 犬 TUBB1")},
        ],
    },
    {
        "match": ["methemoglobinemia"],
        "title": "先天性メトヘモグロビン血症 (Methemoglobinemia / CYB5R3)",
        "severity": "medium",
        "summary": "血液中のヘモグロビンが酸素を運べない型に変わってしまう遺伝性疾患です。",
        "mechanism": "CYB5R3 遺伝子の変異により、酸化型ヘモグロビン（メトヘモグロビン）を還元する酵素が欠損。慢性的なチアノーゼ症状を起こします。",
        "symptoms": "皮膚・粘膜の青紫変色（チアノーゼ）、運動不耐性、疲れやすい。",
        "inheritance": "常染色体劣性。両親キャリアの場合 25% で発症犬。",
        "advice": "発症犬は獣医師の管理下で定期検査が必要。麻酔時は特に注意。",
        "references": [
            {"label": "Wikipedia: メトヘモグロビン血症", "url": _wiki_jp("メトヘモグロビン血症")},
            {"label": "詳細を検索", "url": _google_search("Methemoglobinemia 犬 CYB5R3")},
        ],
    },
    {
        "match": [
            "willebrand type 1", "willebrand's type 1",
            "willebrand disease type 1", "willebrand's disease type 1",
            "vwd1", "vwd-1", "vwd type 1", "vwd type i",
            "willebrand i", "willebrand 1",
        ],
        "title": "フォン・ヴィレブランド病 I型 (vWD1)",
        "severity": "medium",
        "summary": "止血に関わるフォン・ヴィレブランド因子が不足し、出血が止まりにくくなる遺伝性疾患です。",
        "mechanism": "vWF 遺伝子の変異により、血小板を血管壁に結合させる蛋白質が不足。軽度（I型）から重度（II型・III型）まで様々。ドーベルマンで頻発。",
        "symptoms": "外傷時の止血困難、鼻血、歯科処置後の長時間出血、血便など。",
        "inheritance": "常染色体（不完全）優性。1コピーで軽度症状、2コピーでより重度。",
        "advice": "手術・抜歯前に必ず獣医師に申告してください。アスピリン等の抗血小板薬は避ける必要があります。",
        "references": [
            {"label": "Wikipedia: フォン・ヴィレブランド病", "url": _wiki_jp("フォン・ヴィレブランド病")},
            {"label": "詳細を検索", "url": _google_search("vWD von Willebrand 犬")},
        ],
    },
    {
        "match": ["degenerative myelopathy", "\\bdm\\b"],
        "title": "変性性脊髄症 (DM / SOD1)",
        "summary": "高齢期に脊髄が徐々に変性し、後肢の麻痺が進行する遺伝性神経疾患です。",
        "mechanism": "SOD1 遺伝子の変異により神経細胞内に異常蛋白質が蓄積。ヒトのALS（筋萎縮性側索硬化症）と類似のメカニズムで発症します。",
        "symptoms": "8〜14歳頃から後肢のふらつき → 麻痺へ進行。痛みはない場合が多い。最終的には前肢にも進行。",
        "inheritance": "常染色体劣性（不完全浸透）。P/P でも全頭発症するわけではなく、発症率は犬種により異なる。",
        "advice": "P/P 同士の交配は避けることを推奨。発症犬はリハビリ・補助器具で QOL を維持できます。",
        "severity": "high",
        "references": [
            {"label": "Wikipedia: 変性性脊髄症", "url": _wiki_jp("変性性脊髄症")},
            {"label": "詳細を検索", "url": _google_search("Degenerative Myelopathy 犬 SOD1")},
        ],
    },
    {
        "match": ["gm2 gangliosidosis", "gm2", "gangliosidosis gm2"],
        "title": "ガングリオシドーシス GM2 (GM2 / HEXB)",
        "severity": "high",
        "summary": "神経細胞内に脂質が異常蓄積し、進行性の神経障害を起こす重篤な遺伝性疾患です。",
        "mechanism": "HEXB 遺伝子の変異によりリソソーム酵素ヘキソサミニダーゼBが欠損。GM2ガングリオシドが分解されず神経細胞に蓄積。ヒトのテイ・サックス病類似。",
        "symptoms": "若齢で発症し、運動失調・痙攣・視覚障害が進行。多くは1〜2年で死に至る。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/N 同士の交配は厳に避けるべき。発症犬には根本治療はなく、対症療法のみ。",
        "references": [
            {"label": "Wikipedia: テイ・サックス病", "url": _wiki_jp("テイ・サックス病")},
            {"label": "詳細を検索", "url": _google_search("GM2 Gangliosidosis 犬 HEXB")},
        ],
    },
    {
        "match": ["progressive rod cone", "prcd", "progressive retinal atrophy"],
        "title": "進行性網膜萎縮症 (prcd-PRA / PRCD)",
        "severity": "high",
        "summary": "網膜の光受容細胞（杆体・錐体）が徐々に変性し、最終的に失明する遺伝性疾患です。",
        "mechanism": "PRCD 遺伝子の変異により網膜細胞が徐々に死滅。最初に夜盲、次第に昼間視力も失われます。",
        "symptoms": "薄暗い場所での視覚障害（夜盲）→ 周辺視野の喪失 → 完全失明。痛みはなし。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発症は通常 3〜5歳以降のため、繁殖判断には遺伝子検査が重要。失明後も嗅覚・聴覚で適応可能。",
        "references": [
            {"label": "Wikipedia: 網膜色素変性症", "url": _wiki_jp("網膜色素変性症")},
            {"label": "詳細を検索", "url": _google_search("prcd-PRA 犬 進行性網膜萎縮")},
        ],
    },
    {
        "match": ["exercise-induced collapse", "\\beic\\b"],
        "title": "運動誘発性虚脱 (EIC / DNM1)",
        "severity": "medium",
        "summary": "激しい運動の後に突然脱力・崩れる遺伝性疾患です。ラブラドール等で報告。",
        "mechanism": "DNM1 遺伝子の変異により、運動中の神経シナプス伝達が一時的に障害されます。",
        "symptoms": "5〜15分の激しい運動後、後肢の脱力・歩行不能。意識はあり、5〜25分で回復することが多い。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発症犬は激しい運動を避け、ドッグスポーツへの参加は獣医師と相談を。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("EIC 運動誘発性虚脱 犬 DNM1")},
        ],
    },
    # === 神経系・脳系 ===
    {
        "match": ["neonatal encephalopathy", "news", "atf2"],
        "title": "新生児脳症 (NEwS / ATF2)",
        "severity": "high",
        "summary": "プードルに特有の重篤な新生児神経疾患。生後数週間以内に発症し致死的です。",
        "mechanism": "ATF2 遺伝子の変異により神経発達が障害されます。スタンダードプードルで報告。",
        "symptoms": "生後4〜6週から運動失調・痙攣・成長不全。多くは離乳期までに死亡。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "プードル繁殖では P/N 同士の交配を厳に避けること。スタンダードプードルでのキャリア率は数%。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Neonatal Encephalopathy Seizures プードル ATF2")},
        ],
    },
    {
        "match": ["neuronal ceroid lipofuscinosis", "\\bncl\\b"],
        "title": "神経セロイドリポフスチン症 (NCL)",
        "severity": "high",
        "summary": "脳細胞に異常物質（セロイドリポフスチン）が蓄積し、進行性の神経変性を起こす疾患群です。",
        "mechanism": "複数の遺伝子（CLN5, CLN6, CLN8, CTSD など）の変異によりリソソーム機能が障害。多数の亜型が存在し、犬種ごとに原因遺伝子が異なります。",
        "symptoms": "若齢発症の場合: 1〜3歳で行動異常・運動失調・視覚障害→死亡。成犬発症型もあり。",
        "inheritance": "ほとんどが常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/P は予後不良で根本治療なし。繁殖前検査必須。",
        "references": [
            {"label": "Wikipedia: 神経セロイドリポフスチン症", "url": _wiki_jp("神経セロイドリポフスチン症")},
            {"label": "詳細を検索", "url": _google_search("NCL Neuronal Ceroid Lipofuscinosis 犬")},
        ],
    },
    {
        "match": ["late onset ataxia", "\\bloa\\b", "late-onset ataxia"],
        "title": "若年性遅発型運動失調 (LOA / CAPN1)",
        "severity": "high",
        "summary": "若い時期（1〜2歳）から始まる小脳性運動失調。歩行・バランス障害が徐々に進行します。",
        "mechanism": "CAPN1 遺伝子の変異により小脳プルキンエ細胞が変性。ジャック・ラッセル・テリア等で報告。",
        "symptoms": "ふらつき歩行、頭の震え、姿勢制御の困難。痛みはない。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "対症療法のみ。重度になると QOL 維持が困難。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Late Onset Ataxia LOA 犬 CAPN1")},
        ],
    },
    {
        "match": ["spinocerebellar ataxia", "\\bsca\\b"],
        "title": "脊髄小脳変性症 (SCA)",
        "severity": "high",
        "summary": "脊髄と小脳の進行性変性により、運動・バランス障害を起こす遺伝性疾患です。",
        "mechanism": "複数の原因遺伝子（KCNJ10 など）の変異により神経細胞の機能が低下。",
        "symptoms": "歩行のふらつき、捻挫様の転倒、進行性の運動失調。",
        "inheritance": "常染色体劣性。",
        "advice": "発症犬は対症療法のみ。繁殖前検査が予防の鍵。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Spinocerebellar Ataxia 犬")},
        ],
    },
    {
        "match": ["multidrug resistance", "\\bmdr1\\b", "abcb1"],
        "title": "多剤耐性遺伝子 (MDR1 / ABCB1)",
        "severity": "high",
        "summary": "特定の薬剤（イベルメクチン等の駆虫薬・抗がん剤・止瀉薬）に重篤な副作用を起こす遺伝子変異です。",
        "mechanism": "ABCB1（旧名 MDR1）遺伝子の変異により、脳血液関門で薬剤を排出する蛋白質が機能不全。薬剤が脳に蓄積し神経毒性を起こします。コリー系犬種で頻発。",
        "symptoms": "対象薬剤投与後に運動失調・痙攣・昏睡・呼吸停止など重篤な神経症状。",
        "inheritance": "常染色体（不完全）優性。1コピーでもリスクあり、2コピーで重篤。",
        "advice": "**全ての治療・投薬時に獣医師へ MDR1 状態を申告すること**。特にイベルメクチン・ロペラミド・ビンクリスチン等は要注意。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("MDR1 ABCB1 犬 薬剤 イベルメクチン")},
        ],
    },
    # === 眼科系 ===
    {
        "match": ["cone-rod dystrophy", "crd4", "cone rod dystrophy"],
        "title": "錐体杆体ジストロフィー (crd4 / RPGRIP1)",
        "severity": "high",
        "summary": "網膜の錐体（昼間視）が先に変性し、その後杆体（夜間視）も影響を受けるPRAの亜型です。",
        "mechanism": "RPGRIP1 遺伝子の変異により網膜光受容細胞が機能不全。プードルやミニチュアロングヘアードダックスフンドなどで報告。",
        "symptoms": "若齢期から昼間の視覚異常 → 夜盲 → 完全失明。痛みはなし。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "失明後の生活適応は可能。繁殖前検査が重要。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Cone-Rod Dystrophy crd4 犬 RPGRIP1")},
        ],
    },
    {
        "match": ["achromatopsia", "day blindness", "cnga3"],
        "title": "全色盲 / 昼盲 (Achromatopsia / CNGA3)",
        "severity": "medium",
        "summary": "錐体細胞の機能不全により、明るい場所で物が見えなくなる疾患です。",
        "mechanism": "CNGA3 遺伝子の変異により錐体光受容細胞が機能しません。色覚は失われ、明るい光下では視覚が低下。",
        "symptoms": "明るい場所で目を細める・物にぶつかる。薄暗い場所では比較的見える（杆体機能は保たれる）。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "失明と異なり完全な盲目ではないため、明るさを抑えた環境で QOL を維持できる。",
        "references": [
            {"label": "Wikipedia: 全色盲", "url": _wiki_jp("全色盲")},
            {"label": "詳細を検索", "url": _google_search("Achromatopsia 犬 CNGA3 day blindness")},
        ],
    },
    {
        "match": ["collie eye anomaly", "\\bcea\\b"],
        "title": "コリーアイ症候群 (CEA / NHEJ1)",
        "severity": "medium",
        "summary": "網膜・脈絡膜・強膜の形成異常により、視覚障害を起こす遺伝性眼疾患です。コリー系で多発。",
        "mechanism": "NHEJ1 遺伝子の変異により眼球の発達が異常になります。軽度から重度まで連続スペクトラム。",
        "symptoms": "軽度: 無症状〜軽度視覚障害。重度: 出血・網膜剥離・失明。",
        "inheritance": "常染色体劣性（浸透率が変動）。両親キャリアで 25% 発症。",
        "advice": "コリー・シェルティでは検査必須。発症の重症度は個体差が大きい。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Collie Eye Anomaly CEA 犬 NHEJ1")},
        ],
    },
    {
        "match": ["hereditary cataract", "hsf4"],
        "title": "遺伝性白内障 (HSF4)",
        "severity": "medium",
        "summary": "若齢期から水晶体が混濁し、視力低下〜失明に至る遺伝性疾患です。",
        "mechanism": "HSF4 遺伝子の変異により水晶体の蛋白質形成が異常になります。ボストンテリア・スタッフィー等で報告。",
        "symptoms": "若齢期（数ヶ月〜数年）から白内障進行。視覚低下〜失明。",
        "inheritance": "犬種により異なる（劣性〜優性）。",
        "advice": "早期発見で手術により視力回復可能な場合あり。眼科専門医への相談を。",
        "references": [
            {"label": "Wikipedia: 白内障", "url": _wiki_jp("白内障")},
            {"label": "詳細を検索", "url": _google_search("Hereditary Cataract HSF4 犬")},
        ],
    },
    # === 代謝・血液系 ===
    {
        "match": ["hyperuricosuria", "\\bhuu\\b", "slc2a9"],
        "title": "高尿酸尿症 (HUU / SLC2A9)",
        "severity": "medium",
        "summary": "尿酸が異常に高くなり、尿路結石（尿酸石）を生じやすくなる遺伝性疾患です。",
        "mechanism": "SLC2A9 遺伝子の変異により肝臓での尿酸代謝が異常になります。ダルメシアン・ブルドッグなどで高頻度。",
        "symptoms": "頻尿・血尿・排尿困難・腎結石。重度では尿閉・腎機能不全。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発症犬は低プリン体食・水分摂取増・定期的な尿検査が必要。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Hyperuricosuria HUU 犬 ダルメシアン")},
        ],
    },
    {
        "match": ["pyruvate kinase", "\\bpk\\b deficiency", "pklr"],
        "title": "ピルビン酸キナーゼ欠損症 (PK / PKLR)",
        "severity": "high",
        "summary": "赤血球の代謝障害により慢性溶血性貧血を起こす遺伝性疾患です。",
        "mechanism": "PKLR 遺伝子の変異により赤血球内のエネルギー産生が低下。赤血球が早期に破壊されます。",
        "symptoms": "慢性貧血・倦怠・運動不耐・脾腫。多くは2〜5歳で重症化。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "骨髄移植以外の根治治療なし。輸血や対症療法で QOL 維持。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Pyruvate Kinase Deficiency 犬 PKLR")},
        ],
    },
    {
        "match": ["factor vii", "factor 7"],
        "title": "第VII因子欠損症 (Factor VII / F7)",
        "severity": "medium",
        "summary": "血液凝固因子の一つが欠損し、出血傾向を示す軽度〜中等度の凝固障害です。",
        "mechanism": "F7 遺伝子の変異により凝固第VII因子が低下。多くは軽症だが手術時に問題となる。",
        "symptoms": "通常は無症状。外傷・手術時に止血困難。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "手術前には必ず凝固検査と申告を。出血時は新鮮凍結血漿で対応。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Factor VII deficiency 犬 F7")},
        ],
    },
    {
        "match": ["mucopolysaccharidosis", "mps"],
        "title": "ムコ多糖症 (MPS)",
        "severity": "high",
        "summary": "ムコ多糖類が分解されず体内に蓄積し、骨格異常・臓器障害を起こす重篤な代謝疾患です。",
        "mechanism": "リソソーム酵素の遺伝的欠損により多糖類が蓄積。MPS I/VI/VII など複数の亜型。",
        "symptoms": "顔面変形・関節異常・成長障害・心障害・視覚障害など。生命予後不良。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/P は予後不良。繁殖前検査必須。",
        "references": [
            {"label": "Wikipedia: ムコ多糖症", "url": _wiki_jp("ムコ多糖症")},
            {"label": "詳細を検索", "url": _google_search("Mucopolysaccharidosis MPS 犬")},
        ],
    },
    # === 筋・骨格系 ===
    {
        "match": ["centronuclear myopathy", "\\bcnm\\b", "ptpla"],
        "title": "中心核ミオパチー (CNM / PTPLA)",
        "severity": "high",
        "summary": "若齢期からの筋力低下を起こす遺伝性筋疾患です。ラブラドール特発性筋疾患とも。",
        "mechanism": "PTPLA 遺伝子の変異により筋細胞の構造が異常になります。",
        "symptoms": "生後数ヶ月から運動不耐・筋力低下・歩行異常。寒冷悪化。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "対症療法のみ。激しい運動・寒冷を避ける。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Centronuclear Myopathy CNM 犬 ラブラドール")},
        ],
    },
    # === 腎・尿路系 ===
    {
        "match": ["cystinuria"],
        "title": "シスチン尿症 (Cystinuria / SLC3A1, SLC7A9)",
        "severity": "medium",
        "summary": "アミノ酸（シスチン）が尿中に過剰排泄され、結石を生じやすくなる遺伝性疾患です。",
        "mechanism": "SLC3A1 / SLC7A9 遺伝子の変異により腎尿細管でのシスチン再吸収が障害。シスチン結石が形成されます。",
        "symptoms": "頻尿・血尿・排尿困難・尿閉。雄犬で重症化しやすい。",
        "inheritance": "常染色体劣性または X 染色体連鎖（亜型による）。",
        "advice": "発症犬は低蛋白食・尿アルカリ化薬・水分摂取増。雄犬は尿閉に注意。",
        "references": [
            {"label": "Wikipedia: シスチン尿症", "url": _wiki_jp("シスチン尿症")},
            {"label": "詳細を検索", "url": _google_search("Cystinuria 犬 SLC3A1")},
        ],
    },
    {
        "match": ["familial nephropathy", "cocker nephropathy"],
        "title": "家族性腎症 (Familial Nephropathy / COL4A4)",
        "severity": "high",
        "summary": "若齢期から進行性腎不全を起こす遺伝性疾患です。コッカースパニエル等で報告。",
        "mechanism": "COL4A4 遺伝子の変異により糸球体基底膜が脆弱になり、進行性腎機能低下を起こします。",
        "symptoms": "多飲多尿 → 食欲低下・嘔吐 → 末期腎不全。多くは6〜24ヶ月齢で発症。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/P は予後不良。早期発見で食事療法・対症療法で延命可能。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Familial Nephropathy 犬 COL4A4")},
        ],
    },
    # === 皮膚・代謝系 ===
    {
        "match": ["hnpk", "hereditary nasal parakeratosis"],
        "title": "遺伝性鼻過角化症 (HNPK / SUV39H2)",
        "severity": "low",
        "summary": "鼻先の皮膚が異常に厚くなり、亀裂・痛み・痂皮を生じる遺伝性皮膚疾患です。",
        "mechanism": "SUV39H2 遺伝子の変異により鼻皮膚の角化が亢進。ラブラドールで多発。",
        "symptoms": "鼻先の硬化・痂皮・亀裂・出血。痛みあり。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "軟膏・保湿で QOL 維持可能。完治はしない。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("HNPK 鼻過角化 犬 SUV39H2")},
        ],
    },
    {
        "match": ["ichthyosis", "pnpla1"],
        "title": "魚鱗癬 (Ichthyosis / PNPLA1)",
        "severity": "medium",
        "summary": "皮膚が魚の鱗のように剥がれ落ちる遺伝性皮膚疾患です。",
        "mechanism": "PNPLA1 等の遺伝子変異により皮膚の角化過程が異常。ゴールデンレトリーバー等で多発。",
        "symptoms": "全身の白い鱗状の剥がれ・乾燥・痒み。皮脂分泌異常。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "シャンプー療法・保湿で症状コントロール。完治はしない。",
        "references": [
            {"label": "Wikipedia: 魚鱗癬", "url": _wiki_jp("魚鱗癬")},
            {"label": "詳細を検索", "url": _google_search("Ichthyosis 犬 PNPLA1")},
        ],
    },
    {
        "match": ["copper toxicosis", "commd1", "atp7"],
        "title": "銅蓄積性肝障害 (Copper Toxicosis / COMMD1, ATP7A/B)",
        "severity": "high",
        "summary": "肝臓に銅が異常蓄積し、慢性肝炎・肝硬変を起こす遺伝性疾患です。",
        "mechanism": "COMMD1 / ATP7A / ATP7B 遺伝子の変異により銅の排出が障害。ベドリントンテリア・ラブラドール等で報告。",
        "symptoms": "食欲不振・体重減少・腹水・黄疸。進行すると肝不全。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "低銅食・銅キレート薬で進行抑制可能。早期診断が重要。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Copper Toxicosis 銅蓄積 犬 COMMD1")},
        ],
    },
    # === 免疫系 ===
    {
        "match": ["trapped neutrophil syndrome", "\\btns\\b"],
        "title": "好中球機能不全症候群 (TNS / VPS13B)",
        "severity": "high",
        "summary": "白血球（好中球）が骨髄から血液へ移行できず、慢性免疫不全を起こす疾患です。",
        "mechanism": "VPS13B 遺伝子の変異により好中球の遊走機能が障害。ボーダーコリーで多発。",
        "symptoms": "繰り返す感染症・発熱・発育不良。多くは1歳までに死亡。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/P は予後不良。繁殖前検査が予防の鍵。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Trapped Neutrophil Syndrome TNS 犬 VPS13B")},
        ],
    },
    # === Veqta 検査パネル準拠の追加疾患 (PR #45) ===
    # 眼科系（追加）
    {
        "match": ["glaucoma", "緑内障", "primary glaucoma"],
        "title": "原発性緑内障 (Primary Glaucoma / ADAMTS10, ADAMTS17 等)",
        "severity": "high",
        "summary": "眼内圧が異常に上昇し、視神経が圧迫されて失明に至る遺伝性眼疾患です。",
        "mechanism": "前房隅角の閉塞や房水排出経路の異常により眼内圧が上昇。ADAMTS10/17 等の遺伝子変異が関与。ビーグル・コッカー・シーズーなどで多発。",
        "symptoms": "目の充血・角膜混濁・瞳孔散大・痛み・視覚消失。急性発作は緊急事態。",
        "inheritance": "犬種・遺伝子により異なる。多くは常染色体劣性または不完全優性。",
        "advice": "急性発作時は **24時間以内の眼科処置が視力保存の鍵**。定期的な眼圧測定を推奨。",
        "references": [
            {"label": "Wikipedia: 緑内障", "url": _wiki_jp("緑内障")},
            {"label": "詳細を検索", "url": _google_search("犬 緑内障 ADAMTS 原発性")},
        ],
    },
    {
        "match": ["cord1", "cord-1", "rpgrip1 pra"],
        "title": "PRA - CORD1 型 (RPGRIP1)",
        "severity": "high",
        "summary": "進行性網膜萎縮症の一型。錐体（昼間視）が先に変性するタイプです。",
        "mechanism": "RPGRIP1 遺伝子の変異により網膜光受容細胞の機能が失われます。ミニチュアロングヘアダックスフンドで報告。",
        "symptoms": "1〜2歳から昼間の視覚異常 → 進行性失明。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発症前に検査による遺伝子型把握が重要。失明後も嗅覚・聴覚で適応可能。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("PRA CORD1 犬 RPGRIP1")},
        ],
    },
    {
        "match": ["rcd3", "rcd-3", "pde6a pra", "rod cone dysplasia"],
        "title": "PRA - rcd3 型 (PDE6A)",
        "severity": "high",
        "summary": "若齢期から急速に進行する PRA の一型。生後数ヶ月から発症します。",
        "mechanism": "PDE6A 遺伝子の変異により杆体（夜間視）細胞が早期に変性。",
        "symptoms": "生後6ヶ月頃から夜盲 → 1〜2歳で完全失明。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "重度のため早期検査・繁殖選択が重要。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("PRA rcd3 犬 PDE6A")},
        ],
    },
    {
        "match": ["cngb1 pra", "cngb1"],
        "title": "PRA - CNGB1 型",
        "severity": "high",
        "summary": "PRA の遅発型。比較的緩やかな進行を示す型です。",
        "mechanism": "CNGB1 遺伝子の変異により網膜杆体の機能が低下。",
        "symptoms": "中年期から夜盲が始まり、徐々に昼間視覚も低下。",
        "inheritance": "常染色体劣性。",
        "advice": "進行が緩やかなため早期発見で QOL を保ちやすい。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("PRA CNGB1 犬")},
        ],
    },
    {
        "match": ["congenital stationary night blindness", "csnb", "cnsb"],
        "title": "先天性停止性夜盲症 (CSNB / RPE65)",
        "severity": "low",
        "summary": "生まれつき夜間視力が無い遺伝性眼疾患です。昼間は通常通り見えます。",
        "mechanism": "RPE65 遺伝子の変異により網膜の桿体細胞での視物質再生が障害。ブリアード犬で報告。",
        "symptoms": "生後から夜盲・薄暗い場所での視覚困難。進行性ではない（停止性）。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "夜間の照明確保で QOL 維持可能。遺伝子治療研究も進行中。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("CSNB 先天性停止性夜盲 犬 RPE65")},
        ],
    },
    # 神経系（追加）
    {
        "match": ["gm1 gangliosidosis", "gm-1"],
        "title": "ガングリオシドーシス GM1 (GM1 / GLB1)",
        "severity": "high",
        "summary": "GM1 ガングリオシドが神経細胞内に蓄積し進行性神経障害を起こす疾患です。",
        "mechanism": "GLB1 遺伝子の変異によりβ-ガラクトシダーゼ酵素が欠損。GM1 が分解されず蓄積します。シーバ・柴犬・スパニエル等で報告。",
        "symptoms": "若齢発症で運動失調・痙攣・視覚障害が進行。多くは2〜3歳までに死亡。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/N 同士の交配は厳に避ける。発症犬は予後不良。",
        "references": [
            {"label": "Wikipedia: ガングリオシドーシス", "url": _wiki_jp("ガングリオシドーシス")},
            {"label": "詳細を検索", "url": _google_search("GM1 Gangliosidosis 犬 GLB1")},
        ],
    },
    {
        "match": ["myotonia congenita", "先天性筋強直症", "clcn1"],
        "title": "先天性筋強直症 (Myotonia Congenita / CLCN1)",
        "severity": "medium",
        "summary": "筋肉が一度収縮すると弛緩しにくくなる遺伝性筋疾患です。",
        "mechanism": "CLCN1 遺伝子の変異により筋細胞膜のクロライドチャネル機能不全。筋電気活動が異常持続します。",
        "symptoms": "運動開始時のこわばり・転倒・歩行困難。運動を続けると改善（ウォームアップ現象）。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "対症療法のみ。寒冷を避け、ウォームアップを十分にとる。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Myotonia Congenita 犬 CLCN1")},
        ],
    },
    # 血液系（追加）
    {
        "match": [
            "willebrand type 2", "willebrand's type 2",
            "willebrand disease type 2", "willebrand's disease type 2",
            "vwd2", "vwd-2", "vwd type 2", "vwd type ii",
            "willebrand ii", "willebrand 2",
        ],
        "title": "フォン・ヴィレブランド病 II型 (vWD2)",
        "severity": "high",
        "summary": "vWD の中等度〜重度型。I型より出血傾向が強くなります。",
        "mechanism": "vWF 遺伝子の質的変異により、止血因子の機能が低下（量はあるが機能しない）。",
        "symptoms": "外傷・抜歯後の長時間出血、自然出血のリスクあり。",
        "inheritance": "常染色体劣性（不完全浸透）。",
        "advice": "手術前に必ず申告。出血時は新鮮凍結血漿・vWF 製剤が必要。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("von Willebrand Type 2 犬 vWD2")},
        ],
    },
    {
        "match": [
            "willebrand type 3", "willebrand's type 3",
            "willebrand disease type 3", "willebrand's disease type 3",
            "vwd3", "vwd-3", "vwd type 3", "vwd type iii",
            "willebrand iii", "willebrand 3",
        ],
        "title": "フォン・ヴィレブランド病 III型 (vWD3)",
        "severity": "high",
        "summary": "最重度の vWD。出血が止まらず生命に関わる重篤型です。",
        "mechanism": "vWF 蛋白質が完全に欠損。スコティッシュテリア等で報告。",
        "symptoms": "幼齢期から重度出血。歯科処置・外傷で致命的になり得る。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/P 犬は手術リスク極大。輸血製剤の準備が必須。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("von Willebrand Type 3 犬 vWD3")},
        ],
    },
    {
        "match": ["prekallikrein", "プレカリクレイン"],
        "title": "プレカリクレイン欠乏症 (Prekallikrein Deficiency / KLKB1)",
        "severity": "low",
        "summary": "凝固系の前段階因子が欠損し、APTT 検査値が異常を示す凝固障害です。",
        "mechanism": "KLKB1 遺伝子の変異により内因系凝固第一段階が遅延。多くは無症状。",
        "symptoms": "通常は無症状だが、手術時に APTT 延長が見られる。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "手術前検査で APTT 延長が出たら本症を疑う。臨床的影響は少ない。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Prekallikrein Deficiency 犬 KLKB1")},
        ],
    },
    # 代謝系（追加）
    {
        "match": ["glycogen storage disease", "gsd", "グリコーゲン蓄積症"],
        "title": "グリコーゲン蓄積症 (GSD / 複数型)",
        "severity": "high",
        "summary": "グリコーゲンが分解できず体内に蓄積する重篤な代謝疾患です。型により症状が異なります。",
        "mechanism": "GAA (II型), GBE1 (IV型) などの遺伝子変異によりグリコーゲン代謝酵素が欠損。",
        "symptoms": "II型: 心筋・骨格筋の障害、IV型: 肝硬変・神経症状。多くは若齢で重症化。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/P は予後不良。早期診断と対症療法。",
        "references": [
            {"label": "Wikipedia: 糖原病", "url": _wiki_jp("糖原病")},
            {"label": "詳細を検索", "url": _google_search("Glycogen Storage Disease 犬 GSD")},
        ],
    },
    {
        "match": ["cobalamin malabsorption", "コバラミン吸収", "imerslund", "amn cubn"],
        "title": "コバラミン吸収不良症 (Cobalamin Malabsorption / AMN, CUBN)",
        "severity": "medium",
        "summary": "ビタミンB12（コバラミン）の腸吸収障害により神経・血液障害を起こす遺伝性疾患です。",
        "mechanism": "AMN または CUBN 遺伝子の変異により回腸でのコバラミン受容体が機能不全。",
        "symptoms": "成長不良・貧血・神経症状（運動失調・痙攣）。ジャイアントシュナウザー等で報告。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "**コバラミン (B12) 注射で症状改善可能** — 治療可能な疾患のため診断が重要。",
        "references": [
            {"label": "Wikipedia: ビタミンB12欠乏症", "url": _wiki_jp("ビタミンB12欠乏症")},
            {"label": "詳細を検索", "url": _google_search("Cobalamin Malabsorption 犬 AMN CUBN")},
        ],
    },
    # 骨格系（追加）
    {
        "match": ["osteogenesis imperfecta", "骨形成不全", "serpinh1", "col1a"],
        "title": "骨形成不全症 (Osteogenesis Imperfecta / SERPINH1, COL1A1, COL1A2)",
        "severity": "high",
        "summary": "骨が脆くなり、軽度の衝撃でも骨折しやすい遺伝性疾患です。",
        "mechanism": "コラーゲン関連遺伝子 (SERPINH1, COL1A1, COL1A2) の変異により骨の構造蛋白質が異常になります。",
        "symptoms": "若齢期からの繰り返す骨折・歯の異常・関節弛緩。ダックスフンド・ビーグル等で報告。",
        "inheritance": "常染色体劣性（亜型による）。",
        "advice": "発症犬は運動制限・骨折予防が必要。栄養管理（カルシウム・コラーゲン）も重要。",
        "references": [
            {"label": "Wikipedia: 骨形成不全症", "url": _wiki_jp("骨形成不全症")},
            {"label": "詳細を検索", "url": _google_search("Osteogenesis Imperfecta 犬 SERPINH1")},
        ],
    },
    # 消化器系（追加）
    {
        "match": ["gastric and intestinal polyposis", "消化管ポリポーシス", "gastrointestinal polyposis"],
        "title": "消化管ポリポーシス (GP)",
        "severity": "high",
        "summary": "胃・腸にポリープが多発し、出血・腸閉塞のリスクを増す遺伝性疾患です。",
        "mechanism": "遺伝性ポリープ形成傾向。ジャックラッセルテリア等で報告。",
        "symptoms": "繰り返す消化器症状・血便・体重減少・嘔吐。",
        "inheritance": "常染色体劣性または優性。",
        "advice": "定期的な内視鏡検査でポリープ管理。悪性化リスクあり。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Gastric Intestinal Polyposis 犬")},
        ],
    },
    # === Embark/Orivet 標準パネル拡張 (PR #47) ===
    # 神経系（追加）
    {
        "match": ["cerebellar abiotrophy", "小脳皮質変性", "cerebellar degeneration"],
        "title": "小脳皮質変性症 (Cerebellar Abiotrophy)",
        "severity": "high",
        "summary": "小脳のプルキンエ細胞が出生後変性し、運動失調を起こす遺伝性神経疾患です。",
        "mechanism": "犬種により原因遺伝子が異なる (GRM1, SPTBN2 等)。プルキンエ細胞の選択的死滅。",
        "symptoms": "若齢期 (3〜12ヶ月) からふらつき歩行・頭の震え・転倒。進行性。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "対症療法のみ。穏やかな環境での QOL 維持。",
        "references": [
            {"label": "Wikipedia: 小脳萎縮症", "url": _wiki_jp("小脳萎縮症")},
            {"label": "詳細を検索", "url": _google_search("Cerebellar Abiotrophy 犬")},
        ],
    },
    {
        "match": ["sensory neuropathy", "感覚性神経障害", "fam134b"],
        "title": "感覚性神経障害 (SN / FAM134B)",
        "severity": "high",
        "summary": "末梢神経の感覚線維が変性し、痛覚消失や自咬行動を起こす遺伝性疾患です。",
        "mechanism": "FAM134B 遺伝子の変異により末梢感覚神経が変性。ボーダーコリーで報告。",
        "symptoms": "肢端の感覚消失 → 自咬 → 潰瘍・感染。歩行運動には支障なし。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発症犬は自咬予防 (エリザベスカラー等) と感染症管理が必要。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Sensory Neuropathy 犬 FAM134B")},
        ],
    },
    {
        "match": ["globoid cell leukodystrophy", "krabbe", "galc"],
        "title": "球状細胞白質ジストロフィー (Krabbe / GALC)",
        "severity": "high",
        "summary": "中枢・末梢神経のミエリン形成が異常になる重篤な遺伝性代謝疾患です。",
        "mechanism": "GALC 遺伝子の変異によりガラクトセレブロシダーゼ酵素が欠損。ミエリン分解産物が蓄積し神経変性を起こします。",
        "symptoms": "生後 1〜6ヶ月で運動異常・痙攣・成長不全。多くは 1〜2 年で死亡。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "P/P は予後不良。骨髄移植研究中だが実用化はまだ。繁殖前検査必須。",
        "references": [
            {"label": "Wikipedia: クラッベ病", "url": _wiki_jp("クラッベ病")},
            {"label": "詳細を検索", "url": _google_search("Globoid Cell Leukodystrophy 犬 GALC")},
        ],
    },
    {
        "match": ["polyneuropathy", "多発性神経障害", "ndrg1", "arhgef10"],
        "title": "多発性神経障害 (Polyneuropathy / NDRG1, ARHGEF10)",
        "severity": "high",
        "summary": "複数の末梢神経が同時に障害を受け、運動失調・筋萎縮を起こす疾患です。",
        "mechanism": "NDRG1 や ARHGEF10 等の変異により末梢神経の機能が低下。グレーハウンド・アラスカン・マラミュート等で報告。",
        "symptoms": "若齢期から後肢の脱力・歩行異常・筋萎縮。",
        "inheritance": "犬種により異なる（多くは常染色体劣性）。",
        "advice": "対症療法のみ。物理療法・補助具で QOL 維持。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Polyneuropathy 犬 NDRG1")},
        ],
    },
    {
        "match": ["episodic falling", "episodic falling syndrome", "efs"],
        "title": "発作性失神症 (EFS / BCAN)",
        "severity": "medium",
        "summary": "興奮や運動で発作的に筋肉が硬直し転倒する遺伝性疾患です。キャバリア K.C. スパニエルで頻発。",
        "mechanism": "BCAN 遺伝子の変異により神経シグナル伝達が異常になり、運動時に発作を起こします。",
        "symptoms": "興奮・運動・暑さで筋強直 → 転倒・硬直。意識は保持。数秒〜数分で回復。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発作誘因 (興奮・暑さ) を避け、抗痙攣薬で予防可能なケースあり。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Episodic Falling Syndrome 犬 BCAN")},
        ],
    },
    {
        "match": ["l 2 hydroxyglutaric aciduria", "hydroxyglutaric aciduria", "l2hga"],
        "title": "L-2-ヒドロキシグルタル酸尿症 (L2HGA / L2HGDH)",
        "severity": "high",
        "summary": "代謝産物が異常蓄積し神経症状を起こす希少な遺伝性代謝疾患です。スタッフィー系で報告。",
        "mechanism": "L2HGDH 遺伝子の変異により L-2-ヒドロキシグルタル酸の分解酵素が欠損。",
        "symptoms": "若齢期から運動失調・痙攣・行動異常・認知障害。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "対症療法のみ。スタッフォードシャーブルテリア・アメリカン・スタッフォードシャー・テリアで頻発。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("L-2-Hydroxyglutaric Aciduria 犬 L2HGDH")},
        ],
    },
    # 眼科系（追加）
    {
        "match": ["multifocal retinopathy", "cmr1", "cmr2", "best1"],
        "title": "多巣性網膜症 (CMR / BEST1)",
        "severity": "low",
        "summary": "網膜に複数の隆起・剥離が起こる遺伝性眼疾患です。多くは進行が緩やか。",
        "mechanism": "BEST1 遺伝子の変異により網膜色素上皮の機能が異常。マスティフ・ピレネー等で報告。",
        "symptoms": "初期は無症状。中年期から斑状の網膜病変が見られる。視覚障害は軽度なことが多い。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "進行が緩やかなため早期管理で QOL 良好。眼科専門医の定期検査推奨。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Canine Multifocal Retinopathy CMR BEST1")},
        ],
    },
    {
        "match": ["cone degeneration", "achromatopsia cnga3", "day-blindness"],
        "title": "錐体ジストロフィー (Cone Degeneration / CNGB3)",
        "severity": "medium",
        "summary": "網膜の錐体細胞が機能不全になり、明所視・色覚に障害を起こす疾患です。",
        "mechanism": "CNGB3 遺伝子の変異により錐体光受容細胞が変性。アラスカン・マラミュート等で報告。",
        "symptoms": "生後数ヶ月から明所での視覚困難・色覚消失。夜間視は保持。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "失明とは異なる『昼盲』状態。暗い環境で QOL 維持可能。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Cone Degeneration 犬 CNGB3")},
        ],
    },
    {
        "match": ["stargardt", "stgd1", "abca4"],
        "title": "スターガルト病 (STGD1 / ABCA4)",
        "severity": "high",
        "summary": "若齢期から黄斑部の網膜が変性する遺伝性眼疾患です。",
        "mechanism": "ABCA4 遺伝子の変異により網膜色素上皮にリポフスチンが蓄積。ラブラドール等で報告。",
        "symptoms": "1〜2 歳から中心視野の視覚低下 → 進行性。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発症前検査による遺伝子型把握で繁殖選択が重要。",
        "references": [
            {"label": "Wikipedia: スターガルト病", "url": _wiki_jp("スターガルト病")},
            {"label": "詳細を検索", "url": _google_search("Stargardt 犬 ABCA4")},
        ],
    },
    # 皮膚系（追加）
    {
        "match": ["coat color dilution alopecia", "color dilution alopecia", "cda"],
        "title": "毛色希釈性脱毛症 (CDA)",
        "severity": "low",
        "summary": "希釈毛色 (Blue, Lilac 等) を持つ犬で起こる遺伝性脱毛症です。",
        "mechanism": "メラニン顆粒の異常蓄積により毛包が損傷。MLPH 遺伝子変異 (dd) を持つ希釈色犬の一部で発症。",
        "symptoms": "若齢期 (6ヶ月〜2歳) から希釈色部分の脱毛・痂皮・二次感染。非希釈色部位は正常。",
        "inheritance": "希釈遺伝子 dd 保持者のうち一部で発症 (多因子)。",
        "advice": "皮膚ケア・抗生剤治療で症状管理。発症犬の繁殖は再考を。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Color Dilution Alopecia 犬 CDA")},
        ],
    },
    {
        "match": ["footpad hyperkeratosis", "hfh", "肉球角化"],
        "title": "肉球角化亢進症 (HFH / FAM83G)",
        "severity": "medium",
        "summary": "肉球が異常に厚く硬化し、亀裂・痛みを生じる遺伝性皮膚疾患です。",
        "mechanism": "FAM83G 遺伝子の変異により肉球の角化が亢進。アイリッシュ・テリア、ドゴ・ド・ボルドー等で報告。",
        "symptoms": "若齢期から全肉球の硬化・亀裂・痛み。歩行困難を起こすこともある。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "軟膏・保湿で QOL 維持。完治はしない。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Hereditary Footpad Hyperkeratosis 犬 HFH FAM83G")},
        ],
    },
    # 内分泌・腎系（追加）
    {
        "match": ["renal cystadenocarcinoma", "rcnd", "nodular dermatofibrosis", "腎嚢腺癌"],
        "title": "腎嚢腺癌・結節性皮膚線維腫症 (RCND / FLCN)",
        "severity": "high",
        "summary": "腎臓に多発する腫瘍と皮膚結節を起こす遺伝性腫瘍症候群です。ジャーマン・シェパード特有。",
        "mechanism": "FLCN 遺伝子の変異により腫瘍抑制機能が失われ、腎臓・皮膚に腫瘍が発生。",
        "symptoms": "中年期から皮膚に多数の結節 → 後に腎腫瘍・腎不全。",
        "inheritance": "常染色体優性。1コピーで発症リスクあり。",
        "advice": "定期的な腎機能検査・腹部画像診断が早期発見の鍵。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Renal Cystadenocarcinoma RCND 犬 FLCN")},
        ],
    },
    {
        "match": ["hyperphosphatemia", "fgf23", "高リン血症"],
        "title": "家族性高リン血症 (Hyperphosphatemia / FGF23)",
        "severity": "medium",
        "summary": "血中リン濃度が異常に高くなる遺伝性内分泌疾患です。",
        "mechanism": "FGF23 関連遺伝子の変異によりリン代謝が異常。骨・軟組織の異常石灰化を引き起こす。",
        "symptoms": "成長異常・骨格変形・腎機能障害。重症例では予後不良。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発症犬は低リン食・リン結合薬で進行抑制。",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Hyperphosphatemia 犬 FGF23")},
        ],
    },
    # === Embark 準拠の追加疾患 (PR #49) ===
    # 神経・脳系（追加）
    {
        "match": ["necrotizing meningoencephalitis", "nme", "pug encephalitis", "壊死性髄膜脳炎"],
        "title": "壊死性髄膜脳炎 (NME / Pug Encephalitis)",
        "summary": "脳実質に重度の壊死性炎症を起こす重篤な遺伝性自己免疫疾患です。パグ・マルチーズ等で多発。",
        "mechanism": "MHC class II 関連の遺伝子多型による自己免疫反応で、脳・髄膜に壊死病変を形成。",
        "symptoms": "若齢期 (1〜7歳) から痙攣・行動異常・失調・盲目。多くは数週間〜数ヶ月で死亡。",
        "inheritance": "遺伝的素因 + 環境要因（多遺伝子）。リスク遺伝子型保有犬は発症率が高い。",
        "advice": "発症犬は免疫抑制療法で進行抑制可能だが予後不良。リスク遺伝子型犬の繁殖は再考を。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Necrotizing Meningoencephalitis NME 犬 パグ")},
        ],
    },
    {
        "match": ["lafora", "nhlrc1", "ラフォラ"],
        "title": "ラフォラ病 (Lafora Disease / NHLRC1)",
        "summary": "若齢期から進行性ミオクローヌス癲癇を起こす重篤な遺伝性神経疾患です。",
        "mechanism": "NHLRC1 遺伝子の変異によりラフォラ小体（多糖類凝集体）が神経細胞に蓄積。ミニチュアダックスフンド・バセットハウンド等で報告。",
        "symptoms": "5〜10歳から短時間のジャーキング (ミオクローヌス) → 進行性発作・認知低下。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "発症抑制薬で進行を遅らせる可能性。繁殖前検査が予防の鍵。",
        "severity": "high",
        "references": [
            {"label": "Wikipedia: ラフォラ病", "url": _wiki_jp("ラフォラ病")},
            {"label": "詳細を検索", "url": _google_search("Lafora Disease 犬 NHLRC1")},
        ],
    },
    {
        "match": ["narcolepsy", "ナルコレプシー", "hcrtr2"],
        "title": "ナルコレプシー (Narcolepsy / HCRTR2)",
        "summary": "突然の脱力と睡眠発作を起こす遺伝性睡眠障害です。",
        "mechanism": "HCRTR2 遺伝子の変異によりオレキシン受容体が機能不全。覚醒・睡眠の制御が異常になります。ドーベルマン等で報告。",
        "symptoms": "興奮や食事中に突然倒れて眠る・脱力 (カタプレキシー)。意識は保持。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "中枢神経刺激薬・三環系抗うつ薬で症状軽減可能。生命予後は良好。",
        "severity": "medium",
        "references": [
            {"label": "Wikipedia: ナルコレプシー", "url": _wiki_jp("ナルコレプシー")},
            {"label": "詳細を検索", "url": _google_search("Narcolepsy 犬 HCRTR2")},
        ],
    },
    {
        "match": ["neuroaxonal dystrophy", "nad", "pla2g6", "神経軸索ジストロフィー"],
        "title": "神経軸索ジストロフィー (NAD / PLA2G6)",
        "summary": "中枢・末梢神経の軸索が変性し、運動失調が進行する遺伝性神経疾患です。",
        "mechanism": "PLA2G6 遺伝子の変異により神経軸索の膜代謝が異常。スピノーン・パピヨン等で報告。",
        "symptoms": "若齢期 (1〜4歳) から運動失調・転倒・固有感覚異常。進行性。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "対症療法のみ。物理療法で QOL 維持。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Neuroaxonal Dystrophy NAD 犬 PLA2G6")},
        ],
    },
    {
        "match": ["spongiform leukoencephalomyelopathy", "slem", "海綿状白質脳脊髄症"],
        "title": "海綿状白質脳脊髄症 (SLEM)",
        "summary": "脳・脊髄の白質に空胞変性を起こす重篤な若齢期発症の遺伝性疾患です。",
        "mechanism": "ミトコンドリア機能異常により神経軸索ミエリンが変性。シルキー・テリア等で報告。",
        "symptoms": "生後数週間から運動失調・痙攣・成長停止。多くは数ヶ月で死亡。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "予後不良。繁殖前検査必須。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Spongiform Leukoencephalomyelopathy SLEM 犬")},
        ],
    },
    # 免疫系（追加）
    {
        "match": ["severe combined immunodeficiency", "scid", "重症複合免疫不全"],
        "title": "重症複合免疫不全症 (SCID)",
        "summary": "T 細胞・B 細胞が機能せず重度免疫不全を起こす致死性の遺伝性疾患です。",
        "mechanism": "RAG1/RAG2 や DCLRE1C 等の遺伝子変異により T/B 細胞の発達が阻害される。バセットハウンド・ジャックラッセルテリア等で報告。",
        "symptoms": "生後数週間から繰り返す重症感染症・成長不良。骨髄移植以外では多くが幼齢で死亡。",
        "inheritance": "常染色体劣性または X 連鎖。両親キャリアで 25% 発症。",
        "advice": "骨髄移植のみ根治療法。繁殖前検査必須。",
        "severity": "high",
        "references": [
            {"label": "Wikipedia: 重症複合免疫不全症", "url": _wiki_jp("重症複合免疫不全症")},
            {"label": "詳細を検索", "url": _google_search("SCID 犬 重症複合免疫不全")},
        ],
    },
    {
        "match": ["myasthenia gravis", "重症筋無力症", "chrne"],
        "title": "先天性重症筋無力症 (CMG / CHRNE)",
        "summary": "神経筋接合部の機能不全により筋力低下を起こす遺伝性疾患です。",
        "mechanism": "CHRNE 遺伝子の変異によりアセチルコリン受容体が機能不全。シグナル伝達が阻害される。",
        "symptoms": "若齢期から運動不耐・四肢筋力低下・嚥下困難。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "抗コリンエステラーゼ薬で症状改善可能。早期診断・治療が重要。",
        "severity": "medium",
        "references": [
            {"label": "Wikipedia: 重症筋無力症", "url": _wiki_jp("重症筋無力症")},
            {"label": "詳細を検索", "url": _google_search("Myasthenia Gravis 犬 CHRNE")},
        ],
    },
    # 内分泌・発達系（追加）
    {
        "match": ["pituitary dwarfism", "下垂体性小人症", "lhx3"],
        "title": "下垂体性小人症 (Pituitary Dwarfism / LHX3)",
        "summary": "下垂体機能不全により成長ホルモン等が不足し、極端な低身長になる遺伝性疾患です。",
        "mechanism": "LHX3 遺伝子の変異により下垂体前葉の発達が異常。ジャーマンシェパード等で報告。",
        "symptoms": "成長不良・体格小さい・被毛異常・甲状腺機能低下を伴う。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "成長ホルモン・甲状腺ホルモン補充療法で QOL 改善可能。",
        "severity": "medium",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Pituitary Dwarfism 犬 LHX3")},
        ],
    },
    {
        "match": ["persistent mullerian duct syndrome", "pmds", "ミュラー管遺残"],
        "title": "ミュラー管遺残症候群 (PMDS / AMHR2)",
        "summary": "オス犬が子宮・卵管などのメス内性器を持つ遺伝性発達異常です。外見はオス。",
        "mechanism": "AMHR2 遺伝子の変異により抗ミュラー管ホルモンの作用が失われ、メス内性器が退縮しない。ミニチュアシュナウザー等で報告。",
        "symptoms": "外見はオスだが繁殖能力低下・尿路感染・前立腺問題。",
        "inheritance": "常染色体劣性。X 染色体遺伝。",
        "advice": "発症犬は外科的処置を要する場合あり。繁殖前検査が予防の鍵。",
        "severity": "low",
        "references": [
            {"label": "詳細を検索", "url": _google_search("PMDS 犬 ミュラー管 AMHR2")},
        ],
    },
    # 筋骨格系（追加）
    {
        "match": ["limb-girdle muscular dystrophy", "lgmd", "肢帯型筋ジストロフィー"],
        "title": "肢帯型筋ジストロフィー (LGMD)",
        "summary": "肢帯部（肩・腰）の筋肉が進行性に萎縮する遺伝性筋疾患です。",
        "mechanism": "ジストロフィン関連タンパク (DMD, SGCD 等) の変異により筋細胞膜の構造が異常になります。",
        "symptoms": "若齢期から肩・腰の筋萎縮・運動不耐・歩行困難。",
        "inheritance": "X 連鎖劣性または常染色体劣性 (亜型による)。",
        "advice": "対症療法のみ。物理療法・補助具で QOL 維持。",
        "severity": "medium",
        "references": [
            {"label": "Wikipedia: 肢帯型筋ジストロフィー", "url": _wiki_jp("肢帯型筋ジストロフィー")},
            {"label": "詳細を検索", "url": _google_search("Limb-Girdle Muscular Dystrophy 犬 LGMD")},
        ],
    },
    {
        "match": ["skeletal dysplasia 2", "sd2", "骨格異形成2"],
        "title": "骨格異形成 2 型 (Skeletal Dysplasia 2 / COL11A2)",
        "summary": "ラブラドール特有の四肢短縮・骨格異形成を起こす遺伝性疾患です。",
        "mechanism": "COL11A2 遺伝子の変異によりコラーゲン形成が異常。四肢の短縮・前肢の弓形変形。",
        "symptoms": "生後数ヶ月から四肢の短縮・前肢弓形変形・関節異常。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "整形外科的管理・運動制限で QOL 維持。",
        "severity": "medium",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Skeletal Dysplasia 2 犬 COL11A2 ラブラドール")},
        ],
    },
    # 腎・呼吸系（追加）
    {
        "match": ["x-linked hereditary nephropathy", "xlhn", "x連鎖腎症"],
        "title": "X 連鎖性遺伝性腎症 (XLHN / COL4A5)",
        "summary": "オス犬で重度に発症する遺伝性腎症。サモエド等で報告。",
        "mechanism": "X 染色体上の COL4A5 遺伝子変異により糸球体基底膜が脆弱に。",
        "symptoms": "オス: 3〜6ヶ月齢から多飲多尿・進行性腎不全。メス: 軽度症状。",
        "inheritance": "X 連鎖劣性。オスは X 1コピーで発症、メスはキャリア。",
        "advice": "オス犬の P/Y は予後不良。早期発見で食事療法・対症療法。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索", "url": _google_search("X-Linked Hereditary Nephropathy 犬 COL4A5")},
        ],
    },
    {
        "match": ["recurrent inflammatory pulmonary disease", "ripd", "再発性炎症性肺疾患"],
        "title": "再発性炎症性肺疾患 (RIPD / AKNA)",
        "summary": "繰り返す肺炎・気管支炎を起こすロットワイラー特有の遺伝性免疫疾患です。",
        "mechanism": "AKNA 遺伝子の変異により気道粘膜の免疫機能が低下。",
        "symptoms": "若齢期から繰り返す肺感染・呼吸困難・運動不耐。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "抗生剤・気管支拡張剤で対症療法。重症例では長期管理が必要。",
        "severity": "medium",
        "references": [
            {"label": "詳細を検索", "url": _google_search("Recurrent Inflammatory Pulmonary Disease RIPD 犬")},
        ],
    },
    # 眼科系（追加）
    {
        "match": ["rcd1", "pra rcd-1", "pde6b pra", "rod cone dysplasia 1"],
        "title": "PRA - rcd1 型 (PDE6B)",
        "summary": "アイリッシュセッターで報告される重度若齢期 PRA。生後数ヶ月で完全失明。",
        "mechanism": "PDE6B 遺伝子の変異により杆体光受容細胞が急速変性。",
        "symptoms": "生後数ヶ月から夜盲 → 1歳までに完全失明。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "予後不良の重度 PRA。繁殖前検査必須。",
        "severity": "medium",
        "references": [
            {"label": "詳細を検索", "url": _google_search("PRA rcd1 犬 PDE6B アイリッシュセッター")},
        ],
    },
    {
        "match": ["rcd2", "pra rcd-2", "rd3 pra", "rod cone dysplasia 2"],
        "title": "PRA - rcd2 型 (RD3)",
        "summary": "コリー系で報告される若齢期発症 PRA。",
        "mechanism": "RD3 遺伝子の変異により網膜光受容細胞が変性。",
        "symptoms": "生後 6〜12 週から夜盲 → 進行性失明。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。",
        "advice": "重度の若齢期 PRA。繁殖前検査必須。",
        "severity": "medium",
        "references": [
            {"label": "詳細を検索", "url": _google_search("PRA rcd2 犬 RD3 コリー")},
        ],
    },
    # 眼科系（追加）
    {
        "match": ["primary lens luxation", "pll", "水晶体脱臼", "lens luxation", "adamts17 lens", "adamts10 lens"],
        "title": "原発性水晶体脱臼 (PLL / ADAMTS17)",
        "summary": "水晶体を支えるチン小帯が断裂し、水晶体が脱臼する遺伝性眼科疾患。急性緑内障や失明リスクがあります。",
        "mechanism": "ADAMTS17（一部犬種は ADAMTS10）遺伝子変異によりチン小帯タンパク質が機能不全となり、水晶体位置が保持できなくなります。",
        "symptoms": "2〜8 歳での急激な目の充血・眼圧上昇・疼痛。前方脱臼では角膜浮腫、緑内障を続発し失明の危険があります。",
        "inheritance": "常染色体劣性（P/P は発症、P/N はキャリアで発症リスク上昇）。テリア系・ボーダーコリー・シベリアンハスキー等で多報告。",
        "advice": "P/P（ホモ陽性）や P/N 犬は年 1 回以上の眼圧測定を推奨。前方脱臼は眼科的緊急事態のため即時受診を。繁殖には P/N × N/N の組み合わせを推奨。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索 (PLL ADAMTS17)", "url": _google_search("Primary Lens Luxation PLL 犬 ADAMTS17")},
        ],
    },
    # 代謝系（追加）
    {
        "match": ["phosphofructokinase deficiency", "pfk deficiency", "pfkm", "ホスホフルクトキナーゼ欠損", "pfk欠損"],
        "title": "ホスホフルクトキナーゼ欠損症 (PFK / PFKM)",
        "summary": "解糖系の重要酵素 PFK-M が欠損し、溶血性貧血と運動不耐を引き起こす代謝疾患。主にイングリッシュ・スプリンガー・スパニエルで報告。",
        "mechanism": "PFKM 遺伝子の変異により赤血球・筋肉の解糖系酵素活性が著しく低下。赤血球が早期に崩壊し溶血性貧血を生じます。",
        "symptoms": "運動後の筋力低下・黄疸・血色素尿（ヘモグロビン尿）・疲労感。激しい運動や発熱・アルカローシスで溶血発作が誘発されます。",
        "inheritance": "常染色体劣性。両親キャリアで 25% 発症。スプリンガー・スパニエル、アメリカン・コッカー・スパニエル等で多報告。",
        "advice": "激しい運動・過換気（吠え続け等）を避ける。発症犬は溶血発作管理が中心。繁殖では両親の遺伝子検査が必須。",
        "severity": "medium",
        "references": [
            {"label": "詳細を検索 (PFK 犬)", "url": _google_search("Phosphofructokinase Deficiency PFK 犬 溶血性貧血")},
        ],
    },
    # 心臓系（新カテゴリ）
    {
        "match": ["arvc", "arrhythmogenic right ventricular cardiomyopathy", "striatin", "strn", "不整脈源性右室心筋症", "右室心筋症"],
        "title": "不整脈源性右室心筋症 (ARVC / Striatin)",
        "summary": "右心室心筋が脂肪・線維組織に置き換わる遺伝性心筋症。突然死や重篤な不整脈を引き起こすボクサー特有の疾患です。",
        "mechanism": "Striatin（STRN）遺伝子の重複挿入変異により心室筋の細胞間接合タンパク質が機能不全となり、心筋がアポトーシスを起こして脂肪・線維組織に置換されます。",
        "symptoms": "虚脱・失神・突然死。多形性心室頻拍（VT）や心室細動（VF）を示す。若齢〜中年で発症することが多く、無症状のまま突然死する例も。",
        "inheritance": "常染色体優性（不完全浸透）。片方の遺伝子コピーだけでも発症リスクが上昇。ボクサーで最多報告、ドーベルマンにも類似疾患あり。",
        "advice": "ホモ（A/A）は高リスク・繁殖不可推奨。ヘテロ（A/N）は定期的なホルター心電図モニタリングを推奨。抗不整脈薬で管理可能な場合もあるが予後は個体差が大きい。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索 (ARVC ボクサー)", "url": _google_search("ARVC Arrhythmogenic Right Ventricular Cardiomyopathy 犬 ボクサー Striatin")},
        ],
    },
    {
        "match": ["dilated cardiomyopathy", "dcm1", "dcm type 1", "pdk4", "拡張型心筋症", "dcm 犬"],
        "title": "拡張型心筋症 1 型 (DCM1 / PDK4)",
        "summary": "心室壁が薄くなり心臓が拡張する遺伝性心筋症。ドーベルマンで高頻度に報告され、急性心不全や突然死のリスクがあります。",
        "mechanism": "PDK4 遺伝子のイントロン挿入変異が報告されており（特にドーベルマン）、心筋のエネルギー代謝および収縮機能が低下。心室が代償性に拡張し、最終的に収縮不全に至ります。",
        "symptoms": "倦怠感・運動不耐・呼吸困難・咳。心房細動や突然死を起こすこともある。ドーベルマンでは無症状期が長く、ホルター心電図による定期検査が必要。",
        "inheritance": "常染色体優性（ドーベルマン型）または多因子遺伝。大型・巨大犬種（ドーベルマン・アイリッシュ・ウルフハウンド・大型シュナウザー等）で多報告。",
        "advice": "ドーベルマンは 3 歳以降から年 1 回以上のホルター心電図 + 心臓超音波検査を推奨。ACE 阻害剤などで進行を遅らせる可能性がある。繁殖には遺伝子検査＋心臓スクリーニングを組み合わせた評価を。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索 (DCM 犬 PDK4)", "url": _google_search("Dilated Cardiomyopathy DCM 犬 PDK4 ドーベルマン")},
        ],
    },
    # 神経系（追加）
    {
        "match": ["benign familial juvenile epilepsy", "bfje", "lgi2", "良性家族性若年性てんかん", "juvenile epilepsy lagotto"],
        "title": "良性家族性若年性てんかん (BFJE / LGI2)",
        "summary": "ラゴット・ロマニョーロで多発する若齢発症の良性てんかんで、多くは成長とともに自然寛解します。",
        "mechanism": "LGI2 遺伝子の変異により脳内神経回路の興奮抑制バランスが乱れ、焦点性・全般性てんかん発作を引き起こします。",
        "symptoms": "生後 5〜13 週での強直間代発作・焦点発作・精神運動発作。多くは 4 ヶ月齢までに発作が消失し（良性）、神経学的後遺症を残さない例が多い。",
        "inheritance": "常染色体劣性。主にラゴット・ロマニョーロで報告。両親キャリアで 25% 発症。",
        "advice": "発作が重篤または頻繁な場合は抗てんかん薬で対症療法。多くは予後良好だが繁殖前検査で P/P（ホモ陽性）犬の繁殖への使用は避けることを推奨。",
        "severity": "low",
        "references": [
            {"label": "詳細を検索 (BFJE LGI2)", "url": _google_search("Benign Familial Juvenile Epilepsy BFJE 犬 LGI2 Lagotto")},
        ],
    },
]


def _normalize_for_match(text: str) -> str:
    """マッチング用の正規化: lowercase + ハイフン/アンダースコアをスペースに."""
    return re.sub(r"[\-_]", " ", text.lower())


# 疾患をカテゴリへ分類するパターン（タイトルキーワードベース）
# 辞書ページでグルーピング表示に使用
DISEASE_CATEGORIES = [
    ("🦴 骨格・関節系",   ["椎間板", "骨軟骨", "短足", "骨形成不全"]),
    ("🧠 神経・脳系",     ["脳症", "リポフスチン", "運動失調", "脊髄小脳", "多剤耐性", "ガングリオシドーシス", "変性性脊髄症", "筋強直症", "小脳皮質", "感覚性神経", "ジストロフィー", "白質ジストロフィー", "Krabbe", "多発性神経", "発作性失神", "ヒドロキシグルタル", "髄膜脳炎", "ラフォラ", "ナルコレプシー", "神経軸索", "海綿状白質", "てんかん"]),
    ("👁 眼科系",         ["緑内障", "夜盲", "コリーアイ", "白内障", "全色盲", "錐体杆体", "進行性網膜萎縮", "PRA", "多巣性網膜症", "CMR", "錐体ジストロフィー", "スターガルト", "rcd", "水晶体脱臼"]),
    ("💜 心臓系",         ["心筋症", "不整脈"]),
    ("🩸 血液・凝固系",   ["フォン・ヴィレブランド", "ピルビン酸", "第VII", "プレカリクレイン", "血小板", "メトヘモグロビン", "フルクトキナーゼ"]),
    ("🧪 代謝・内分泌系", ["尿酸尿", "コバラミン", "グリコーゲン蓄積", "ムコ多糖", "銅蓄積", "高リン血症", "リン血症"]),
    ("💪 筋・運動系",     ["運動誘発性", "中心核ミオパチー", "肢帯型筋", "重症筋無力症", "骨格異形成"]),
    ("🫘 腎・泌尿器系",   ["シスチン尿", "腎症", "腎嚢腺癌"]),
    ("🧴 皮膚・被毛系",   ["鼻過角化", "魚鱗癬", "希釈性脱毛", "肉球角化"]),
    ("🛡 免疫系",         ["好中球", "重症複合免疫不全", "再発性炎症性肺"]),
    ("🌱 発達・内分泌系", ["下垂体性小人症", "ミュラー管"]),
    ("🫃 消化器系",       ["消化管"]),
]


def get_disease_category(entry: dict) -> str:
    """エントリのタイトルからカテゴリを推定。マッチしない場合は『その他』"""
    title = entry.get("title", "")
    for cat, patterns in DISEASE_CATEGORIES:
        for p in patterns:
            if p in title:
                return cat
    return "📋 その他"


# 重症度ヒューリスティック: KB 本文のキーワードから 3段階に分類
# entry.get("severity") で明示指定があればそれを優先
_SEVERITY_KEYWORDS = {
    "high":   ["予後不良", "致死", "死亡", "生命に関わる", "生命予後", "1〜2 年で死亡", "1〜2年で死亡",
               "重篤", "失明", "完全失明"],
    "medium": ["対症療法のみ", "進行性", "リスク大幅", "重症", "重度", "進行抑制", "発症リスク",
               "歩行困難", "繰り返す感染"],
    "low":    ["通常は無症状", "通常無症状", "QOL 維持可能", "完治はしない", "症状コントロール",
               "症状管理"],
}


def get_disease_severity(entry: dict) -> str:
    """KB エントリから重症度を推定する: 'high' / 'medium' / 'low'.

    entry.get('severity') が明示されていればそれを使う（手動オーバーライド可）。
    未指定なら summary/mechanism/symptoms/advice のテキストから推定。
    """
    explicit = entry.get("severity")
    if explicit in ("high", "medium", "low"):
        return explicit
    text = " ".join([
        entry.get("summary", ""),
        entry.get("mechanism", ""),
        entry.get("symptoms", ""),
        entry.get("advice", ""),
    ])
    for level in ("high", "medium", "low"):
        for kw in _SEVERITY_KEYWORDS[level]:
            if kw in text:
                return level
    return "medium"  # デフォルト


SEVERITY_LABELS = {
    "high":   {"label": "高リスク", "color": "#dc2626", "bg": "#fee2e2", "emoji": "🔴"},
    "medium": {"label": "中リスク", "color": "#c2410c", "bg": "#fef3c7", "emoji": "🟡"},
    "low":    {"label": "低リスク", "color": "#166534", "bg": "#dcfce7", "emoji": "🟢"},
}


# 症状ベースの絞り込みインデックス（理解できるコンセプト）
# ユーザーが『うちの犬は X の症状がある』から関連疾患を探せる。
# 各シンボトムは KB の match パターンの集合にマップ。
SYMPTOM_INDEX = [
    {
        "id": "hindlimb",
        "label": "🦵 後肢の麻痺・歩行困難",
        "match_patterns": [
            "degenerative myelopathy", "chondrodystrophy", "cddy", "ivdd",
            "centronuclear myopathy", "limb-girdle muscular dystrophy", "myasthenia gravis",
            "episodic falling", "exercise-induced collapse",
        ],
    },
    {
        "id": "vision",
        "label": "👁 視覚障害・失明",
        "match_patterns": [
            "progressive rod cone", "prcd", "progressive retinal atrophy",
            "cone-rod dystrophy", "rcd1", "rcd2", "rcd3", "cngb1 pra",
            "achromatopsia", "day blindness", "stargardt",
            "collie eye anomaly", "hereditary cataract", "hsf4",
            "glaucoma", "csnb", "cone degeneration", "multifocal retinopathy",
            "congenital stationary night blindness",
        ],
    },
    {
        "id": "bleeding",
        "label": "🩸 出血傾向・止血困難",
        "match_patterns": [
            "von willebrand", "vwd",
            "factor vii", "factor 7",
            "prekallikrein", "macrothrombocytopenia",
        ],
    },
    {
        "id": "neuro",
        "label": "🧠 痙攣・神経症状",
        "match_patterns": [
            "neonatal encephalopathy", "news",
            "neuronal ceroid lipofuscinosis", "ncl",
            "gangliosidosis", "gm1", "gm2",
            "globoid cell leukodystrophy", "krabbe",
            "lafora", "necrotizing meningoencephalitis", "nme",
            "spinocerebellar ataxia", "cerebellar abiotrophy",
            "late onset ataxia", "neuroaxonal dystrophy",
            "spongiform leukoencephalomyelopathy",
        ],
    },
    {
        "id": "kidney",
        "label": "🫘 多飲多尿・腎機能異常",
        "match_patterns": [
            "hyperuricosuria", "huu",
            "cystinuria",
            "familial nephropathy",
            "x-linked hereditary nephropathy",
            "renal cystadenocarcinoma",
        ],
    },
    {
        "id": "skin",
        "label": "🧴 皮膚異常・脱毛",
        "match_patterns": [
            "hnpk", "hereditary nasal parakeratosis",
            "ichthyosis",
            "coat color dilution alopecia", "cda",
            "footpad hyperkeratosis", "hfh",
        ],
    },
    {
        "id": "skeletal",
        "label": "🦴 骨格・関節異常",
        "match_patterns": [
            "osteochondrodysplasia", "skeletal dysplasia",
            "osteogenesis imperfecta",
            "chondrodysplasia", "cdpa",
        ],
    },
    {
        "id": "metabolic",
        "label": "🧪 代謝・成長異常",
        "match_patterns": [
            "glycogen storage disease",
            "mucopolysaccharidosis",
            "cobalamin malabsorption",
            "copper toxicosis",
            "hyperphosphatemia",
            "pituitary dwarfism",
            "pyruvate kinase",
        ],
    },
    {
        "id": "drug",
        "label": "💉 薬剤過敏症",
        "match_patterns": [
            "multidrug resistance", "mdr1", "abcb1",
            "methemoglobinemia",
        ],
    },
    {
        "id": "immune",
        "label": "🛡 免疫異常・感染症",
        "match_patterns": [
            "trapped neutrophil",
            "severe combined immunodeficiency", "scid",
            "recurrent inflammatory pulmonary",
        ],
    },
]


def filter_by_symptom(entries: list, symptom_id: str) -> list:
    """症状 ID で疾患をフィルタリング。

    各疾患エントリの match パターンと、症状の match_patterns 間で
    部分文字列マッチがあればその疾患を含める。
    """
    sym = next((s for s in SYMPTOM_INDEX if s["id"] == symptom_id), None)
    if not sym:
        return entries
    sym_patterns = [p.lower() for p in sym["match_patterns"]]
    result = []
    for entry in entries:
        entry_patterns = [p.lower() for p in entry.get("match", [])]
        for sp in sym_patterns:
            if any(sp in ep or ep in sp for ep in entry_patterns):
                result.append(entry)
                break
    return result


def group_diseases_by_category(entries: list) -> list:
    """疾患リストをカテゴリ別にグループ化し、定義順で返す。
    Returns: [(category_name, [entries...]), ...]"""
    buckets = {cat: [] for cat, _ in DISEASE_CATEGORIES}
    buckets["📋 その他"] = []
    order = [cat for cat, _ in DISEASE_CATEGORIES] + ["📋 その他"]
    for entry in entries:
        cat = get_disease_category(entry)
        buckets[cat].append(entry)
    return [(cat, buckets[cat]) for cat in order if buckets[cat]]


def get_disease_detail(test_name: str) -> Optional[dict]:
    """疾患名から詳細解説を取得する。"""
    if not test_name:
        return None
    name_norm = _normalize_for_match(test_name)
    for entry in DISEASE_KB:
        for pattern in entry["match"]:
            # 簡易マッチング: substring または '\b...\b' 単語境界
            if pattern.startswith("\\b") and pattern.endswith("\\b"):
                if re.search(pattern, name_norm):
                    return entry
            elif pattern in name_norm:
                return entry
    return None


# ============================================================
# 形質（毛色など）詳細解説 KB
# ============================================================

TRAIT_KB = [
    {
        "match": ["e locus", "mc1r"],
        "title": "E座位 (MC1R) — 黒系色素のスイッチ",
        "summary": "コートに黒/茶色色素（eumelanin）を作るかどうかを決める『マスタースイッチ』です。",
        "mechanism": "MC1R 遺伝子が活性なら黒/茶色素を産生。e/e ホモ（劣性）では完全に活性が失われ、コートは赤/黄/クリーム/ホワイトのみになります（鼻パッドの色素は残る）。",
        "phenotype": "E/E, E/e: コートに黒/茶色素を発現可。K座位・A座位の支配を受ける。\\ne/e: コートはクリーム〜アプリコット〜レッド（KITLG が濃淡を決定）。",
        "inheritance": "常染色体劣性（e/e ホモ接合で発現）。E/e キャリア間の交配から 25% の確率で e/e 子犬が生まれます。",
        "advice": "e/e でも鼻・パッド・アイリムの色は B 座位で決まります（黒 or ブラウン）。",
        "references": [
            {"label": "詳細を検索 (E locus)", "url": _google_search("E locus MC1R 犬 毛色")},
        ],
    },
    {
        "match": ["k locus", "cbd103"],
        "title": "K座位 (CBD103) — ドミナントブラック",
        "summary": "黒一色（ソリッド）かどうかを決める優性遺伝子です。",
        "mechanism": "KB（優性）が 1コピーでもあると A 座位の表現が抑制され、コートは単色になります。ky/ky ではアグーチ模様（A座位）が現れます。kbr はブリンドル。",
        "phenotype": "KB/_ : ソリッド（黒・茶・希釈色 など、E と B で決まる）\\nky/ky : A 座位の模様（セーブル・タンポイント等）\\nkbr/_  : ブリンドル",
        "inheritance": "常染色体優性（KB は 1コピーで優性。ky/ky が劣性ホモ）。優性序列: KB > kbr > ky。",
        "advice": "ソリッドカラーを残したいなら KB を維持。模様を出したいなら ky/ky × ky/ky に交配。",
        "references": [
            {"label": "詳細を検索 (K locus)", "url": _google_search("K locus CBD103 犬 ドミナントブラック")},
        ],
    },
    {
        "match": ["a locus", "agouti", "asip"],
        "title": "A座位 (ASIP) — アグーチ（模様）パターン",
        "summary": "K座位が ky/ky のときに発現する『毛色の模様』を決める座位です。",
        "mechanism": "ay > aw > at > a の優性順位。ay=セーブル、aw=ワイルドセーブル、at=タンポイント、a=リセッシブブラック。",
        "phenotype": "ay/_ : フォーン/セーブル\\naw/_ : ワイルドセーブル\\nat/_ : ブラックタン/トライカラー（ドーベルマン的）\\na/a : リセッシブブラック（単色黒）",
        "inheritance": "複対立遺伝子の優性序列: ay > aw > at > a。K座位が ky/ky の場合にのみ表現型に現れます。",
        "advice": "見た目の模様パターンは A 座位＋K 座位＋E 座位の組み合わせで決まります。",
        "references": [
            {"label": "詳細を検索 (A locus)", "url": _google_search("A locus ASIP 犬 アグーチ")},
        ],
    },
    {
        "match": ["b locus", "tyrp1", "brown"],
        "title": "B座位 (TYRP1) — ブラウン色素",
        "summary": "黒色素を「黒」のまま発現するか「茶色」に変換するかを決めます。",
        "mechanism": "TYRP1 遺伝子の機能が失われると（bb ホモ）、すべての黒色素がブラウンに変換されます。e/e の場合はコートに eumelanin が無いため B はコートに影響せず、鼻・パッド色素のみに影響します。",
        "phenotype": "B/_ : 通常通り黒色素\\nbb : 全ての黒がブラウン（チョコレート/レバー）。ee と組み合わせるとコートはクリーム〜アプリコットだが鼻はブラウン。",
        "inheritance": "常染色体劣性（bb ホモ接合で発現）。B/b キャリア間の交配から 25% の確率でチョコ子犬が生まれます。",
        "advice": "チョコレート色を残したい場合は bb 必須。",
        "references": [
            {"label": "詳細を検索 (B locus)", "url": _google_search("B locus TYRP1 犬 ブラウン チョコレート")},
        ],
    },
    {
        "match": ["d locus", "dilute", "mlph"],
        "title": "D座位 (MLPH) — 希釈遺伝子",
        "summary": "色素の濃度を薄める（希釈する）遺伝子。黒→青/ブルー、茶→ライラック/イザベラ、黄→シャンパンになります。",
        "mechanism": "MLPH 遺伝子の機能不全（dd ホモ）でメラニン顆粒が均一に分布せず、淡い色になります。",
        "phenotype": "D/_ : 通常通り\\ndd : 希釈。Black→Blue, Brown→Lilac/Isabella, Yellow→Champagne",
        "inheritance": "常染色体劣性（dd ホモ接合で希釈発現）。D/d キャリア間の交配から 25% の確率で希釈色子犬が生まれます。",
        "advice": "ワイマラナーの『ねずみ色』、フレンチブルドッグの『ブルー』などは dd によるもの。",
        "references": [
            {"label": "詳細を検索 (D locus)", "url": _google_search("D locus MLPH 犬 dilute ブルー")},
        ],
    },
    {
        "match": ["\\bm locus\\b", "merle"],
        "title": "M座位 (PMEL17) — マールパターン",
        "summary": "コートに不規則な色のまだら（マール）を作る遺伝子。M/M（ダブルマール）は重大な健康リスクあり。",
        "mechanism": "PMEL17 遺伝子の変異により、色素細胞の機能が部分的に失われ、まだら模様になります。M/M は失明・聴覚障害のリスクが高い。",
        "phenotype": "m/m : マールなし\\nM/m : マール表現型\\nM/M : ダブルマール（白割合増・視聴覚障害リスク大）",
        "inheritance": "常染色体優性（不完全優性）。M/m ヘテロでマール発現。M/M ホモは視聴覚障害リスク大。M/m × M/m 交配は 25% が M/M になるため厳禁。",
        "advice": "**M/m × M/m の交配は厳禁**。25% の確率でダブルマール子犬が生まれます。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索 (Merle)", "url": _google_search("Merle locus PMEL17 犬 マール")},
        ],
    },
    {
        "match": ["s locus", "pied", "mitf", "piebald"],
        "title": "S座位 (MITF) — パイド/パーティカラー",
        "summary": "コートに白い部分（白斑）を作るかどうかを決めます。",
        "mechanism": "MITF 遺伝子のプロモーター変異により、色素細胞の分布が制限されコートに白い領域ができます。",
        "phenotype": "S/S : 白斑なし or 最小\\nS/sp : 軽度の白斑\\nsp/sp : パイド/パーティカラー（白の割合が高い）",
        "inheritance": "常染色体（不完全優性）。sp アレルは半劣性: S/sp で軽度白斑、sp/sp でパーティカラー発現。",
        "advice": "パーティプードルなどは sp/sp。S 座位だけでなく Irish spotting 等の他遺伝子も白の表現に関与。",
        "references": [
            {"label": "詳細を検索 (S locus)", "url": _google_search("S locus MITF 犬 パイド ピーバルド")},
        ],
    },
    {
        "match": ["furnishings", "rspo2"],
        "title": "ファーニシング (RSPO2)",
        "summary": "眉毛・髭・飾り毛などの『ふさふさ』を作る遺伝子です。",
        "mechanism": "RSPO2 遺伝子の挿入変異が顔の毛量を増やします。",
        "phenotype": "F/F or F/N: ファーニシングあり（テリア・ドゥードゥル系）\\nN/N: スムースコート",
        "inheritance": "常染色体優性（不完全優性）。F/N ヘテロでも発現。F/F ホモの方がより濃いファーニシングになる場合があります。",
        "advice": "ドゥードゥル系（ラブラドゥードル等）の見た目に大きく影響。",
        "references": [
            {"label": "詳細を検索 (Furnishings)", "url": _google_search("Furnishings RSPO2 犬 ファーニシング")},
        ],
    },
    {
        "match": ["curly coat", "krt71", "curl"],
        "title": "巻き毛遺伝子 (KRT71)",
        "summary": "コートが直毛か巻き毛かを決める遺伝子です。",
        "mechanism": "KRT71 遺伝子の変異がカール毛を形成します。",
        "phenotype": "C/C or C/N: 巻き毛\\nN/N: 直毛",
        "inheritance": "常染色体優性（不完全優性）。C/N ヘテロでも巻き毛が発現。C/C ホモではより強いカールになります。",
        "advice": "プードル・ビションフリーゼ等は C/C ホモ。F 座位（ファーニシング）と組み合わせると様々な毛質に。",
        "references": [
            {"label": "詳細を検索 (Curly)", "url": _google_search("Curly coat KRT71 犬 巻き毛")},
        ],
    },
    # === Orivet パネル準拠の追加形質 (PR #49) ===
    {
        "match": ["l locus", "hair length", "fgf5", "毛長", "long hair"],
        "title": "L座位 (FGF5) — 被毛長",
        "summary": "短毛か長毛かを決める遺伝子です。長毛は劣性。",
        "mechanism": "FGF5 遺伝子の変異により被毛のサイクルが変化。L/L は短毛、l/l は長毛になります。",
        "phenotype": "L/L: 短毛\\nL/l: 短毛（キャリア）\\nl/l: 長毛",
        "inheritance": "常染色体劣性（l/l ホモ接合で長毛）。L/l キャリア間の交配から 25% の確率で長毛子犬が生まれます。",
        "advice": "プードル・ヨークシャーテリア・パピヨン等は l/l ホモ。FGF5 だけでなく KRT71 (Curly)・RSPO2 (Furnishings) との組み合わせで多様な毛質に。",
        "references": [
            {"label": "詳細を検索 (Hair Length)", "url": _google_search("FGF5 犬 被毛長 短毛 長毛")},
        ],
    },
    {
        "match": ["shedding", "mc5r", "shed", "抜け毛"],
        "title": "SD座位 (MC5R) — 抜け毛量",
        "summary": "コートの抜け毛量を決める遺伝子です。多くの犬種で 1〜2 アレルを保有。",
        "mechanism": "MC5R 遺伝子の変異が被毛の脱落サイクルに影響。SD/SD ホモは抜け毛多、N/N は少なめ。",
        "phenotype": "SD/SD: 抜け毛多（重ねシェッディング）\\nSD/N: 中程度\\nN/N: 抜け毛少なめ",
        "inheritance": "常染色体（共優性/相加的）。SD アレル数に比例して抜け毛量が増えます。",
        "advice": "プードル・ドゥードゥル系は N/N で抜け毛が少なく『ハイポアレジェニック』とされる場合あり。100% 無毛ではない点に注意。",
        "references": [
            {"label": "詳細を検索 (Shedding)", "url": _google_search("MC5R 犬 抜け毛 シェッディング")},
        ],
    },
    {
        "match": ["bob tail", "natural bob tail", "brachyury", "t-gene", "短尾"],
        "title": "BT座位 (Brachyury / T 遺伝子) — 自然短尾",
        "summary": "生まれつき尻尾が短い形質を決める遺伝子です。ホモ (BT/BT) は致死。",
        "mechanism": "Brachyury (T) 遺伝子の変異により尻尾の発達が短くなります。ホモ接合は胚致死。",
        "phenotype": "BT/BT: 胚致死（生まれない）\\nBT/N: 自然短尾\\nN/N: 通常の尾長",
        "inheritance": "常染色体優性（致死ホモ）。BT/N ヘテロで短尾発現。**BT/BT ホモは胚致死** のため、BT/N × BT/N 交配は 25% の胎児が発育しません。",
        "advice": "**BT/BT 同士の交配は厳禁** — 受胎しても胚致死で出生しません。コーギー・ボブテイル・ボクサー等で頻発。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索 (Bob Tail)", "url": _google_search("Brachyury 犬 自然短尾 BT")},
        ],
    },
    {
        "match": ["em locus", "melanistic mask", "メラニスティックマスク", "黒マスク"],
        "title": "Em座位 (MC1R) — メラニスティックマスク",
        "summary": "顔に黒いマスク模様を形成する遺伝子。MC1R の特定バリアント。",
        "mechanism": "E座位 (MC1R) の Em バリアントが優性で、顔面に黒い色素を集中させます。E座位本体とは別の変異。",
        "phenotype": "Em/Em or Em/E: 黒マスクあり\\nE/E (Em なし): マスクなし",
        "inheritance": "常染色体優性（E座位の Em アレルが優性）。Em/E または Em/Em のいずれでも発現。e/e 犬（eumelanin なし）では発現しません。",
        "advice": "ジャーマンシェパード・パグ・ボクサー等で典型。ee 犬では発現しません（eumelanin がコートに無いため）。",
        "references": [
            {"label": "詳細を検索 (Mask)", "url": _google_search("Em MC1R 犬 メラニスティックマスク 黒マスク")},
        ],
    },
    {
        "match": ["g locus", "greying", "progressive greying", "退色", "シルバー"],
        "title": "G座位 (Greying / PMEL17) — 進行性退色",
        "summary": "子犬期は色付きで生まれ、成犬になると退色する遺伝子。プードルのシルバー・ベージュ色の原因。",
        "mechanism": "G_ は成犬期に毛色のメラニンを段階的に失わせます。1〜2歳までに退色が進行。",
        "phenotype": "g/g: 退色なし\\nG/g or G/G: 成犬で退色\\n  Black + G_  → Silver\\n  Brown + G_  → Silver Beige\\n  Blue + G_   → Silver（淡）",
        "inheritance": "常染色体優性。G/g ヘテロでも退色が発現します。G/G ホモとの表現型差は軽微です。",
        "advice": "シルバープードルは生まれた時は黒。1〜2歳までに段階的に退色していきます。M座位 (Merle) とは別遺伝子（同じ PMEL17 でも変異位置が異なる）。",
        "references": [
            {"label": "詳細を検索 (Greying)", "url": _google_search("Greying 犬 シルバー 退色 PMEL17")},
        ],
    },
    # === 犬種特異形質遺伝子（プードル以外） ===
    {
        "match": ["ridge", "ridgeback", "dorsal ridge", "fgf3 fgf4 fgf19", "リッジ", "リッジバック"],
        "title": "Ridge 座位 (FGF3-FGF4-FGF19 重複) — 背筋リッジ",
        "summary": "ローデシアン・リッジバック特有の、背中に逆向きの毛流が生じる形質。FGF3/FGF4/FGF19 領域の重複が原因。",
        "mechanism": "16番染色体の FGF3-FGF4-FGF19 領域に約 133kb の重複が発生し、背中の毛流が逆方向に成長します。",
        "phenotype": "R/R: リッジあり（ただしダーモイドサイナス [dermoid sinus] のリスク上昇）\\nR/r: リッジあり\\nr/r: リッジなし（FCI スタンダード違反）",
        "inheritance": "常染色体優性。R/r ヘテロでリッジ発現。**R/R ホモはダーモイドサイナスリスク上昇** のため、R/r × r/r 交配が推奨されます。",
        "advice": "R/R ホモ接合では**ダーモイドサイナス**（神経管閉鎖不全による皮下嚢胞）の発症率が上がります。R/r ヘテロ × r/r で繁殖し、リッジを保持しつつホモ接合を避けるのが推奨。タイランドリッジバック、フーピアン (Phu Quoc) リッジバックでも同じ変異。",
        "references": [
            {"label": "詳細を検索 (Ridge)", "url": _google_search("FGF4 ridgeback dermoid sinus")},
        ],
    },
    {
        "match": ["hairless", "foxi3", "ヘアレス", "無毛", "中国冠毛犬", "シャイクレ"],
        "title": "Hairless 座位 (FOXI3) — 無毛形質",
        "summary": "チャイニーズ・クレステッド、メキシカン・ヘアレス（ショロイツクィントレ）、ペルービアン・ヘアレスの無毛を決める遺伝子。ホモは胚致死。",
        "mechanism": "FOXI3 遺伝子の挿入変異が優性。歯や毛包の発達に関与し、変異により被毛・歯の生成が抑制されます。H/H ホモ接合は胚段階で致死。",
        "phenotype": "H/H: 胚致死（生まれない）\\nH/h: 無毛 / 部分無毛・歯の欠損あり\\nh/h: 完全な被毛 (powderpuff / coated)",
        "inheritance": "常染色体優性（致死ホモ）。H/h ヘテロで無毛発現。**H/H ホモは胚致死**。H/h × H/h 交配は 25% の胎児が発育しません。",
        "advice": "**H/H 同士の交配は厳禁** — 全頭出生しない。チャイニーズ・クレステッドでは無毛 (H/h) × パウダーパフ (h/h) 交配が標準。皮膚保護・日焼け止め・防寒が必要な犬種。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索 (Hairless)", "url": _google_search("FOXI3 hairless dog Chinese Crested Xolo")},
        ],
    },
    {
        "match": ["i locus", "intense red", "mfsd12", "intensity", "レッド濃度", "イエロー濃度"],
        "title": "I 座位 (MFSD12) — 赤/黄色の濃度",
        "summary": "E座位 e/e のフェオメラニン（赤・黄系）の濃さを決める遺伝子。ラブラドール・プードル・ゴールデンの色の濃淡を説明。",
        "mechanism": "MFSD12 遺伝子の変異がメラノサイトでのフェオメラニン生成量を調節。i/i ホモで色素が薄くなりクリーム/ホワイト寄りになります。",
        "phenotype": "I/I: 濃い赤・ディープレッド\\nI/i: 中間（アプリコット〜レッド）\\ni/i: 淡いクリーム・ホワイト寄り",
        "inheritance": "常染色体（不完全優性/相加的）。I/I が最濃色、I/i が中間、i/i が最淡色。",
        "advice": "イングリッシュ・クリーム・ゴールデンの淡色や、プードルのホワイト〜クリーム〜アプリコット〜レッドの幅広さの一因。E座位 ee と組み合わせて色の濃淡を予測する際に重要。E_ 犬（黒系優位）では coat 色への影響は限定的。",
        "references": [
            {"label": "詳細を検索 (Intense Red)", "url": _google_search("MFSD12 dog intense red phaeomelanin")},
        ],
    },
    {
        "match": ["cocoa", "hps3", "non-standard chocolate", "テスティング・チョコ", "ココア"],
        "title": "Cocoa 座位 (HPS3) — 非標準チョコレート",
        "summary": "フレンチブルドッグで報告された B座位 (TYRP1) とは別の劣性チョコレート色。",
        "mechanism": "HPS3 遺伝子の変異が劣性で発現。BB（または Bb）でも co/co ホモであればチョコレート色になります。B座位 bb の従来チョコとは独立した遺伝子。",
        "phenotype": "Co/Co or Co/co: 通常色\\nco/co: チョコレート（B 座位とは独立に発現）",
        "inheritance": "常染色体劣性（co/co ホモ接合で発現）。Co/co キャリア間の交配から 25% の確率で Cocoa チョコレートが生まれます。",
        "advice": "フレンチブルドッグ・オーストラリアンシェパード等で報告。B座位だけ検査して 'チョコにならないはず' の交配でも、co/co × co/co で予想外のチョコが出るため、希少色のブリーダーは両方検査推奨。CDA リスクは未確立。",
        "references": [
            {"label": "詳細を検索 (Cocoa)", "url": _google_search("HPS3 cocoa French Bulldog chocolate")},
        ],
    },
    {
        "match": ["harlequin", "h locus", "psmb7", "ハーレクイン"],
        "title": "H座位 (PSMB7) — ハーレクイン",
        "summary": "グレートデーン特有の白地に黒斑のハーレクイン模様を決める遺伝子。マールと相互作用。ホモ致死。",
        "mechanism": "PSMB7 遺伝子の変異が優性。M座位 (Merle) のメルル領域の色素を完全に脱色させ、白地に黒斑のパターンを作ります。M_ + H_ の組み合わせでのみ発現。H/H は胚致死。",
        "phenotype": "H/H: 胚致死\\nH/h + M/m: ハーレクイン（白地に黒斑）\\nH/h + m/m: 表現型に変化なし（ハーレクイン非発現の保因）\\nh/h: ハーレクイン非発現",
        "inheritance": "常染色体優性（致死ホモ）。H/h + M/m の組み合わせでのみハーレクイン表現。**H/H ホモは胚致死**。M座位 (Merle) との同時検査が必須。",
        "advice": "**H/H × H/H 交配は禁忌** — 全頭胚致死。さらに M/M（ダブルマール）× ハーレクインの組み合わせは深刻な発達異常リスク。グレートデーンのハーレクインブリーダーは PSMB7 と PMEL17 (M) の両方を検査する必要があります。",
        "severity": "high",
        "references": [
            {"label": "詳細を検索 (Harlequin)", "url": _google_search("PSMB7 harlequin Great Dane merle")},
        ],
    },
    {
        "match": ["roan", "r locus", "usher", "ローン"],
        "title": "Roan 座位 — ローン（混色斑点）",
        "summary": "成犬になると白い部分に色付き斑点が散らばる『ローン』形質。スパニエル系・オーストラリアン・キャトルドッグ等で典型。",
        "mechanism": "原因遺伝子は未確定（USH2A 周辺が候補）。優性遺伝で、白斑領域に色素細胞が遅れて移行することで成犬期にティッキング模様が出現します。",
        "phenotype": "Rn/Rn or Rn/rn: ローンあり（成犬で発現）\\nrn/rn: ローンなし（白斑のまま）",
        "inheritance": "常染色体優性（原因遺伝子未確定）。Rn/rn ヘテロでも発現。S座位 (パイド/白斑) の存在が前提です。",
        "advice": "イングリッシュ・コッカー・スパニエル、ジャーマン・ショートヘアード・ポインター、オーストラリアン・キャトルドッグの『ブルーヒーラー』『レッドヒーラー』はこのアレル発現。子犬期は白いが、4〜6 週齢から斑点が出始めます。",
        "references": [
            {"label": "詳細を検索 (Roan)", "url": _google_search("dog roan USH2A coat pattern")},
        ],
    },
    {
        "match": ["ticking", "t locus", "ティッキング", "ダルメシアン斑点"],
        "title": "T座位 — ティッキング（斑点）",
        "summary": "白地に小さな色付き斑点が散らばる形質。ダルメシアンの斑点もこの遺伝子による。",
        "mechanism": "優性遺伝。原因遺伝子は未確定だが、S座位 sp/sp の白斑領域に色素細胞が遅れて移行することで斑点が出現します。Roan より斑点が小さく密度が低い。",
        "phenotype": "T/T or T/t: ティッキングあり（白地に小斑点）\\nt/t: 純白の白斑のまま",
        "inheritance": "常染色体優性（原因遺伝子未確定）。T/t ヘテロでも発現。S座位の白斑領域に対してのみ作用します。",
        "advice": "ダルメシアン・イングリッシュ・セッター・ジャック・ラッセル・テリア等で典型。ダルメシアンでは追加で SLC2A9（高尿酸尿症 HUU）のリスクアレルが固定しており、HUU 検査と併用が標準。",
        "references": [
            {"label": "詳細を検索 (Ticking)", "url": _google_search("dog ticking T locus Dalmatian")},
        ],
    },
    {
        "match": ["alx4", "blue eyes", "blue eye duplication", "青目", "ブルーアイ"],
        "title": "ALX4 — ブルーアイ（青目）",
        "summary": "シベリアン・ハスキー特有の青目を決める遺伝子重複。マールやメルル以外の青目原因。",
        "mechanism": "18番染色体の ALX4 近傍に約 98kb の重複が発生し、優性遺伝で虹彩の色素生成を抑制。M座位 (Merle) や S座位 (Pied) 経由の青目とは独立の機構です。",
        "phenotype": "BE/BE or BE/be: 青目（片目または両目）\\nbe/be: 通常色（茶〜琥珀）",
        "inheritance": "常染色体優性（遺伝子重複）。BE/be ヘテロでも青目または片目が発現します。",
        "advice": "シベリアン・ハスキーの青目・オッドアイの主因。マールや白斑経由の青目と異なり、視聴覚障害リスクとの関連は報告されていません。オーストラリアンシェパードのマール経由青目とは区別が必要。",
        "references": [
            {"label": "詳細を検索 (ALX4 Blue Eyes)", "url": _google_search("ALX4 blue eyes Siberian Husky")},
        ],
    },
    # === 毛色濃度・パターン補助遺伝子 ===
    {
        "match": ["kitlg", "fox red", "intense yellow", "フォックスレッド", "ディープレッド"],
        "title": "KITLG — フォックスレッド / 黄色濃度",
        "summary": "黄ラブラドールの『フォックスレッド』など、フェオメラニン（黄/赤系）の発色濃度を決める補助遺伝子。",
        "mechanism": "KITLG 遺伝子近傍の調節領域変異がメラノサイトの分化と色素生成を調節。e/e (E座位) 犬で coat の黄色濃度に強く影響します。",
        "phenotype": "Intense allele 2 コピー: フォックスレッド〜ディープレッド\\nIntense allele 1 コピー: 中間色（中程度のイエロー）\\n通常型 2 コピー: 淡いイエロー〜クリーム",
        "inheritance": "常染色体（相加的）。Intense アレル数に比例して赤色濃度が上がります。",
        "advice": "ラブラドール、ノバスコシア・ダックトーリング・レトリーバー、レッド系ゴールデンで重要。I座位 (MFSD12) と組み合わせて使うと黄〜赤色の濃淡をより正確に予測できます。健康影響なし。",
        "references": [
            {"label": "詳細を検索 (KITLG)", "url": _google_search("KITLG fox red Labrador phaeomelanin")},
        ],
    },
    {
        "match": ["saddle tan", "raly", "saddle pattern", "サドルタン", "サドル模様"],
        "title": "Saddle Tan (RALY) — サドル模様",
        "summary": "生まれた時はブラックタン（タンポイント）だが、成長と共に背中の黒色が後退し『サドル』模様になる遺伝子。ジャーマンシェパードの典型パターン。",
        "mechanism": "RALY 遺伝子の調節領域変異が優性。A座位 at/at の犬で、成長に伴い背中の eumelanin 領域が縮小して褐色（タン）が広がります。",
        "phenotype": "I/I or I/i: 成犬でサドルタン（背中のみ黒、四肢・顔は褐色）\\ni/i: ブラックタン（黒主体のまま）",
        "inheritance": "常染色体優性（不完全優性）。I/i ヘテロでもサドルタンが発現。A座位 at/at であることが前提条件。",
        "advice": "ジャーマンシェパード、エアデール・テリア、ビーグル、ベルジアン・タービュレン等で典型。子犬期はブラックタンに見えても、6ヶ月〜2歳でサドル模様に変化します。A座位 at/at が必須前提。",
        "references": [
            {"label": "詳細を検索 (Saddle Tan)", "url": _google_search("RALY saddle tan dog German Shepherd")},
        ],
    },
    {
        "match": ["domino", "grizzle", "ea allele", "mc1r ea", "ドミノ", "グリズル"],
        "title": "Domino / Grizzle (MC1R Ea) — 顔マスクの色抜け",
        "summary": "サイトハウンド系特有の『顔のマスク領域の色が抜けて薄く見える』パターン。アフガン、サルーキ、ボルゾイ、シベリアンハスキーで報告。",
        "mechanism": "MC1R の Ea (Edomino) バリアントが eumelanin 産生をマスク領域で部分抑制。E座位の派生アレルで、E > Eg > Ea > e の優位序列。",
        "phenotype": "Ea/Ea or Ea/e: ドミノ/グリズル（顔の黒抜け、淡色マスク）\\nEa/E_ (E 優性): 通常",
        "inheritance": "E座位の複対立遺伝子。優性序列: E > Eg > Ea > e。Ea/e または Ea/Ea のみで発現（E/_ または Em/_ が存在すると抑制）。",
        "advice": "アフガンハウンドの『ドミノ』、サルーキの『グリズル』、シベリアンハスキーの『アグーチ』様パターンの一因。Em マスク（黒マスク）とは反対方向の作用。E座位の検査で Ea バリアントを含めない検査もあるため、サイトハウンドブリーダーは検査内容を確認推奨。",
        "references": [
            {"label": "詳細を検索 (Domino)", "url": _google_search("MC1R domino grizzle Afghan Saluki Ea")},
        ],
    },
    {
        "match": ["improper coat", "ic locus", "smooth doodle", "incorrect coat", "インプロパーコート"],
        "title": "Improper Coat (IC / RSPO2 variant)",
        "summary": "ポーチュギーズ・ウォーター・ドッグや Doodle 系で出現する『スムースな顔（ファーニシングなし）+ 短毛』の劣性表現型。",
        "mechanism": "RSPO2 のファーニシング変異の保有なし（ic/ic ホモ）が原因。F 座位の f/f に相当し、顔毛・眉毛・髭の発達が抑制されます。",
        "phenotype": "IC/IC or IC/ic: 標準的なファーニッシュコート（眉・髭あり）\\nic/ic: スムースコート（顔毛なし、ラブ/ゴールデン的な外観）",
        "inheritance": "常染色体劣性（ic/ic ホモ接合でスムースコート発現）。IC/ic キャリア間の交配から 25% の確率でインプロパーコートが生まれます。",
        "advice": "ポーチュギーズ・ウォーター・ドッグ、ラゴット・ロマニョーロでは血統書失格、Doodle 系（ラブラドゥードル/ゴールデンドゥードル）では『hypoallergenic 失敗』と扱われます。F座位 (RSPO2) ファーニシング検査と同時に確認推奨。健康影響なし。",
        "references": [
            {"label": "詳細を検索 (Improper Coat)", "url": _google_search("improper coat IC RSPO2 Portuguese Water Dog Doodle")},
        ],
    },
    {
        "match": ["albinism", "oca", "tyr", "slc45a2", "oculocutaneous albinism", "アルビニズム", "白皮症"],
        "title": "Albinism (TYR / SLC45A2) — 眼皮膚白皮症",
        "summary": "メラニン生成が極度に低下する遺伝性白皮症。被毛・皮膚が白く、虹彩が淡色〜赤目、視覚障害を伴うことが多い。",
        "mechanism": "TYR (チロシナーゼ) や SLC45A2 などのメラニン合成経路遺伝子の劣性変異により、メラニン生成が極端に低下。優性ホワイト（マールやパイドの白）とは区別される真のアルビニズム。",
        "phenotype": "A/A or A/oca: 通常色\\noca/oca: アルビノ（被毛・皮膚白、赤目〜淡色虹彩、羞明・視覚低下）",
        "inheritance": "常染色体劣性（oca/oca ホモ接合で発現）。A/oca キャリア間の交配から 25% の確率でアルビノ子犬が生まれます。",
        "advice": "ドーベルマン (Z-factor)・ペキニーズ・ラサ・アプソ等で報告。⚠️ **健康影響あり**: 紫外線過敏（皮膚がん発症率上昇）・視覚障害・羞明。アルビノ犬の繁殖は倫理的に推奨されません。屋外活動時の遮光保護と定期皮膚検診が必須。",
        "references": [
            {"label": "詳細を検索 (Albinism)", "url": _google_search("dog albinism TYR SLC45A2 Doberman Z-factor")},
        ],
    },
]


def _slugify(text: str) -> str:
    """URL-safe スラッグに変換。

    例: 'Chondrodystrophy with IVDD' → 'chondrodystrophy-with-ivdd'
    """
    text = text.lower()
    # 区切り記号 → 空白（プラス・スラッシュ・括弧・アンダースコア等）
    text = re.sub(r"[+_/&,()]", " ", text)
    # 全角空白・non-breaking space → 半角空白
    text = re.sub(r"[ 　]", " ", text)
    # アルファベット・数字・空白・ハイフン以外を削除
    text = re.sub(r"[^a-z0-9\s\-]", "", text)
    # 連続空白・ハイフン → 単一ハイフン
    text = re.sub(r"[\s\-]+", "-", text.strip())
    return text.strip("-")


def make_entry_slug(entry: dict) -> str:
    """KB エントリから URL スラッグを生成。

    最初の ASCII-friendly な match パターンを優先。
    日本語のみの場合は title からフォールバック。
    """
    for pat in entry.get("match", []):
        clean = pat.replace("\\b", "").strip()
        # ASCII 主体のパターンを採用
        if re.match(r"^[a-zA-Z0-9\s\-_]+$", clean):
            slug = _slugify(clean)
            if slug:
                return slug
    # フォールバック: title から生成
    title = entry.get("title", "")
    # 括弧内の英略を優先抽出 (例: "CDDY+IVDD" 等)
    m = re.search(r"\(([A-Za-z0-9\s\-+/]+)\)", title)
    if m:
        return _slugify(m.group(1))
    return _slugify(title) or "entry"


def _build_slug_index(entries: list) -> dict:
    """slug → entry の辞書を生成。重複時はサフィックスで一意化。"""
    index = {}
    for e in entries:
        slug = make_entry_slug(e)
        base = slug
        i = 2
        while slug in index:
            slug = f"{base}-{i}"
            i += 1
        e["_slug"] = slug  # エントリ自身に slug をキャッシュ
        index[slug] = e
    return index


# モジュールロード時に slug インデックスを生成
DISEASE_SLUG_INDEX = _build_slug_index(DISEASE_KB)
TRAIT_SLUG_INDEX = _build_slug_index(TRAIT_KB)


# ============================================================
# 英訳オーバーレイ (kb_en.py から疾患・形質エントリへマージ)
# ============================================================
# AI 自動翻訳。獣医監修要。lang=en で参照可能。

try:
    from kb_en import (
        DISEASE_EN, TRAIT_EN,
        SEVERITY_LABELS_EN, CATEGORY_LABELS_EN, SYMPTOM_LABELS_EN,
    )
    # 既存エントリに _en フィールドとしてマージ
    for slug, en_data in DISEASE_EN.items():
        if slug in DISEASE_SLUG_INDEX:
            DISEASE_SLUG_INDEX[slug]["_en"] = en_data
    for slug, en_data in TRAIT_EN.items():
        if slug in TRAIT_SLUG_INDEX:
            TRAIT_SLUG_INDEX[slug]["_en"] = en_data
    HAS_EN_KB = True
except ImportError:
    HAS_EN_KB = False
    SEVERITY_LABELS_EN = {}
    CATEGORY_LABELS_EN = {}
    SYMPTOM_LABELS_EN = {}


def get_entry_field(entry: dict, field: str, lang: str = "ja") -> str:
    """エントリから指定フィールドの値を言語に応じて取得。

    lang='en' なら _en[field] を試み、無ければ日本語フィールドにフォールバック。
    """
    if lang == "en" and entry and "_en" in entry:
        en_val = entry["_en"].get(field)
        if en_val:
            return en_val
    return entry.get(field, "") if entry else ""


def get_disease_kb_localized(lang: str = "ja") -> list:
    """言語に応じた疾患 KB を返す（_en があれば優先、無ければ JA）"""
    if lang != "en":
        return DISEASE_KB
    result = []
    for entry in DISEASE_KB:
        if "_en" in entry:
            merged = {**entry, **entry["_en"]}
            # match と severity 等 EN にないものは JA から保持
            merged["match"] = entry["match"]
            merged["_slug"] = entry.get("_slug")
            if "severity" in entry:
                merged["severity"] = entry["severity"]
            merged["references"] = entry.get("references", [])
            result.append(merged)
        else:
            # 英訳なし: そのまま日本語版を返す（注記をタイトルに追加）
            result.append(entry)
    return result


def get_trait_kb_localized(lang: str = "ja") -> list:
    """言語に応じた形質 KB を返す"""
    if lang != "en":
        return TRAIT_KB
    result = []
    for entry in TRAIT_KB:
        if "_en" in entry:
            merged = {**entry, **entry["_en"]}
            merged["match"] = entry["match"]
            merged["_slug"] = entry.get("_slug")
            merged["references"] = entry.get("references", [])
            result.append(merged)
        else:
            result.append(entry)
    return result


# ============================================================
# ガイド記事フレームワーク（マーケティング・SEO 流入用）
# ============================================================
# 各ガイド: slug, title, summary, sections, related_disease_slugs, related_trait_slugs
# sections は heading + body_html のリスト（簡易構造）
# 監修者・公開日・更新日は Orivet 側で正式化予定

GUIDES = [
    {
        "slug": "how-to-read-orivet-results",
        "title": "Orivet 遺伝子検査結果の読み方ガイド",
        "summary": "Orivet 遺伝子検査PDFを受け取ったときに、結果の見方・用語の意味・次のアクションを分かりやすく解説します。",
        "category": "🔰 初心者向け",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "📄 検査結果 PDF に何が書かれているか",
                "body": (
                    "Orivet 遺伝子検査 PDF には大きく分けて『健康疾患』と『形質（毛色等）』の2種類の結果が記載されています。"
                    "各項目に対して『N/N（ノーマル）』『P/N（キャリア）』『P/P（ポジティブ）』のいずれかが示されます。"
                ),
            },
            {
                "heading": "🟢 N/N — ノーマル（陰性）",
                "body": (
                    "両親から受け継いだ遺伝子の両方が正常な状態です。"
                    "その疾患を発症する遺伝的リスクはなく、子犬にもキャリア遺伝子を渡しません。"
                ),
            },
            {
                "heading": "🟡 P/N — キャリア（保因犬）",
                "body": (
                    "片方の親から変異遺伝子を受け継いだヘテロ接合体。"
                    "ほとんどの疾患（常染色体劣性遺伝の場合）では**発症しません**が、子犬に変異遺伝子を50%の確率で渡します。"
                    "繁殖時は P/P または P/N 相手を避けることで、発症犬の出生を防げます。"
                ),
            },
            {
                "heading": "🔴 P/P — ポジティブ（発症リスクあり）",
                "body": (
                    "両親から変異遺伝子を受け継いだホモ接合体。"
                    "常染色体劣性疾患では**発症します**。発症時期や症状の重さは疾患により異なります。"
                    "獣医師による定期的な健康チェックを推奨します。"
                ),
            },
            {
                "heading": "💡 次にすべきこと",
                "body": (
                    "1. ポジティブ (P/P) の項目があれば、まず**獣医師に相談**してください。発症前に対策できる疾患も多くあります。<br>"
                    "2. キャリア (P/N) の場合は**繁殖計画**を慎重に。同じ変異を持つ犬同士の交配は避けるべきです。<br>"
                    "3. 結果をシェアしたい場合は、Excel ダウンロード機能でドキュメント化できます。<br>"
                    "4. **辞書ページ**で各疾患の詳細解説・遺伝様式・アドバイスを確認できます。"
                ),
            },
        ],
        "related_disease_slugs": ["chondrodystrophy", "degenerative-myelopathy", "progressive-rod-cone"],
        "related_trait_slugs": [],
    },
    {
        "slug": "coi-basics",
        "title": "COI（近親交配係数）入門 — 数字の意味を理解する",
        "summary": "COI（Coefficient of Inbreeding）が何を意味するか、どこから危険な水準か、人間関係に置き換えた直感的な解説。",
        "category": "🐕 繁殖計画",
        "reading_time": "4 分",
        "sections": [
            {
                "heading": "📊 COI とは",
                "body": (
                    "COI（Coefficient of Inbreeding）は、父親と母親が共通祖先を持つ場合、その子犬が両親から同じ遺伝子を2コピー受け継ぐ確率を示す指標です。"
                    "1922年に Sewall Wright が確立した古典的な指標で、犬の繁殖判断の基本となっています。"
                ),
            },
            {
                "heading": "🎚 段階の目安（人間関係換算）",
                "body": (
                    "<table style='width:100%;border-collapse:collapse;'>"
                    "<tr><th style='text-align:left;padding:6px 10px;background:#f3f4f6;'>COI</th><th style='text-align:left;padding:6px 10px;background:#f3f4f6;'>人間関係換算</th></tr>"
                    "<tr><td style='padding:6px 10px;color:#22c55e;'><strong>0%</strong></td><td>完全に無関係な両親同士</td></tr>"
                    "<tr><td style='padding:6px 10px;color:#22c55e;'><strong>〜6.25%</strong></td><td>いとこ婚相当 — 一般的に許容範囲</td></tr>"
                    "<tr><td style='padding:6px 10px;color:#eab308;'><strong>6.25〜12.5%</strong></td><td>半兄妹婚相当 — 免疫力・繁殖力低下傾向</td></tr>"
                    "<tr><td style='padding:6px 10px;color:#ef4444;'><strong>12.5〜25%</strong></td><td>兄妹婚・親子婚相当 — 遺伝性疾患リスク大幅増加</td></tr>"
                    "<tr><td style='padding:6px 10px;color:#dc2626;'><strong>25%超</strong></td><td>近親婚の繰り返し — 劣性疾患の発症率が指数関数的に上昇</td></tr>"
                    "</table>"
                ),
            },
            {
                "heading": "⚠️ 高 COI 犬の健康リスク",
                "body": (
                    "高 COI 犬では、隠れていた劣性遺伝病が顕在化しやすくなります。また免疫機能・繁殖能力・寿命の低下も報告されています。"
                    "JKC・FCI 等の育種団体は『COI 6.25% 以下』を推奨するケースが多く、特に種牡犬の選定時には重要な指標です。"
                ),
            },
            {
                "heading": "🔧 ツールで COI を算出する",
                "body": (
                    "本サービスの**繁殖シミュレーター**では、3世代の血統情報を入力するだけで Wright の方法による COI を自動算出できます。"
                    "共通祖先がどこに何回出現しているかも可視化されるため、繁殖計画の意思決定に直接活用できます。"
                ),
            },
            {
                "heading": "🔬 血統ベース COI とヘテロ接合率（ゲノム多様性）の違い",
                "body": (
                    "『COI』と名のつく数値には、実は<strong>大きく分けて 2 種類</strong>あり、測定しているものが異なります。<br><br>"
                    "<strong>① 血統ベース COI（本ツールの算出方法）</strong><br>"
                    "血統書から共通祖先を辿り、子犬が同じ遺伝子を 2 コピー受け継ぐ確率を<strong>予測</strong>します。"
                    "交配『前』に血統書だけで計算できるのが最大の利点。"
                    "ただし血統書の世代数・精度に依存し、記載のない祖先の重複は反映されません。"
                    "JKC・FCI 等の伝統的な繁殖指針はこちらを前提にしています。<br><br>"
                    "<strong>② ヘテロ接合率（Orivet などの DNA 検査）</strong><br>"
                    "数万箇所の SNP（一塩基多型）を<strong>実測</strong>し、ゲノム全体でヘテロ接合（両親由来の遺伝子が異なる）の割合を求めます。"
                    "血統書の誤りや記載漏れに左右されず、実際のゲノム状態を直接反映するのが強み。"
                    "ただし DNA 検査が必要で、交配『前』の予測には使えません。<br><br>"
                    "⚠️ <strong>本ツールの COI 値と、Orivet の検査結果（ヘテロ接合率）の数値は一致しません。</strong>"
                    "これは誤りではなく、『予測 vs 実測』『血統 vs ゲノム』という<strong>別々の指標</strong>だからです。"
                    "両者は競合ではなく補完関係にあり、繁殖判断では『血統由来のリスク予測（COI）』と『実測のゲノム多様性（ヘテロ接合率）』を併用するのが理想的です。"
                ),
            },
        ],
        "related_disease_slugs": [],
        "related_trait_slugs": [],
    },
    {
        "slug": "color-genetics-basics",
        "title": "犬の毛色遺伝子の基本 — 8座位の役割",
        "summary": "犬の毛色は8つの主要な遺伝子座（E/K/A/B/D/M/S/G）の組み合わせで決まります。各座位の役割を簡潔に解説。",
        "category": "🎨 毛色遺伝学",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🎨 毛色の決まり方の階層構造",
                "body": (
                    "犬の毛色は単一の遺伝子ではなく、複数の座位の組み合わせで決まります。"
                    "まず E座位が『黒系色素を作れるか』を決め、次に K座位が『単色か模様か』を決定。"
                    "B 座位が『黒 or 茶』、D 座位が『希釈の有無』、M/S は模様、G は退色を司ります。"
                ),
            },
            {
                "heading": "🔌 E座位（MC1R）— 色素のマスタースイッチ",
                "body": (
                    "E_ なら黒/茶色素を coat に発現可能。e/e ホモでは coat はクリーム〜アプリコット〜レッドのみ。"
                    "ただし e/e でも鼻・パッド・アイリムには色素が出るため B 座位の影響を受けます。"
                ),
            },
            {
                "heading": "🎯 K座位（CBD103）— ドミナントブラック",
                "body": (
                    "KB_ は単色（ソリッド）になり、A座位の発現を抑制。"
                    "ky/ky では A 座位の模様（セーブル・タンポイント等）が現れます。"
                    "kbr_ はブリンドル。"
                ),
            },
            {
                "heading": "🎭 A座位（ASIP）— アグーチ模様",
                "body": (
                    "K = ky/ky のときに発現。優性順位 ay > aw > at > a。"
                    "ay = フォーン/セーブル、aw = ワイルドセーブル、at = ブラックタン/トライカラー、a/a = リセッシブブラック。"
                ),
            },
            {
                "heading": "🍫 B座位（TYRP1）— ブラウン色素",
                "body": (
                    "B_ なら通常の黒色素。bb で全ての黒色素がブラウン（チョコレート/レバー）に。"
                    "ee 犬では B はコート色に影響せず、鼻・パッド色素のみに作用します。"
                ),
            },
            {
                "heading": "💧 D座位（MLPH）— 希釈",
                "body": (
                    "dd で色素濃度が希釈：Black → Blue, Brown → Lilac/Isabella, Yellow → Champagne。"
                    "ワイマラナーやフレンチブルドッグのブルーはこれが原因。"
                ),
            },
            {
                "heading": "🎨 M / S / G 座位",
                "body": (
                    "**M座位 (PMEL17)**: マール模様。M/M はダブルマールで視聴覚障害リスク大。<br>"
                    "**S座位 (MITF)**: パイド/パーティカラー。<br>"
                    "**G座位 (Greying)**: シルバープードルのような『成犬で退色する』形質。"
                ),
            },
        ],
        "related_disease_slugs": [],
        "related_trait_slugs": ["e-locus", "k-locus", "a-locus", "b-locus", "d-locus"],
    },
    {
        "slug": "breeders-checklist",
        "title": "ブリーダー向け：繁殖計画チェックリスト",
        "summary": "健康な子犬を生むために、繁殖前に必ず確認すべき項目をチェックリスト形式で整理。",
        "category": "🐕 繁殖計画",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "✅ 繁殖前に必ず確認すべきこと",
                "body": (
                    "□ 両親の遺伝子検査結果がある（少なくとも 12〜14 項目）<br>"
                    "□ 両親が同じ変異の P/P または P/N でないか（劣性疾患の発症犬を避ける）<br>"
                    "□ COI が許容範囲内（理想 6.25% 以下）<br>"
                    "□ M座位（Merle）の場合、M/m × M/m を避ける（ダブルマール禁忌）<br>"
                    "□ BT座位（自然短尾）の場合、BT/BT 同士は胚致死<br>"
                    "□ vWD・MDR1 等のキャリア結果を獣医師と共有<br>"
                    "□ 両親の血統書を 3〜5 世代まで確認"
                ),
            },
            {
                "heading": "⚠️ 繁殖を再考すべきケース",
                "body": (
                    "<strong>1. 両親が同じ高リスク疾患のキャリア</strong>: 25% の確率で発症犬。<br>"
                    "<strong>2. M/m × M/m</strong>: ダブルマールリスク。失明・聴覚障害。<br>"
                    "<strong>3. COI 12.5% 超</strong>: 健康問題増加・繁殖力低下リスク大。<br>"
                    "<strong>4. 親に重度の遺伝性疾患歴</strong>: 子犬への遺伝可能性大。<br>"
                    "<strong>5. 検査結果なしでの繁殖</strong>: リスクを把握できないまま生まれる子犬への責任問題。"
                ),
            },
            {
                "heading": "📋 推奨検査パネル（最低限）",
                "body": (
                    "犬種により推奨検査は異なりますが、以下は多くの犬種で重要:<br>"
                    "・<strong>DM（変性性脊髄症）</strong> — SOD1<br>"
                    "・<strong>CDDY+IVDD</strong> — FGF4 / 椎間板疾患<br>"
                    "・<strong>vWD I/II/III</strong> — 凝固因子<br>"
                    "・<strong>prcd-PRA</strong> — 進行性網膜萎縮<br>"
                    "・<strong>MDR1</strong> — 薬剤過敏症<br>"
                    "犬種固有の疾患（例: プードルの NEwS、ラブラドールの CNM 等）は **辞書ページで犬種別に確認**してください。"
                ),
            },
            {
                "heading": "📊 シミュレーターで事前検証",
                "body": (
                    "繁殖シミュレーターを使えば、両親候補の遺伝子型から子犬の遺伝子型確率を事前算出できます。"
                    "色シミュレーション・健康リスク予測・COI 計算を一括で確認しましょう。"
                ),
            },
        ],
        "related_disease_slugs": ["chondrodystrophy", "degenerative-myelopathy", "progressive-rod-cone"],
        "related_trait_slugs": [],
    },
    {
        "slug": "severity-explained",
        "title": "重症度（🔴🟡🟢）の判定基準について",
        "summary": "辞書ページの重症度バッジが何を基準にしているか、どう活用すべきかを解説。",
        "category": "🔰 初心者向け",
        "reading_time": "3 分",
        "sections": [
            {
                "heading": "🚦 3段階の重症度",
                "body": (
                    "本サービスでは、辞書・レポートの各疾患に重症度バッジを表示しています:<br>"
                    "🔴 <strong>高リスク</strong>: 予後不良・致死性が高い、または生命に関わる疾患<br>"
                    "🟡 <strong>中リスク</strong>: 進行性または対症療法が必要だが、QOL を維持しながら生活可能<br>"
                    "🟢 <strong>低リスク</strong>: 通常無症状または軽微、限定的な注意で済む"
                ),
            },
            {
                "heading": "🤖 判定方法",
                "body": (
                    "判定は2段階で行います:<br>"
                    "1. KB エントリの本文テキストから自動推定（『予後不良』『致死』等のキーワード）<br>"
                    "2. 誤判定が発生したエントリは手動で `severity` フィールドを明示指定（オーバーライド）"
                ),
            },
            {
                "heading": "⚠️ 重要な免責",
                "body": (
                    "重症度は**一般的な傾向**を示すものであり、特定個体の予後を保証するものではありません。"
                    "実際の症状の重さは:<br>"
                    "・犬種<br>"
                    "・遺伝子型（P/N キャリア vs P/P 発症犬）<br>"
                    "・併発疾患<br>"
                    "・環境要因<br>"
                    "により大きく異なります。診断・治療判断は必ず**獣医師にご相談ください**。"
                ),
            },
            {
                "heading": "🔍 重症度フィルターの活用",
                "body": (
                    "辞書ページの 🚦 フィルターで、重症度別の疾患を一覧できます。"
                    "繁殖計画では特に <strong>高リスク疾患</strong> の検査を優先することをおすすめします。"
                ),
            },
        ],
        "related_disease_slugs": [],
        "related_trait_slugs": [],
    },
    # === 犬種別ガイド (PR #62) ===
    {
        "slug": "poodle-genetic-health-guide",
        "title": "プードル飼い主・ブリーダー向け遺伝子検査ガイド",
        "summary": "スタンダード/ミニチュア/トイ/タイニープードルで特に注意すべき遺伝性疾患・毛色遺伝子をまとめたガイド。",
        "category": "🐩 犬種別",
        "reading_time": "7 分",
        "sections": [
            {
                "heading": "🐩 プードルで特に重要な遺伝病",
                "body": (
                    "プードルは全体的に健康な犬種ですが、いくつか特有の遺伝性疾患があります:<br>"
                    "・<strong>NEwS (新生児脳症)</strong> — スタンダードプードル特有。生後 4〜6 週で致死。両親キャリア × キャリア交配は厳禁。<br>"
                    "・<strong>prcd-PRA (進行性網膜萎縮症)</strong> — 全サイズ。中年期から失明。<br>"
                    "・<strong>vWD1 (フォン・ヴィレブランド病 I型)</strong> — 出血傾向。手術前に申告。<br>"
                    "・<strong>HSF4 白内障</strong> — 一部ラインで報告。<br>"
                    "・<strong>DM (変性性脊髄症)</strong> — 大型プードルで報告例あり。"
                ),
            },
            {
                "heading": "🎨 プードルの毛色遺伝学",
                "body": (
                    "プードルは多彩な毛色を持つ犬種で、以下の座位が重要:<br>"
                    "・<strong>E座位 (MC1R)</strong> — クリーム/アプリコット/レッドの基本色制御<br>"
                    "・<strong>K座位 (CBD103)</strong> — ブラック/ブラウン(チョコ) のソリッド色<br>"
                    "・<strong>B座位 (TYRP1)</strong> — bb でブラウン (チョコ/レバー) になる<br>"
                    "・<strong>D座位 (MLPH)</strong> — dd でブルー/シルバービーグへ希釈<br>"
                    "・<strong>G座位 (Greying)</strong> — シルバープードルの原因。生まれは黒、成犬で退色<br>"
                    "・<strong>S座位 (MITF)</strong> — sp/sp でパーティカラー"
                ),
            },
            {
                "heading": "✂ プードル特有の被毛形質",
                "body": (
                    "・<strong>C/C (KRT71)</strong> — 巻き毛（プードルは全て C/C ホモ）<br>"
                    "・<strong>F/F (RSPO2)</strong> — ファーニシング（眉・髭・飾り毛）<br>"
                    "・<strong>l/l (FGF5)</strong> — 長毛（プードルは全て l/l ホモ）<br>"
                    "・<strong>N/N (MC5R)</strong> — 抜け毛少なめ（『ハイポアレジェニック』と称される根拠）<br>"
                    "これら全てがプードル特有の『くるくる長い毛・抜け毛少ない』被毛を形成します。"
                ),
            },
            {
                "heading": "🐕 サイズ別の注意点",
                "body": (
                    "<strong>スタンダードプードル</strong>: NEwS リスク特有。股関節形成不全・SARDS も注意。COI が低い系統選択を推奨。<br>"
                    "<strong>ミニチュアプードル</strong>: 膝蓋骨脱臼・てんかんの素因あり。<br>"
                    "<strong>トイ/タイニープードル</strong>: 低血糖・水頭症・歯の問題が多い。極端な小型化を避けた繁殖を。"
                ),
            },
            {
                "heading": "💡 プードル繁殖の推奨ステップ",
                "body": (
                    "1. <strong>両親の遺伝子検査</strong>を最低 8 項目（プードル標準パネル）<br>"
                    "2. <strong>COI 6.25% 以下</strong>を維持する系統選択<br>"
                    "3. <strong>NEwS / prcd-PRA / vWD1</strong> の P/N × P/N 交配を絶対避ける<br>"
                    "4. M座位 (Merle) を持つ犬は他のメルル犬と交配しない<br>"
                    "5. 子犬の遺伝子検査を生後早期に実施"
                ),
            },
        ],
        "related_disease_slugs": ["neonatal-encephalopathy", "progressive-rod-cone", "willebrand-type-1", "hereditary-cataract", "degenerative-myelopathy"],
        "related_trait_slugs": ["e-locus", "k-locus", "b-locus", "d-locus", "g-locus", "curly-coat", "furnishings", "l-locus"],
    },
    {
        "slug": "labrador-genetic-health-guide",
        "title": "ラブラドール飼い主・ブリーダー向け遺伝子検査ガイド",
        "summary": "ラブラドール・レトリーバーで特に注意すべき遺伝性疾患・毛色遺伝子をまとめたガイド。EIC や HNPK 等の犬種特有疾患も解説。",
        "category": "🐕 犬種別",
        "reading_time": "7 分",
        "sections": [
            {
                "heading": "🐕 ラブラドールで特に重要な遺伝病",
                "body": (
                    "ラブラドールは人気犬種ゆえに遺伝病情報が豊富です:<br>"
                    "・<strong>EIC (運動誘発性虚脱)</strong> — ラブ特有。激しい運動後に脱力。<br>"
                    "・<strong>CNM (中心核ミオパチー)</strong> — 若齢期から筋力低下。<br>"
                    "・<strong>prcd-PRA</strong> — 中年期失明。<br>"
                    "・<strong>HNPK (鼻過角化症)</strong> — 鼻の硬化・亀裂。ラブで頻発。<br>"
                    "・<strong>銅蓄積性肝障害 (COMMD1)</strong> — 肝臓に銅が異常蓄積。<br>"
                    "・<strong>HUU (高尿酸尿症)</strong> — 一部ラインで報告。<br>"
                    "・<strong>SD2 (骨格異形成 2型)</strong> — 四肢の異常な短縮。<br>"
                    "・<strong>CDDY+IVDD</strong> — 椎間板疾患リスク（短足ラブで特に）"
                ),
            },
            {
                "heading": "🎨 ラブラドールの3つの基本色",
                "body": (
                    "<strong>ブラック (E_, B_)</strong>: 標準色。E と B の両方が機能。<br>"
                    "<strong>イエロー (e/e)</strong>: ee ホモで黒/茶色素が coat に出ない。"
                    "明るいクリームから濃いフォックスレッドまで幅広い（KITLG が濃淡決定）。<br>"
                    "<strong>チョコレート (bb)</strong>: bb ホモで黒色素がブラウンに変換。<br>"
                    "<strong>シルバー/チャコール</strong>: dd (希釈) によるブルー/シャンパン。"
                    "ラブのシルバーは AKC で論争的。Em (マスク) も発現する場合あり。"
                ),
            },
            {
                "heading": "🦴 ラブラドール特有の被毛・体型形質",
                "body": (
                    "・<strong>L/L (FGF5)</strong> — 短毛（ラブは全て L/L）。たまに l/l で出るのは『ファジー』ラブと呼ばれる長毛。<br>"
                    "・<strong>SD/SD (MC5R)</strong> — シェッディング多。ラブは抜け毛が多い犬種。<br>"
                    "・<strong>F無し (RSPO2 N/N)</strong> — ファーニシングなし、スムースコート。<br>"
                    "・<strong>Em/E (MC1R variant)</strong> — 一部ラブで黒マスクが見られる。"
                ),
            },
            {
                "heading": "💡 ラブラドール繁殖の推奨パネル",
                "body": (
                    "最低限テストすべき項目:<br>"
                    "・EIC、CNM、prcd-PRA、HNPK、CDDY、CNM、HUU、Copper Toxicosis、Centronuclear Myopathy<br>"
                    "ラブラドール特有の遺伝病が多いため、Embark や Orivet の『Labrador panel』を活用しましょう。"
                ),
            },
        ],
        "related_disease_slugs": ["exercise-induced-collapse", "centronuclear-myopathy", "progressive-rod-cone", "hnpk", "copper-toxicosis", "hyperuricosuria", "skeletal-dysplasia-2", "chondrodystrophy"],
        "related_trait_slugs": ["e-locus", "b-locus", "d-locus", "l-locus", "shedding", "em-locus"],
    },
    {
        "slug": "doodle-genetic-health-guide",
        "title": "ドゥードゥル系犬種（Goldendoodle/Labradoodle 等）の遺伝子検査ガイド",
        "summary": "プードル × Golden/Labrador 系のミックス犬では両親犬種の遺伝病パネル両方が必要。被毛予測も複雑。",
        "category": "🐾 犬種別",
        "reading_time": "7 分",
        "sections": [
            {
                "heading": "🐾 ドゥードゥル系犬種とは",
                "body": (
                    "プードルと他犬種の交配で生まれる F1 / F1B / F2 等のミックス犬種:<br>"
                    "・<strong>Goldendoodle</strong> — プードル × ゴールデン・レトリーバー<br>"
                    "・<strong>Labradoodle / Australian Labradoodle</strong> — プードル × ラブラドール<br>"
                    "・<strong>Bernedoodle</strong> — プードル × バーニーズ・マウンテン・ドッグ<br>"
                    "・<strong>Sheepadoodle</strong> — プードル × オールド・イングリッシュ・シープドッグ<br>"
                    "・<strong>Cavapoo</strong> — プードル × キャバリア K.C. スパニエル<br>"
                    "各組み合わせで、両親犬種固有の遺伝病パネルが必要になります。"
                ),
            },
            {
                "heading": "🩺 ドゥードゥル繁殖で重要な遺伝病パネル",
                "body": (
                    "両親犬種の遺伝病を網羅的に検査:<br>"
                    "<strong>プードル側</strong>: prcd-PRA、NEwS (スタンダードの場合)、vWD1、HSF4<br>"
                    "<strong>ゴールデン側</strong>: prcd-PRA、Ichthyosis、心臓関連<br>"
                    "<strong>ラブラドール側</strong>: EIC、CNM、HNPK、HUU、Copper Toxicosis<br>"
                    "<strong>バーニーズ側</strong>: DM、骨肉腫リスク、組織球性肉腫<br>"
                    "<strong>キャバリア側</strong>: 僧帽弁疾患、Macrothrombocytopenia、EFS、DM<br>"
                    "両親が同じ変異の P/N を持つ場合、ドゥードゥルでも 25% 発症します。"
                ),
            },
            {
                "heading": "🎨 被毛タイプの予測",
                "body": (
                    "ドゥードゥル系の被毛は『プードルらしさ』を決める2座位の組み合わせで決まる:<br>"
                    "<strong>C/C (KRT71 巻き毛)</strong>: カーリー<br>"
                    "<strong>C/N</strong>: ウェーブ<br>"
                    "<strong>N/N</strong>: ストレート<br>"
                    "<strong>F/F (RSPO2 ファーニシング)</strong>: 顔毛しっかり、抜け毛少なめ<br>"
                    "<strong>F/N</strong>: 中間<br>"
                    "<strong>N/N (Improper Coat)</strong>: ラブ/ゴールデン的なスムースな顔<br>"
                    "F1 (純血種同士の初代) は F/N × C/N が多く、被毛タイプにバラつきが出やすい。<br>"
                    "<strong>『ハイポアレジェニック』を狙うなら F/F + C/C + N/N (Shedding)</strong> が理想。"
                ),
            },
            {
                "heading": "💡 ドゥードゥルブリーダーへのアドバイス",
                "body": (
                    "1. <strong>両親両方</strong>の遺伝子検査が必須（片方だけは無意味）<br>"
                    "2. <strong>COI 計算</strong>は両親が異犬種でも、共通の祖先がいる場合（同系プードル等）に上昇<br>"
                    "3. <strong>被毛タイプは確率予測</strong>のみ — F/F C/C N/N の両親同士でも、F1 子犬は被毛が混ざる<br>"
                    "4. <strong>初代 (F1) より F1B が予測しやすい</strong>（F1 × プードル戻し交配）<br>"
                    "5. 健康とブリーダーの倫理が優先 — 被毛より遺伝病パネルを優先"
                ),
            },
        ],
        "related_disease_slugs": ["progressive-rod-cone", "ichthyosis", "exercise-induced-collapse", "centronuclear-myopathy", "hnpk", "degenerative-myelopathy", "macrothrombocytopenia"],
        "related_trait_slugs": ["curly-coat", "furnishings", "shedding", "l-locus", "e-locus", "b-locus"],
    },
    # === 追加犬種別ガイド (Orivet JP 対応犬種 + 日本人気犬種) ===
    {
        "slug": "shiba-genetic-health-guide",
        "title": "柴犬の遺伝子検査・健康ガイド",
        "summary": "日本犬の代表・柴犬で重要な遺伝病（GM1ガングリオシドーシス等）・毛色遺伝学を解説。Orivet 検査対応犬種。",
        "category": "🐕 犬種別",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🐕 柴犬で特に重要な遺伝病",
                "body": (
                    "柴犬は比較的健康な犬種ですが、特有の遺伝病があります:<br>"
                    "・<strong>GM1 ガングリオシドーシス (GLB1)</strong> — 柴犬で確立した致死性神経疾患。両親キャリア × キャリア交配で 25% 発症。<br>"
                    "・<strong>緑内障</strong> — 中年期から発症リスク。失明予防に定期眼圧測定。<br>"
                    "・<strong>アトピー性皮膚炎</strong> — 遺伝的素因 + 環境要因。<br>"
                    "・<strong>膝蓋骨脱臼</strong> — 中型犬として一般的なリスク。<br>"
                    "・<strong>GM2 ガングリオシドーシス</strong> — 一部ラインで報告（GM1 ほど頻発しない）"
                ),
            },
            {
                "heading": "🎨 柴犬の毛色遺伝学",
                "body": (
                    "柴犬の4毛色は以下の組み合わせ:<br>"
                    "<strong>赤</strong>: ay/_ ky/ky E_ B_ (アグーチ・セーブル発現)<br>"
                    "<strong>黒褐 (黒タン)</strong>: at/at ky/ky E_ B_ (タンポイント)<br>"
                    "<strong>胡麻 (ごま)</strong>: ay/at + 個体差で胡麻状の混色<br>"
                    "<strong>白</strong>: e/e (劣性) でクリーム/白系。少数派。<br>"
                    "<strong>ウラジロ (裏白)</strong>: 一般的な裏白模様は S 座位とは別の遺伝子。"
                ),
            },
            {
                "heading": "💡 柴犬繁殖の推奨",
                "body": (
                    "・<strong>GM1 検査必須</strong> — 日本のブリーダーは特に注意<br>"
                    "・四毛色（赤・黒褐・胡麻・白）の維持には A/E/B 座位の遺伝管理<br>"
                    "・小型化の極端化は健康問題を増やすため避ける"
                ),
            },
        ],
        "related_disease_slugs": ["gm1-gangliosidosis", "gm2-gangliosidosis", "glaucoma"],
        "related_trait_slugs": ["a-locus", "e-locus", "b-locus", "k-locus"],
    },
    {
        "slug": "akita-genetic-health-guide",
        "title": "秋田犬の遺伝子検査・健康ガイド",
        "summary": "秋田犬で重要な遺伝病（DM・免疫疾患等）と特有の遺伝形質を解説。Orivet 検査対応犬種。",
        "category": "🐕 犬種別",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "🐕 秋田犬で特に重要な遺伝病",
                "body": (
                    "・<strong>DM (変性性脊髄症)</strong> — 秋田犬は SOD1 リスクアレルを保有することが多い。<br>"
                    "・<strong>VKH 様症候群</strong> — 自己免疫疾患（眼・皮膚色素脱失）。秋田で多い。<br>"
                    "・<strong>進行性網膜萎縮症 (PRA)</strong> — 一部系統で報告。<br>"
                    "・<strong>股関節形成不全</strong> — 大型犬として注意。<br>"
                    "・<strong>甲状腺機能低下症</strong> — 中高齢期から発症。"
                ),
            },
            {
                "heading": "🎨 秋田犬の毛色",
                "body": (
                    "<strong>赤</strong>: ay/_ + 一般的なウラジロ模様<br>"
                    "<strong>白</strong>: e/e でクリーム/白系<br>"
                    "<strong>胡麻</strong>: アグーチによる混色<br>"
                    "<strong>虎毛 (ブリンドル)</strong>: kbr アレル発現（日本秋田より米国秋田で多い）"
                ),
            },
            {
                "heading": "💡 秋田犬繁殖の推奨",
                "body": (
                    "・<strong>DM 検査</strong>を成犬時に必ず実施。P/P 犬同士の交配は避ける。<br>"
                    "・自己免疫疾患の素因を考慮した系統選択<br>"
                    "・COI を低く保つ（秋田犬は犬種内多様性が低めなため）"
                ),
            },
        ],
        "related_disease_slugs": ["degenerative-myelopathy", "progressive-rod-cone"],
        "related_trait_slugs": ["a-locus", "e-locus", "k-locus"],
    },
    {
        "slug": "shar-pei-genetic-health-guide",
        "title": "シャー・ペイの遺伝子検査・健康ガイド",
        "summary": "シワが特徴のシャー・ペイで重要な家族性シャーペイ熱・皮膚疾患を解説。Orivet 検査対応犬種。",
        "category": "🐕 犬種別",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "🐕 シャー・ペイで特に重要な遺伝病",
                "body": (
                    "・<strong>家族性シャーペイ熱 (FSF)</strong> — 周期的な発熱と関節腫脹。アミロイドーシスのリスクと関連。シャー・ペイ特有。<br>"
                    "・<strong>進行性腎臓アミロイドーシス</strong> — FSF の長期合併症。<br>"
                    "・<strong>瞼内反症 / 外反症</strong> — シワ多い犬種で多発。<br>"
                    "・<strong>パッド過角化症</strong> — 皮膚バリア機能の遺伝的低下。<br>"
                    "・<strong>POAG（原発性開放隅角緑内障）</strong>"
                ),
            },
            {
                "heading": "🎨 シャー・ペイの毛色・被毛",
                "body": (
                    "<strong>毛色</strong>: 黒・ブラウン・クリーム・赤・チョコ等の多彩な色<br>"
                    "<strong>被毛タイプ</strong>:<br>"
                    "・ホースコート (短い・硬い)<br>"
                    "・ブラッシュコート (やや長め)<br>"
                    "・ベアーコート (長毛・FCI 非認可)<br>"
                    "L 座位 (FGF5) の l/l がベアーコートの原因。"
                ),
            },
            {
                "heading": "💡 シャー・ペイ繁殖の推奨",
                "body": (
                    "・FSF / アミロイドーシスの家族歴を必ず確認<br>"
                    "・極端なシワ強調を避けた繁殖（皮膚疾患リスク）<br>"
                    "・若齢期からの定期的な腎機能検査"
                ),
            },
        ],
        "related_disease_slugs": ["glaucoma"],
        "related_trait_slugs": ["l-locus", "e-locus", "b-locus"],
    },
    {
        "slug": "chin-genetic-health-guide",
        "title": "狆（チン）の遺伝子検査・健康ガイド",
        "summary": "日本の伝統小型犬・狆で重要な遺伝病と特性。Orivet 検査対応犬種。",
        "category": "🐕 犬種別",
        "reading_time": "4 分",
        "sections": [
            {
                "heading": "🐕 狆で特に重要な遺伝病",
                "body": (
                    "・<strong>ガングリオシドーシス GM2</strong> — 一部ラインで報告。重篤な神経疾患。<br>"
                    "・<strong>頭蓋骨変形（Brachycephalic Syndrome）</strong> — 短頭種症候群。呼吸器・眼科リスク。<br>"
                    "・<strong>水頭症</strong> — 小型犬として注意。<br>"
                    "・<strong>膝蓋骨脱臼</strong> — 小型犬として一般的。<br>"
                    "・<strong>白内障</strong> — 中年期以降。"
                ),
            },
            {
                "heading": "🎨 狆の毛色",
                "body": (
                    "<strong>白黒 (Black & White)</strong>: at/at + 白斑 (sp/sp)<br>"
                    "<strong>白赤 (Red & White)</strong>: ay/_ + 白斑<br>"
                    "白の割合が多いのが特徴で、S 座位 sp/sp の発現が顕著。"
                ),
            },
            {
                "heading": "💡 狆繁殖の推奨",
                "body": (
                    "・短頭種特有の呼吸器負担を考慮した骨格繁殖<br>"
                    "・小型化を極端に進めない<br>"
                    "・GM2 等の遺伝子検査"
                ),
            },
        ],
        "related_disease_slugs": ["gm2-gangliosidosis"],
        "related_trait_slugs": ["a-locus", "e-locus", "s-locus"],
    },
    {
        "slug": "dachshund-genetic-health-guide",
        "title": "ダックスフンドの遺伝子検査・健康ガイド",
        "summary": "短足の代表・ダックスフンドで特に注意すべき椎間板疾患 (CDDY+IVDD)・PRA を解説。",
        "category": "🐕 犬種別",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🐕 ダックスフンドで特に重要な遺伝病",
                "body": (
                    "・<strong>CDDY+IVDD (椎間板疾患)</strong> — ダックス系で最重要。短足の原因 CDPA に加え、椎間板リスクの CDDY を保有。<br>"
                    "・<strong>CORD1 / PRA</strong> — ミニチュアロングヘアード Dx で特に。失明リスク。<br>"
                    "・<strong>ラフォラ病</strong> — 進行性ミオクローヌス癲癇。ミニチュア Dx で報告。<br>"
                    "・<strong>骨形成不全症 (OI)</strong> — ダックスで報告例あり。<br>"
                    "・<strong>てんかん</strong> — 多因子。"
                ),
            },
            {
                "heading": "🎨 ダックスフンドの毛色・被毛",
                "body": (
                    "<strong>毛色</strong>: ブラックタン・チョコタン・レッド・クリーム・シルバーダップル (Merle)・ピーバルド 等<br>"
                    "<strong>被毛タイプ</strong>: スムース (短毛)・ロングヘアード (l/l)・ワイヤーヘアード (F/F + Curl 混合)<br>"
                    "⚠️ <strong>ダップル × ダップル交配は厳禁</strong> (M/M ダブルマール = 失明・難聴)"
                ),
            },
            {
                "heading": "💡 ダックスフンド飼育の推奨",
                "body": (
                    "・<strong>椎間板予防</strong>: 体重管理・階段の昇降制限・ジャンプ禁止<br>"
                    "・<strong>CDDY 検査</strong>: 全頭推奨<br>"
                    "・<strong>PRA / Lafora 検査</strong>: ミニチュアロングで必須<br>"
                    "・ダップル交配規則の厳守"
                ),
            },
        ],
        "related_disease_slugs": ["chondrodystrophy", "cord1", "progressive-rod-cone", "lafora", "osteogenesis-imperfecta"],
        "related_trait_slugs": ["m-locus", "s-locus", "l-locus", "curly-coat", "furnishings"],
    },
    {
        "slug": "french-bulldog-genetic-health-guide",
        "title": "フレンチブルドッグの遺伝子検査・健康ガイド",
        "summary": "日本で人気のフレブルで重要な短頭種疾患・遺伝病・ブルー(dd)関連 CDA を解説。",
        "category": "🐕 犬種別",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🐕 フレンチブルドッグで特に重要な遺伝病",
                "body": (
                    "・<strong>短頭種気道症候群 (BOAS)</strong> — 呼吸器負担。麻酔リスク高。<br>"
                    "・<strong>椎骨形成異常 (Hemivertebrae)</strong> — フレブル特有の脊椎奇形。<br>"
                    "・<strong>CDA (毛色希釈性脱毛症)</strong> — ブルー (dd) フレブルで頻発。<br>"
                    "・<strong>HUU (高尿酸尿症)</strong> — 一部ラインで報告。<br>"
                    "・<strong>多発性軟骨外骨腫</strong><br>"
                    "・<strong>白内障 / チェリーアイ</strong>"
                ),
            },
            {
                "heading": "🎨 フレンチブルドッグの毛色",
                "body": (
                    "<strong>標準色 (FCI 認可)</strong>: フォーン・ブリンドル・パイド<br>"
                    "<strong>非認可色</strong>: ブルー (dd)・チョコ (bb)・ライラック (bb dd)・マール (M/_)<br>"
                    "ブルー系は CDA リスクが付随。マールは FCI 非認可 + 健康リスク（M/M ダブルマール厳禁）"
                ),
            },
            {
                "heading": "💡 フレブル繁殖の推奨",
                "body": (
                    "・<strong>BOAS スコアリング</strong>を実施した個体での繁殖<br>"
                    "・<strong>椎骨レントゲン</strong>で奇形チェック<br>"
                    "・帝王切開率が高いため、ブリーダーは産科対応必須<br>"
                    "・希釈色 (dd) を選ぶ場合は CDA リスクを認識"
                ),
            },
        ],
        "related_disease_slugs": ["coat-color-dilution-alopecia", "hyperuricosuria", "hereditary-cataract"],
        "related_trait_slugs": ["d-locus", "b-locus", "m-locus", "k-locus"],
    },
    {
        "slug": "cavalier-genetic-health-guide",
        "title": "キャバリア K.C. スパニエルの遺伝子検査・健康ガイド",
        "summary": "キャバリア特有の僧帽弁疾患 (MVD)・先天性巨大血小板減少症・EFS を解説。",
        "category": "🐕 犬種別",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🐕 キャバリアで特に重要な遺伝病",
                "body": (
                    "・<strong>僧帽弁疾患 (MVD)</strong> — キャバリア最大の死因。中年期から発症。心臓検査必須。<br>"
                    "・<strong>キアリ様奇形 / SM (脊髄空洞症)</strong> — 頭部の発達異常で神経症状。MRI 検査必要。<br>"
                    "・<strong>EFS (発作性失神症 / BCAN)</strong> — 興奮で発作。キャバリア特有。<br>"
                    "・<strong>Macrothrombocytopenia (TUBB1)</strong> — 巨大血小板。多くは無症状。<br>"
                    "・<strong>DM (変性性脊髄症)</strong> — 一部ラインで報告。<br>"
                    "・<strong>白内障 (HSF4)</strong>"
                ),
            },
            {
                "heading": "🎨 キャバリアの4毛色",
                "body": (
                    "<strong>ブレナム (赤白)</strong>: ay/_ + 白斑<br>"
                    "<strong>トライカラー (黒赤白)</strong>: at/at + 白斑<br>"
                    "<strong>ルビー (赤)</strong>: e/e (劣性レッド)<br>"
                    "<strong>ブラック・アンド・タン</strong>: at/at"
                ),
            },
            {
                "heading": "💡 キャバリア繁殖の推奨",
                "body": (
                    "・<strong>MVD 検査</strong>を成犬時に実施。心臓スコアでブリーダー選定。<br>"
                    "・<strong>SM の MRI スクリーニング</strong>（コストは高いが推奨）<br>"
                    "・EFS / Macrothrombo / DM の遺伝子検査<br>"
                    "・キャバリアは犬種内 COI が高めなので、低 COI 系統選択が重要"
                ),
            },
        ],
        "related_disease_slugs": ["episodic-falling", "macrothrombocytopenia", "degenerative-myelopathy", "hereditary-cataract"],
        "related_trait_slugs": ["a-locus", "e-locus", "s-locus"],
    },
    {
        "slug": "border-collie-genetic-health-guide",
        "title": "ボーダーコリーの遺伝子検査・健康ガイド",
        "summary": "知能高い牧羊犬・ボーダーコリーで多数の遺伝子検査が必要な理由とパネルを解説。",
        "category": "🐕 犬種別",
        "reading_time": "7 分",
        "sections": [
            {
                "heading": "🐕 ボーダーコリーで特に重要な遺伝病",
                "body": (
                    "ボーダーコリーは遺伝子検査が最も充実した犬種の1つ:<br>"
                    "・<strong>CEA (コリーアイ症候群)</strong> — 眼球の発達異常。<br>"
                    "・<strong>TNS (好中球機能不全症候群)</strong> — 重度免疫不全、幼齢致死。<br>"
                    "・<strong>NCL (神経セロイドリポフスチン症)</strong> — 進行性神経変性。<br>"
                    "・<strong>MDR1 (多剤耐性)</strong> — イベルメクチン等の薬剤過敏症。<br>"
                    "・<strong>感覚性神経障害 (SN / FAM134B)</strong> — 自咬リスク。<br>"
                    "・<strong>CL (Ceroid Lipofuscinosis 各型)</strong><br>"
                    "・<strong>DM</strong>、<strong>てんかん</strong>"
                ),
            },
            {
                "heading": "🎨 ボーダーコリーの多彩な毛色",
                "body": (
                    "<strong>標準色</strong>: ブラック&ホワイト・レッド&ホワイト・トライカラー・ブルー&ホワイト<br>"
                    "<strong>マール</strong>: ブルーマール・レッドマール・スレートマール<br>"
                    "<strong>レア色</strong>: ライラック・シール・ブリンドル<br>"
                    "⚠️ <strong>マール × マール交配は厳禁</strong>（M/M ダブルマール）"
                ),
            },
            {
                "heading": "💡 ボーダーコリー繁殖の推奨",
                "body": (
                    "<strong>必須パネル（最低限）</strong>:<br>"
                    "CEA / TNS / NCL / MDR1 / IGS (B12)<br>"
                    "<strong>推奨パネル</strong>:<br>"
                    "DM / SN / 全 CL タイプ / EAOD（早期発症成犬難聴）<br>"
                    "<strong>運動犬種としての注意</strong>:<br>"
                    "高運動量犬種なので、関節検査と運動誘発性虚脱の確認も。"
                ),
            },
        ],
        "related_disease_slugs": ["collie-eye-anomaly", "trapped-neutrophil-syndrome", "neuronal-ceroid-lipofuscinosis", "multidrug-resistance", "sensory-neuropathy", "degenerative-myelopathy", "cobalamin-malabsorption"],
        "related_trait_slugs": ["m-locus", "s-locus", "k-locus"],
    },
    {
        "slug": "german-shepherd-genetic-health-guide",
        "title": "ジャーマンシェパードの遺伝子検査・健康ガイド",
        "summary": "知能・運動能力高いジャーマンシェパード特有の DM・RCND・下垂体性小人症等を解説。",
        "category": "🐕 犬種別",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🐕 ジャーマンシェパードで特に重要な遺伝病",
                "body": (
                    "・<strong>DM (変性性脊髄症)</strong> — 中高齢期から後肢麻痺進行。SOD1 保有率が高い犬種。<br>"
                    "・<strong>RCND (腎嚢腺癌・結節性皮膚線維腫症 / FLCN)</strong> — ジャーマンシェパード特有の腫瘍症候群。<br>"
                    "・<strong>下垂体性小人症 (LHX3)</strong> — 成長異常。<br>"
                    "・<strong>股関節形成不全</strong> — 大型犬として注意。OFA / PennHIP 検査推奨。<br>"
                    "・<strong>肘関節形成不全</strong>、<strong>てんかん</strong><br>"
                    "・<strong>膵外分泌不全 (EPI)</strong>"
                ),
            },
            {
                "heading": "🎨 ジャーマンシェパードの毛色",
                "body": (
                    "<strong>標準色</strong>: ブラック&タン（黒マスク Em/_）・セーブル<br>"
                    "<strong>レア色</strong>: 全黒 (a/a)・全白 (e/e + 色素脱失)・パンダ<br>"
                    "ブラック&タンはアグーチ at/at + Em マスクの組み合わせ。"
                ),
            },
            {
                "heading": "💡 ジャーマンシェパード繁殖の推奨",
                "body": (
                    "・<strong>DM 検査</strong>を成犬時に。P/P × P/P は厳禁。<br>"
                    "・<strong>股関節レントゲン</strong>を 12 ヶ月齢以降に。<br>"
                    "・RCND の家族歴確認<br>"
                    "・大型犬の宿命として COI を低く保つことが重要"
                ),
            },
        ],
        "related_disease_slugs": ["degenerative-myelopathy", "renal-cystadenocarcinoma", "pituitary-dwarfism", "x-linked-hereditary-nephropathy"],
        "related_trait_slugs": ["a-locus", "em-locus", "k-locus", "b-locus"],
    },
    {
        "slug": "mini-schnauzer-genetic-health-guide",
        "title": "ミニチュアシュナウザーの遺伝子検査・健康ガイド",
        "summary": "ミニシュナで特に重要な PMDS・MAC（高脂血症）・若年性白内障等を解説。",
        "category": "🐕 犬種別",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "🐕 ミニシュナで特に重要な遺伝病",
                "body": (
                    "・<strong>PMDS (ミュラー管遺残症候群 / AMHR2)</strong> — オス犬が子宮・卵管を残す発達異常。ミニシュナで報告。<br>"
                    "・<strong>高脂血症 / 膵炎</strong> — 遺伝的素因。食事管理が重要。<br>"
                    "・<strong>若年性白内障 (HSF4)</strong> — 一部ラインで報告。<br>"
                    "・<strong>糖尿病</strong> — 中高齢期から発症リスク。<br>"
                    "・<strong>尿石症</strong>（シュウ酸カルシウム結石）<br>"
                    "・<strong>進行性網膜萎縮症 (PRA)</strong>"
                ),
            },
            {
                "heading": "🎨 ミニシュナの毛色",
                "body": (
                    "<strong>FCI 認可色</strong>: ソルト&ペッパー (G座位による退色)・ブラック&シルバー (at/at + G)・ブラック・ホワイト (e/e)<br>"
                    "<strong>非認可色</strong>: チョコレート (bb)・パーティカラー (sp/sp)"
                ),
            },
            {
                "heading": "💡 ミニシュナ繁殖の推奨",
                "body": (
                    "・<strong>低脂肪食</strong>での飼育（膵炎予防）<br>"
                    "・<strong>HSF4 / PRA 検査</strong><br>"
                    "・若年性白内障の家族歴確認<br>"
                    "・PMDS のオス犬の繁殖は再考"
                ),
            },
        ],
        "related_disease_slugs": ["persistent-mullerian-duct-syndrome", "hereditary-cataract", "progressive-rod-cone"],
        "related_trait_slugs": ["a-locus", "g-locus", "b-locus", "s-locus"],
    },
    # === 追加犬種別ガイド（人気 8 犬種） ===
    {
        "slug": "golden-retriever-genetic-health-guide",
        "title": "ゴールデン・レトリーバーの遺伝子検査・健康ガイド",
        "summary": "ゴールデン・レトリーバーで特に重要な GR-PRA・Ichthyosis・関節疾患・腫瘍リスクを解説。",
        "category": "🐕 犬種別",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🐕 ゴールデン・レトリーバーで特に重要な遺伝病",
                "body": (
                    "・<strong>GR-PRA1 / GR-PRA2 / prcd-PRA</strong> — ゴールデン特有の3種類の進行性網膜萎縮症。中年期からの失明リスク。<br>"
                    "・<strong>Ichthyosis (PNPLA1)</strong> — 鱗状の皮膚・フケ。ゴールデンで頻発。<br>"
                    "・<strong>HSF4 白内障</strong> — 若年性白内障。<br>"
                    "・<strong>変性性脊髄症 (DM)</strong> — SOD1 リスクアレル保有。<br>"
                    "・<strong>股関節・肘関節形成不全</strong> — 大型犬として注意。<br>"
                    "・<strong>感覚性神経障害 (SN)</strong> — 自咬・神経麻痺。一部ラインで報告。<br>"
                    "・<strong>腫瘍素因</strong> — 血管肉腫・リンパ腫・骨肉腫（多因子だが家族歴重要）"
                ),
            },
            {
                "heading": "🎨 ゴールデンの毛色",
                "body": (
                    "ゴールデンは E座位 e/e の劣性レッドが固定された犬種:<br>"
                    "・<strong>標準色</strong>: 全頭が e/e でクリーム〜濃いゴールド<br>"
                    "・<strong>KITLG / I 座位</strong>: 黄色濃度の主因。イングリッシュ・クリームの淡色から濃いレッドまで幅広い<br>"
                    "・<strong>稀色</strong>: ブラック（黒）が出る場合はゴールデン以外の血が混入（K + E_）"
                ),
            },
            {
                "heading": "💡 ゴールデン繁殖の推奨パネル",
                "body": (
                    "・<strong>3種類の PRA 全て</strong>（GR-PRA1 / GR-PRA2 / prcd）<br>"
                    "・<strong>Ichthyosis</strong> — 健康な被毛維持に必須<br>"
                    "・<strong>DM / HSF4 / SN</strong><br>"
                    "・<strong>股関節 OFA / PennHIP</strong> 12ヶ月齢以降<br>"
                    "・腫瘍歴の家族歴確認（特に血管肉腫の高い犬種内発生率）"
                ),
            },
        ],
        "related_disease_slugs": ["progressive-rod-cone", "ichthyosis", "hereditary-cataract", "degenerative-myelopathy", "sensory-neuropathy"],
        "related_trait_slugs": ["e-locus", "kitlg", "i-locus", "k-locus"],
    },
    {
        "slug": "welsh-corgi-genetic-health-guide",
        "title": "ウェルシュ・コーギーの遺伝子検査・健康ガイド",
        "summary": "ペンブローク／カーディガンで重要な DM・椎間板疾患・自然短尾（BT）を解説。",
        "category": "🐕 犬種別",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🐕 ウェルシュ・コーギーで特に重要な遺伝病",
                "body": (
                    "・<strong>変性性脊髄症 (DM)</strong> — ペンブロークは SOD1 リスクアレル頻度が最も高い犬種の1つ。発症率高。<br>"
                    "・<strong>CDDY + IVDD（椎間板疾患）</strong> — 短足犬種として椎間板リスクが付随。<br>"
                    "・<strong>フォン・ヴィレブランド病 I型 (vWD1)</strong> — 一部ラインで報告。<br>"
                    "・<strong>運動誘発性虚脱 (EIC)</strong> — ペンブロークで報告例あり。<br>"
                    "・<strong>進行性網膜萎縮症 (prcd-PRA)</strong> — カーディガンで報告。<br>"
                    "・<strong>股関節形成不全</strong> — 重い体重と短足のため注意。"
                ),
            },
            {
                "heading": "🎨 コーギーの毛色・形質",
                "body": (
                    "<strong>ペンブローク標準色</strong>: レッド (ay)・セーブル (ay)・トライカラー (at/at)・フォーン<br>"
                    "<strong>カーディガン</strong>: 上記 + ブリンドル (kbr)・ブルーマール (M/m)<br>"
                    "<strong>自然短尾 (BT 座位 / Brachyury)</strong>: ペンブロークの一部は生まれつき短尾。BT/BT は胚致死のため BT/N × N/N が標準。<br>"
                    "<strong>カーディガンは長尾</strong>: BT 検査不要。"
                ),
            },
            {
                "heading": "💡 コーギー繁殖の推奨",
                "body": (
                    "・<strong>DM 検査必須</strong> — ペンブロークでは特に。P/P × P/P は厳禁、可能なら P/N × N/N で頻度を下げる<br>"
                    "・<strong>CDDY 検査</strong>と椎間板予防の生活管理（体重・階段制限）<br>"
                    "・<strong>BT 検査</strong>: ペンブロークは BT/BT 交配を必ず避ける<br>"
                    "・<strong>カーディガンのマール × マール厳禁</strong>（ダブルマール失明・難聴）"
                ),
            },
        ],
        "related_disease_slugs": ["degenerative-myelopathy", "chondrodystrophy", "willebrand-type-1", "exercise-induced-collapse", "progressive-rod-cone"],
        "related_trait_slugs": ["a-locus", "k-locus", "m-locus", "bob-tail"],
    },
    {
        "slug": "pomeranian-genetic-health-guide",
        "title": "ポメラニアンの遺伝子検査・健康ガイド",
        "summary": "ポメで特に重要な脱毛症 X・気管虚脱・膝蓋骨脱臼・歯科疾患を解説。",
        "category": "🐕 犬種別",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "🐕 ポメラニアンで特に重要な遺伝病",
                "body": (
                    "・<strong>脱毛症 X (Alopecia X)</strong> — 体幹両側の対称性脱毛。原因遺伝子は研究中だが家族性顕著。<br>"
                    "・<strong>膝蓋骨脱臼</strong> — 小型犬として高頻度。<br>"
                    "・<strong>気管虚脱</strong> — 小型犬特有のリスク。<br>"
                    "・<strong>低血糖</strong> — 子犬期の重大リスク。<br>"
                    "・<strong>歯科疾患</strong> — 早期歯石・歯周病。<br>"
                    "・<strong>白内障 (HSF4)</strong>・<strong>進行性網膜萎縮症</strong><br>"
                    "・<strong>水頭症</strong>"
                ),
            },
            {
                "heading": "🎨 ポメラニアンの多彩な毛色",
                "body": (
                    "ポメは犬種内で最も多彩な毛色を持つ犬種の1つ:<br>"
                    "<strong>標準色</strong>: オレンジ (ay)・クリーム (e/e)・セーブル・ブラック&タン (at/at)・ブラック・チョコ (bb)・ブルー (dd)<br>"
                    "<strong>パーティカラー (sp/sp)</strong>・<strong>マール (M/_)</strong> — マール × マール厳禁<br>"
                    "<strong>I 座位 / KITLG</strong>: オレンジの濃度に影響"
                ),
            },
            {
                "heading": "💡 ポメ飼育の推奨",
                "body": (
                    "・<strong>子犬期の低血糖管理</strong> — 頻繁な少量給餌<br>"
                    "・<strong>気管虚脱予防</strong>: 首輪ではなくハーネス推奨<br>"
                    "・<strong>歯磨き習慣化</strong><br>"
                    "・<strong>PRA / HSF4 検査</strong><br>"
                    "・脱毛症 X 発症時は獣医師に相談（去勢・避妊で改善する場合あり）"
                ),
            },
        ],
        "related_disease_slugs": ["hereditary-cataract", "progressive-rod-cone"],
        "related_trait_slugs": ["a-locus", "e-locus", "b-locus", "d-locus", "m-locus", "s-locus", "i-locus", "kitlg"],
    },
    {
        "slug": "chihuahua-genetic-health-guide",
        "title": "チワワの遺伝子検査・健康ガイド",
        "summary": "チワワで特に重要な水頭症・膝蓋骨脱臼・低血糖・歯科疾患を解説。",
        "category": "🐕 犬種別",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "🐕 チワワで特に重要な遺伝病",
                "body": (
                    "・<strong>水頭症</strong> — チワワで最頻発。モレラ（前頭骨の閉鎖不全）を伴うことも。<br>"
                    "・<strong>膝蓋骨脱臼</strong> — 小型犬として高頻度。<br>"
                    "・<strong>低血糖</strong> — 子犬期の重大リスク。<br>"
                    "・<strong>歯科疾患</strong> — 乳歯遺残・歯周病。<br>"
                    "・<strong>僧帽弁疾患 (MVD)</strong> — 中高齢期から。<br>"
                    "・<strong>気管虚脱</strong><br>"
                    "・<strong>進行性網膜萎縮症 (PRA)</strong>"
                ),
            },
            {
                "heading": "🎨 チワワの毛色・被毛",
                "body": (
                    "<strong>被毛タイプ</strong>: スムース (短毛・L/L)、ロング (l/l)<br>"
                    "<strong>毛色</strong>: 多彩 — フォーン (ay)・チョコ (bb)・ブルー (dd)・ブラックタン (at/at)・パーティ (sp/sp)・マール (M/_)<br>"
                    "<strong>マール × マール厳禁</strong>（チワワでも FCI 非認可かつ健康リスク）"
                ),
            },
            {
                "heading": "💡 チワワ飼育の推奨",
                "body": (
                    "・<strong>子犬期の低血糖管理</strong><br>"
                    "・<strong>頭部の保護</strong> — モレラ部の打撲リスクに注意<br>"
                    "・<strong>体重管理</strong> — 過体重は膝・気管に負担<br>"
                    "・<strong>歯科ケア</strong>を生涯にわたって継続<br>"
                    "・心臓検査（MVD スコアリング）を成犬期から"
                ),
            },
        ],
        "related_disease_slugs": ["progressive-rod-cone", "hereditary-cataract"],
        "related_trait_slugs": ["a-locus", "b-locus", "d-locus", "m-locus", "s-locus", "l-locus"],
    },
    {
        "slug": "yorkshire-terrier-genetic-health-guide",
        "title": "ヨークシャー・テリアの遺伝子検査・健康ガイド",
        "summary": "ヨーキーで特に重要な門脈体循環シャント・気管虚脱・歯科疾患を解説。",
        "category": "🐕 犬種別",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "🐕 ヨーキーで特に重要な遺伝病",
                "body": (
                    "・<strong>門脈体循環シャント (PSS)</strong> — ヨーキーで高頻度。発育不良・神経症状。<br>"
                    "・<strong>気管虚脱</strong> — 小型犬特有のリスクとして頻発。<br>"
                    "・<strong>膝蓋骨脱臼</strong><br>"
                    "・<strong>レッグ・カルベ・ペルテス病</strong> — 大腿骨頭壊死。小型犬で頻発。<br>"
                    "・<strong>歯科疾患</strong> — 乳歯遺残・早期歯周病。<br>"
                    "・<strong>低血糖</strong>（子犬期）<br>"
                    "・<strong>進行性網膜萎縮症</strong>"
                ),
            },
            {
                "heading": "🎨 ヨーキーの毛色・被毛",
                "body": (
                    "<strong>標準色</strong>: ブラック&タン → 成犬で『スチール・ブルー&タン』に変化（G座位の退色作用）<br>"
                    "<strong>被毛</strong>: 長毛 (l/l)・直毛（カーリーではない）<br>"
                    "<strong>ファーニシング (F/F)</strong>: 顔毛・髭が豊富<br>"
                    "子犬期は黒主体、生後 1〜3 年で背中部が退色してブルーシルバーに変化していきます。"
                ),
            },
            {
                "heading": "💡 ヨーキー飼育の推奨",
                "body": (
                    "・<strong>胆汁酸検査</strong>を子犬期に — PSS の早期発見<br>"
                    "・<strong>気管虚脱予防</strong>: ハーネス使用、肥満予防<br>"
                    "・<strong>歯科ケア</strong>を生涯継続<br>"
                    "・<strong>子犬期の低血糖管理</strong><br>"
                    "・成犬期の段階的な毛色変化は正常"
                ),
            },
        ],
        "related_disease_slugs": ["progressive-rod-cone"],
        "related_trait_slugs": ["a-locus", "g-locus", "l-locus", "furnishings"],
    },
    {
        "slug": "pug-genetic-health-guide",
        "title": "パグの遺伝子検査・健康ガイド",
        "summary": "パグ特有の壊死性髄膜脳炎 (PDE)・短頭種疾患・皮膚疾患を解説。",
        "category": "🐕 犬種別",
        "reading_time": "5 分",
        "sections": [
            {
                "heading": "🐕 パグで特に重要な遺伝病",
                "body": (
                    "・<strong>壊死性髄膜脳炎 (NME / PDE — Pug Dog Encephalitis)</strong> — パグ特有の致死性脳炎。発作・進行性神経症状。<br>"
                    "・<strong>短頭種気道症候群 (BOAS)</strong> — 呼吸器負担、麻酔リスク。<br>"
                    "・<strong>椎骨形成異常 (Hemivertebrae)</strong> — 巻き尾犬種の脊椎奇形。<br>"
                    "・<strong>パグ脊髄症 (Pug Myelopathy)</strong> — 後肢失調。<br>"
                    "・<strong>色素性角膜炎 (Pigmentary Keratitis)</strong> — 角膜への色素沈着。<br>"
                    "・<strong>乾燥性角結膜炎 (KCS / ドライアイ)</strong><br>"
                    "・<strong>皮膚襞皮膚炎</strong>"
                ),
            },
            {
                "heading": "🎨 パグの毛色",
                "body": (
                    "FCI 認可色は4色のみ:<br>"
                    "<strong>フォーン</strong>: ay + Em（黒マスク）<br>"
                    "<strong>アプリコット</strong>: ay + KITLG/I 座位による濃淡<br>"
                    "<strong>シルバー</strong>: 一部に G 座位の退色<br>"
                    "<strong>ブラック</strong>: KB_ + E_ または a/a 劣性ブラック<br>"
                    "全パグは E座位の Em バリアントを持ち、黒マスクが特徴。"
                ),
            },
            {
                "heading": "💡 パグ飼育の推奨",
                "body": (
                    "・<strong>NME 検査</strong>（パグ協会推奨マーカー）<br>"
                    "・<strong>BOAS スコアリング</strong>を実施した個体での繁殖<br>"
                    "・<strong>椎骨レントゲン</strong>で奇形確認<br>"
                    "・<strong>角膜・眼科検診</strong>を定期的に<br>"
                    "・<strong>皮膚襞の清拭</strong>を日常的に<br>"
                    "・高温多湿環境は厳禁（熱中症リスク大）"
                ),
            },
        ],
        "related_disease_slugs": ["necrotizing-meningoencephalitis", "hereditary-cataract"],
        "related_trait_slugs": ["a-locus", "em-locus", "k-locus", "g-locus", "kitlg"],
    },
    {
        "slug": "siberian-husky-genetic-health-guide",
        "title": "シベリアン・ハスキーの遺伝子検査・健康ガイド",
        "summary": "ハスキーで重要な PRA・白内障・甲状腺機能低下症と、ALX4 ブルーアイ等の毛色遺伝学を解説。",
        "category": "🐕 犬種別",
        "reading_time": "6 分",
        "sections": [
            {
                "heading": "🐕 シベリアン・ハスキーで特に重要な遺伝病",
                "body": (
                    "・<strong>進行性網膜萎縮症 (X-linked PRA)</strong> — ハスキーで報告される X 連鎖性 PRA。<br>"
                    "・<strong>遺伝性白内障 (HSF4)</strong> — 若年性白内障。<br>"
                    "・<strong>甲状腺機能低下症</strong> — 自己免疫性、中高齢期から。<br>"
                    "・<strong>緑内障</strong> — 一部ラインで報告。<br>"
                    "・<strong>亜鉛応答性皮膚症</strong> — 北方犬種に多い皮膚異常。亜鉛吸収障害。<br>"
                    "・<strong>多発性血管症</strong>・<strong>てんかん</strong>"
                ),
            },
            {
                "heading": "🎨 ハスキーの毛色・目色の遺伝学",
                "body": (
                    "ハスキーは目色・毛色のバリエーションが非常に豊富:<br>"
                    "<strong>毛色</strong>: ブラック&ホワイト・アグーチ&ホワイト (aw)・グレー&ホワイト・レッド&ホワイト (e/e)・サブル<br>"
                    "<strong>パイド模様</strong>: S 座位 sp/sp が固定<br>"
                    "<strong>ALX4 ブルーアイ</strong>: ハスキーの青目・オッドアイの主原因。マールや白斑経由とは別機構。<br>"
                    "<strong>Domino (MC1R Ea)</strong>: 顔マスクの色抜けパターンの一因。<br>"
                    "<strong>マールは非認可</strong>: ハスキーには本来 M 座位は存在しない。"
                ),
            },
            {
                "heading": "💡 ハスキー飼育の推奨",
                "body": (
                    "・<strong>HSF4 / PRA 検査</strong><br>"
                    "・<strong>甲状腺機能検査</strong>を定期的に<br>"
                    "・<strong>ALX4 ブルーアイ検査</strong>: 青目の遺伝メカニズム把握用<br>"
                    "・運動量の確保（運動犬種としての本能）<br>"
                    "・夏期の高温対策（北方犬種のため熱に弱い）<br>"
                    "・<strong>マールの混入が疑われる個体</strong>は M 座位検査推奨"
                ),
            },
        ],
        "related_disease_slugs": ["hereditary-cataract", "progressive-rod-cone", "glaucoma"],
        "related_trait_slugs": ["a-locus", "e-locus", "s-locus", "alx4", "domino", "m-locus"],
    },
    {
        "slug": "australian-shepherd-genetic-health-guide",
        "title": "オーストラリアン・シェパードの遺伝子検査・健康ガイド",
        "summary": "オーシーで重要な MDR1・CEA・PRA・マール、Cocoa（HPS3）等の毛色遺伝学を解説。",
        "category": "🐕 犬種別",
        "reading_time": "7 分",
        "sections": [
            {
                "heading": "🐕 オーシーで特に重要な遺伝病",
                "body": (
                    "・<strong>MDR1 (多剤耐性)</strong> — オーシーで最頻発の薬剤過敏症遺伝子。イベルメクチン・ロペラミド等の禁忌。<br>"
                    "・<strong>CEA (コリーアイ症候群)</strong> — 眼球発達異常。<br>"
                    "・<strong>進行性網膜萎縮症 (prcd-PRA)</strong><br>"
                    "・<strong>遺伝性白内障 (HSF4)</strong><br>"
                    "・<strong>変性性脊髄症 (DM)</strong> — SOD1 リスクアレル保有率高。<br>"
                    "・<strong>てんかん</strong> — 多因子。<br>"
                    "・<strong>神経セロイドリポフスチン症 (NCL)</strong>"
                ),
            },
            {
                "heading": "🎨 オーシーの多彩な毛色",
                "body": (
                    "<strong>標準色</strong>: ブラック・レッド・ブルーマール (M/m)・レッドマール<br>"
                    "<strong>+ タンポイント (at/at)・+ 白斑 (sp/sp)</strong>: 計16通り以上の組み合わせ<br>"
                    "<strong>Cocoa (HPS3)</strong>: 一部ラインで報告されている劣性チョコ。B 座位とは別。<br>"
                    "<strong>⚠️ マール × マール厳禁</strong>: M/M ダブルマールで失明・難聴のリスク。"
                ),
            },
            {
                "heading": "💡 オーシー飼育の推奨",
                "body": (
                    "・<strong>MDR1 検査必須</strong>: 全頭推奨。獣医師に必ず結果を共有<br>"
                    "・<strong>CEA / PRA / HSF4 / DM</strong> の遺伝子検査<br>"
                    "・<strong>M 座位検査</strong>: マール交配計画の必須条件<br>"
                    "・<strong>Cocoa 検査</strong>: 希少色ブリーダーは追加検討<br>"
                    "・運動量・知的刺激の確保（牧羊犬種としての本能）"
                ),
            },
        ],
        "related_disease_slugs": ["multidrug-resistance", "collie-eye-anomaly", "progressive-rod-cone", "hereditary-cataract", "degenerative-myelopathy", "neuronal-ceroid-lipofuscinosis"],
        "related_trait_slugs": ["a-locus", "k-locus", "m-locus", "s-locus", "cocoa", "b-locus"],
    },
]

# Slug ベースで guides を引けるよう辞書化
GUIDES_INDEX = {g["slug"]: g for g in GUIDES}


# 犬種ガイドの slug → 犬種表示名（JA / EN）
# 5 基礎ガイド（how-to-read 等）は犬種に紐付かないので除外。
GUIDE_BREED_NAMES = {
    "poodle-genetic-health-guide":             {"ja": "プードル", "en": "Poodle"},
    "labrador-genetic-health-guide":           {"ja": "ラブラドール", "en": "Labrador"},
    "doodle-genetic-health-guide":             {"ja": "ドゥードゥル系", "en": "Doodle breeds"},
    "shiba-genetic-health-guide":              {"ja": "柴犬", "en": "Shiba Inu"},
    "akita-genetic-health-guide":              {"ja": "秋田犬", "en": "Akita Inu"},
    "shar-pei-genetic-health-guide":           {"ja": "シャー・ペイ", "en": "Shar-Pei"},
    "chin-genetic-health-guide":               {"ja": "狆", "en": "Japanese Chin"},
    "dachshund-genetic-health-guide":          {"ja": "ダックスフンド", "en": "Dachshund"},
    "french-bulldog-genetic-health-guide":     {"ja": "フレンチブルドッグ", "en": "French Bulldog"},
    "cavalier-genetic-health-guide":           {"ja": "キャバリア", "en": "Cavalier KCS"},
    "border-collie-genetic-health-guide":      {"ja": "ボーダーコリー", "en": "Border Collie"},
    "german-shepherd-genetic-health-guide":    {"ja": "ジャーマンシェパード", "en": "German Shepherd"},
    "mini-schnauzer-genetic-health-guide":     {"ja": "ミニチュアシュナウザー", "en": "Miniature Schnauzer"},
    "golden-retriever-genetic-health-guide":   {"ja": "ゴールデン", "en": "Golden Retriever"},
    "welsh-corgi-genetic-health-guide":        {"ja": "ウェルシュ・コーギー", "en": "Welsh Corgi"},
    "pomeranian-genetic-health-guide":         {"ja": "ポメラニアン", "en": "Pomeranian"},
    "chihuahua-genetic-health-guide":          {"ja": "チワワ", "en": "Chihuahua"},
    "yorkshire-terrier-genetic-health-guide":  {"ja": "ヨークシャー・テリア", "en": "Yorkshire Terrier"},
    "pug-genetic-health-guide":                {"ja": "パグ", "en": "Pug"},
    "siberian-husky-genetic-health-guide":     {"ja": "シベリアン・ハスキー", "en": "Siberian Husky"},
    "australian-shepherd-genetic-health-guide": {"ja": "オーストラリアン・シェパード", "en": "Australian Shepherd"},
}


# 逆引きインデックス: disease_slug → [guides], trait_slug → [guides]
def _build_guide_reverse_index():
    """各疾患/形質 slug がどのガイドから参照されているかの逆引き辞書を構築。

    関連ガイドリンクを疾患/形質個別ページに表示するため。
    また、disease/trait slug → [breed dict] の逆引きも生成する。
    """
    disease_to_guides = {}
    trait_to_guides = {}
    disease_to_breeds = {}
    trait_to_breeds = {}
    for g in GUIDES:
        slug = g["slug"]
        breed = GUIDE_BREED_NAMES.get(slug)
        for d_slug in g.get("related_disease_slugs", []):
            disease_to_guides.setdefault(d_slug, []).append(g)
            if breed:
                bucket = disease_to_breeds.setdefault(d_slug, [])
                if {"slug": slug, **breed} not in bucket:
                    bucket.append({"slug": slug, **breed})
        for t_slug in g.get("related_trait_slugs", []):
            trait_to_guides.setdefault(t_slug, []).append(g)
            if breed:
                bucket = trait_to_breeds.setdefault(t_slug, [])
                if {"slug": slug, **breed} not in bucket:
                    bucket.append({"slug": slug, **breed})
    return disease_to_guides, trait_to_guides, disease_to_breeds, trait_to_breeds


GUIDES_BY_DISEASE, GUIDES_BY_TRAIT, BREEDS_BY_DISEASE, BREEDS_BY_TRAIT = _build_guide_reverse_index()


# ============================================================
# 初心者向け解説オーバーレイ（simple_explainers.py）
# ============================================================
# 専門用語に不慣れな飼い主さん向けのコンテンツ層。既存の DISEASE_KB /
# TRAIT_KB / GUIDES に oneliner / daily_impact / misconceptions / tldr /
# faq を上乗せする。テンプレートは値の有無を見て条件付きで描画する。
try:
    from simple_explainers import DISEASE_SIMPLE, TRAIT_SIMPLE, GUIDE_EXTRAS, GENETICS_TOOLTIPS
    for entry in DISEASE_KB:
        slug = entry.get("_slug")
        if slug in DISEASE_SIMPLE:
            entry["_simple"] = DISEASE_SIMPLE[slug]
    for entry in TRAIT_KB:
        slug = entry.get("_slug")
        if slug in TRAIT_SIMPLE:
            entry["_simple"] = TRAIT_SIMPLE[slug]
    for g in GUIDES:
        if g["slug"] in GUIDE_EXTRAS:
            extras = GUIDE_EXTRAS[g["slug"]]
            if "tldr" in extras:
                g["tldr"] = extras["tldr"]
            if "faq" in extras:
                g["faq"] = extras["faq"]
    HAS_SIMPLE_EXPLAINERS = True
except ImportError:
    HAS_SIMPLE_EXPLAINERS = False
    GENETICS_TOOLTIPS = {}


# 犬種文字列 → 犬種ガイド検出のためのキーワード辞書
# 「トイプードル」「ミニチュアダックスフンド」のようなサイズ違いも吸収するため
# 部分一致のキーワードと、追加で英名のキーワードも含める。
_BREED_GUIDE_KEYWORDS = {
    "poodle-genetic-health-guide":             ["プードル", "poodle"],
    "labrador-genetic-health-guide":           ["ラブラドール", "labrador"],
    "doodle-genetic-health-guide":             ["ドゥードゥル", "doodle", "ラブラドゥードル", "ゴールデンドゥードル"],
    "shiba-genetic-health-guide":              ["柴犬", "shiba"],
    "akita-genetic-health-guide":              ["秋田", "akita"],
    "shar-pei-genetic-health-guide":           ["シャー", "shar-pei", "shar pei"],
    "chin-genetic-health-guide":               ["狆", "chin"],
    "dachshund-genetic-health-guide":          ["ダックス", "dachshund"],
    "french-bulldog-genetic-health-guide":     ["フレンチ", "フレブル", "french bulldog"],
    "cavalier-genetic-health-guide":           ["キャバリア", "cavalier"],
    "border-collie-genetic-health-guide":      ["ボーダーコリー", "border collie"],
    "german-shepherd-genetic-health-guide":    ["ジャーマン", "シェパード", "german shepherd"],
    "mini-schnauzer-genetic-health-guide":     ["シュナウザー", "schnauzer"],
    "golden-retriever-genetic-health-guide":   ["ゴールデン", "golden retriever"],
    "welsh-corgi-genetic-health-guide":        ["コーギー", "corgi"],
    "pomeranian-genetic-health-guide":         ["ポメ", "pomeranian"],
    "chihuahua-genetic-health-guide":          ["チワワ", "chihuahua"],
    "yorkshire-terrier-genetic-health-guide":  ["ヨーキー", "ヨークシャー", "yorkshire"],
    "pug-genetic-health-guide":                ["パグ", "pug"],
    "siberian-husky-genetic-health-guide":     ["ハスキー", "husky"],
    "australian-shepherd-genetic-health-guide": ["オーストラリアン", "オーシー", "australian shepherd", "aussie"],
}


def detect_breed_guides(breed_strings) -> list:
    """犬種文字列のリストから該当する犬種ガイドを検出。

    PDF から抽出された breed フィールド（"POODLE (トイプードル)" 等）に対し、
    GUIDE_BREED_NAMES に対応する候補を抽出して返す。複数犬の解析時はマージ。

    Returns:
        [{"slug": "...", "ja": "...", "en": "..."}, ...]  重複排除済
    """
    if isinstance(breed_strings, str):
        breed_strings = [breed_strings]
    if not breed_strings:
        return []
    joined = " ".join(str(s) for s in breed_strings if s).lower()
    seen = set()
    result = []
    for slug, keywords in _BREED_GUIDE_KEYWORDS.items():
        if any(kw.lower() in joined for kw in keywords):
            if slug in seen:
                continue
            seen.add(slug)
            name = GUIDE_BREED_NAMES.get(slug, {})
            result.append({"slug": slug, "ja": name.get("ja", slug), "en": name.get("en", slug)})
    return result


# ============================================================
# ガイド記事の英訳オーバーレイ
# ============================================================
try:
    from guides_en import GUIDES_EN
    for g in GUIDES:
        en_data = GUIDES_EN.get(g["slug"])
        if en_data:
            g["_en"] = en_data
    HAS_EN_GUIDES = True
except ImportError:
    HAS_EN_GUIDES = False
    GUIDES_EN = {}


def get_guide_localized(slug: str, lang: str = "ja") -> Optional[dict]:
    """slug + lang からローカライズ済みガイド dict を返す。

    lang='en' で英訳が存在すれば merge した dict を返す。無ければ日本語版。
    related_*_slugs 等の言語非依存フィールドは保持。
    """
    guide = GUIDES_INDEX.get(slug)
    if not guide:
        return None
    if lang == "en" and "_en" in guide:
        merged = {**guide, **guide["_en"]}
        merged["slug"] = guide["slug"]
        merged["related_disease_slugs"] = guide.get("related_disease_slugs", [])
        merged["related_trait_slugs"] = guide.get("related_trait_slugs", [])
        return merged
    return guide


def get_guides_localized(lang: str = "ja") -> list:
    """ガイド一覧をローカライズして返す（一覧表示用）"""
    if lang != "en":
        return GUIDES
    result = []
    for g in GUIDES:
        if "_en" in g:
            merged = {**g, **g["_en"]}
            merged["slug"] = g["slug"]
            merged["related_disease_slugs"] = g.get("related_disease_slugs", [])
            merged["related_trait_slugs"] = g.get("related_trait_slugs", [])
            result.append(merged)
        else:
            result.append(g)
    return result


def get_trait_detail(test_name: str) -> Optional[dict]:
    """形質名から詳細解説を取得する。"""
    if not test_name:
        return None
    name_norm = _normalize_for_match(test_name)
    for entry in TRAIT_KB:
        for pattern in entry["match"]:
            if pattern.startswith("\\b") and pattern.endswith("\\b"):
                if re.search(pattern, name_norm):
                    return entry
            elif pattern in name_norm:
                return entry
    return None


def render_detail_html(detail: dict) -> str:
    """KB エントリ dict を折りたためる <details> HTML に変換。"""
    if not detail:
        return ""
    refs_html = ""
    for ref in detail.get("references", []):
        url = _h(ref.get("url", "#"))
        label = _h(ref.get("label", "リンク"))
        refs_html += f'<a href="{url}" target="_blank" rel="noopener noreferrer" class="kb-ref-link">{label} ↗</a>'

    sections = []
    for field_key, label in [
        ("summary", "📋 概要"),
        ("mechanism", "🧬 メカニズム"),
        ("symptoms", "⚠️ 症状"),
        ("phenotype", "🎨 表現型"),
        ("inheritance", "🧪 遺伝様式"),
        ("advice", "💡 アドバイス"),
    ]:
        text = detail.get(field_key)
        if text:
            sections.append(f'<div class="kb-section"><strong>{label}</strong><div>{_h(text).replace(chr(10), "<br>")}</div></div>')

    body = "".join(sections)
    if refs_html:
        body += f'<div class="kb-section kb-refs"><strong>🔗 参考リンク</strong><div>{refs_html}</div></div>'

    return f"""<details class="kb-detail">
<summary>📖 詳しい解説を見る</summary>
<div class="kb-body">
<h4 class="kb-title">{_h(detail.get("title", ""))}</h4>
{body}
</div>
</details>"""


def is_valid_health_test(test_name: str) -> bool:
    """有効な健康検査項目かどうかを判定（PDFの文字化けやゴミデータを除外）"""
    if not test_name or len(test_name) < 3:
        return False
    name_lower = test_name.lower()
    for blacklisted in HEALTH_TEST_BLACKLIST:
        if blacklisted in name_lower:
            return False
    # 文字化けチェック: 制御文字や置換文字が含まれている場合
    if re.search(r'[\ufffd\x00-\x08\x0b\x0c\x0e-\x1f]', test_name):
        return False
    return True


# ============================================================
# 既知の血統書データ
# ============================================================

KNOWN_PEDIGREES = {
    "seven": Pedigree(
        dog_name="SMASH JP SEVEN NIGHT",
        breed="POODLE (トイプードル)",
        registration="JKC-PT -32565/25",
        sex="MALE",
        dob="2025年4月14日",
        color="BLACK",
        microchip="392149002585861",
        breeder="TOSHINORI OMURA, FUJISHI",
        owner="KENTARO KAMIDE, MINAMISOMASHI",
        sire=Ancestor(position="sire", name="SMASH JP NIGHT DANCER",
                       registration="JKC-PT -41545/23", titles="CH/24.11, J.CH",
                       color="BLK", dna_number="JP006738/24", microchip="392144000844198",
                       dob="2023年4月30日"),
        dam=Ancestor(position="dam", name="SMASH JP NEZUKO",
                      registration="JKC-PT -70721/20", titles="",
                      color="WH", dna_number="JP003236/24",
                      dob="2020年10月27日"),
        ss=Ancestor(position="ss", name="SMASH JP HIKARU",
                     registration="JKC-PT -30198/22", titles="CH/23.6, C1B-J, J.CH, CH(PH1), WJW22, AAO.CH",
                     color="BLK", dna_number="JP004083/23", microchip="392144000441530"),
        sd=Ancestor(position="sd", name="SMASH JP JEG VIL AT VI",
                     registration="JKC-PT -70711/20", titles="CH/21.11",
                     color="BLK", dna_number="JP007839/21", microchip="392144000312770"),
        ds=Ancestor(position="ds", name="SMASH JP PERFECT HUMAN",
                     registration="JKC-PT -33197/18", titles="CH/17.4",
                     color="WH", dna_number="JP002506/17", microchip="392148014113831"),
        dd=Ancestor(position="dd", name="SMASH JP TOGENYAN",
                     registration="JKC-PT -12866/15", titles="CH/18.7",
                     color="WH", dna_number="JP004270/16"),
        sss=Ancestor(position="sss", name="SMASH JP BLINDING LIGHTS",
                      registration="JKC-PT -47888/20", titles="CH/21.5",
                      color="BLK", dna_number="JP003001/21", microchip="392144000310643"),
        ssd=Ancestor(position="ssd", name="SMASH JP TIK TOK",
                      registration="JKC-PT -42345/18", titles="INT.CH, CH/19.5, GCH(PH1), AAO.CH, SEA.CH",
                      color="BLK", dna_number="JP002875/18", microchip="392144000158862"),
        sds=Ancestor(position="sds", name="SMASH JP ONE OF US",
                      registration="JKC-PT -00866/19", titles="C.I.B., CH/19.11",
                      color="BR", dna_number="JP007567/19", microchip="392145000477714"),
        sdd=Ancestor(position="sdd", name="SMASH JP LONDON WIND",
                      registration="JKC-PT -15146/18", titles="",
                      color="BLK", microchip="392144000313005"),
        dss=Ancestor(position="dss", name="SMASH JP MOON WALK",
                      registration="JKC-PT -67987/06", titles="INT.CH, CH/07.11, CH(AM, SWE, CRO), G.CH(AM), WW11",
                      color="WH", dna_number="JP013774/07", microchip="392143000042098"),
        dsd=Ancestor(position="dsd", name="SMASH JP CINDERELA",
                      registration="JKC-PT -37596/14", titles="CH/15.5",
                      color="WH", dna_number="JP003076/15", microchip="392141000628131"),
        dds=Ancestor(position="dds", name="SMASH JP BLIZARD",
                      registration="JKC-PT -02878/13", titles="CH/14.1",
                      color="WH", dna_number="JP000541/14", microchip="392141000595003"),
        ddd=Ancestor(position="ddd", name="SMASH JP THE POWER OF DREAMS",
                      registration="JKC-PT -21576/11", titles="CH/12.1",
                      color="WH", dna_number="JP000419/12"),
    ),
}


# ████████████████████████████████████████████████████████████
# PART 1: Orivet PDF 解析エンジン
# ████████████████████████████████████████████████████████████

def extract_all_text(pdf_path: str) -> str:
    """PDFから全ページのテキストを抽出（テーブル抽出も併用）

    pdfplumber の extract_text() はテーブルセル境界で文字が
    失われることがある（例: at/at → —）。extract_tables() で
    セル単位のデータも取得し、テキスト末尾に追加することで
    後処理での遺伝子型補完を可能にする。
    """
    texts = []
    table_texts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
            # テーブルデータもセル単位で抽出して補完用テキストに追加
            try:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if row:
                                cells = [str(c).strip() for c in row if c]
                                if cells:
                                    table_texts.append(" | ".join(cells))
            except Exception:
                pass  # テーブル抽出失敗時はextract_textのみで続行

    result = "\n\n".join(texts)
    if table_texts:
        result += "\n\n--- TABLE DATA ---\n" + "\n".join(table_texts)
    return result


def parse_animal_details(text: str) -> dict:
    """動物の基本情報を抽出"""
    info = {}
    patterns = {
        "registered_name": r"Registered\s+Name\s*:?\s*(.+?)(?:\n|$)",
        "pet_name": r"Pet\s+Name\s*:?\s*(.+?)(?:\n|$)",
        "breed": r"Breed\s*:?\s*:?\s*(.+?)(?:\n|$)",
        "microchip": r"Microchip\s+Number\s*:?\s*(\d[\d\s]*\d)",
        "sex": r"Sex\s*:?\s*:?\s*(.+?)(?:\n|$)",
        "dob": r"Date\s+of\s+Birth\s*:?\s*(.+?)(?:\n|$)",
        "colour": r"Colour\s*:?\s*(.+?)(?:\n|$)",
        "case_number": r"Case\s+Number\s*:?\s*(\S+)",
        "owner_name": r"(?:Owner|Name)\s*:?\s*([\w\s]+Kamide|[\w\s]+(?:san|sama))",
        "test_date": r"Date\s+of\s+Test\s*:?\s*(.+?)(?:\n|$)",
        "test_requested": r"Test\s+Requested\s*:?\s*(.+?)(?:\n|$)",
        "approved_collection": r"Approved\s+Collection\s*(?:Method)?\s*:?\s*(Yes|No)",
        "sample_type": r"Sample\s+Type\s*:?\s*(\S+)",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            val = match.group(1).strip()
            val = re.sub(r'\s+', ' ', val).strip()
            if val and val.lower() not in (':', ''):
                info[key] = val

    if "pet_name" not in info:
        m = re.search(r"Animal\s+Name\s*:?\s*(.+?)(?:\n|$)", text, re.IGNORECASE)
        if m:
            info["pet_name"] = m.group(1).strip()

    if "owner_name" not in info:
        m = re.search(r"Name\s*:\s*(.+?)(?:\n|$)", text)
        if m:
            info["owner_name"] = m.group(1).strip()

    return info


def classify_result(result_text: str) -> str:
    """結果テキストからステータスを分類"""
    text_upper = result_text.upper()
    if "POSITIVE (P/P)" in text_upper or "TWO COPIES" in text_upper:
        return "positive"
    elif "CARRIER (P/N)" in text_upper or "ONE COPY OF THE" in text_upper:
        return "carrier"
    elif "NORMAL (N/N)" in text_upper or "NO VARIANT DETECTED" in text_upper:
        return "normal"
    elif "POSITIVE HETEROZYGOUS" in text_upper:
        return "carrier"
    elif "POSITIVE" in text_upper and "P/P" in text_upper:
        return "positive"
    elif "CARRIER" in text_upper:
        return "carrier"
    elif "NORMAL" in text_upper:
        return "normal"
    else:
        return "trait"


def get_japanese_name(test_name: str) -> str:
    """検査名から日本語名を取得"""
    name_lower = test_name.lower()
    for key, jp_name in TEST_NAME_JP.items():
        if key in name_lower:
            return jp_name
    return ""


def get_category_jp(category: str) -> str:
    """カテゴリーの日本語名を取得"""
    cat_lower = category.lower()
    for key, jp_name in CATEGORY_JP.items():
        if key in cat_lower:
            return jp_name
    return category


def extract_genotype(result_text: str, test_name: str = "") -> str:
    """結果テキストから遺伝子型を抽出（検査名に応じた優先順位付き）"""
    name_lower = test_name.lower() if test_name else ""

    # 検査名に特化した遺伝子型パターン（優先）
    specific_patterns = {
        "a locus": [r'(at/at|ay/at|ay/ay|a/a|aw/at|aw/aw|ay/aw)'],
        "b locus": [r'(Bb|BB|bb)\b'],
        "d locus": [r'(D/D|D/d|d/d)\b'],
        "dilute": [r'(D/D|D/d|d/d)\b'],
        "e locus": [r'(E/e|e/e|E/E|Em/E|Em/e)\b'],
        "em ": [r'(En/En|EM/EM|EM/En|EM/e)'],
        "mc1r": [r'(En/En|EM/EM|EM/En|EM/e)'],
        "melanistic mask": [r'(En/En|EM/EM|EM/En|EM/e)'],
        "k locus": [
            r'(KB\s*/\s*KB|K/K|KB\s*/\s*ky|KB\s*/\s*kbr|ky\s*/\s*ky|kbr\s*/\s*ky|kbr\s*/\s*kbr)',
            # OCR連結テキストで "ONE COPY DOMINANT BLACK (KB)" + "(ky)" のようなパターン
            r'DOMINANT\s+BLACK\s*\(KB\).*?\b(ky|kbr)\b',
            # "(ky )" 単独 — K Locus コンテキスト内
            r'\b(KB|ky|kbr)\s*\)',
        ],
        "m locus": [r'(m/m|M/m|M/M)'],
        "merle": [r'(m/m|M/m|M/M)'],
        "curly": [r'(Cu/Cu|Cu/N)', r'(N/N)'],
        "furnishings": [r'(F/F|F/f|f/f)'],
        "rspo2": [r'(F/F|F/f|f/f)'],
        "pied": [r'(sp/sp|S/sp|S/S)'],
        "brown tyrp1": [r'(BL/BL|BL/bs|bs/bs)'],
        "tyrp1": [r'(BL/BL|BL/bs|bs/bs)'],
        "cdpa": [r'\b([PN])/([PN])\b'],
        "chondrodysplasia": [r'\b([PN])/([PN])\b'],
    }

    # 検査名に特化したパターンを優先的に試行
    matched_specific = False
    for key, patterns in specific_patterns.items():
        if key in name_lower:
            matched_specific = True
            for p in patterns:
                m = re.search(p, result_text, re.IGNORECASE)
                if m:
                    if m.lastindex and m.lastindex >= 2:
                        return f"{m.group(1)}/{m.group(2)}"
                    # スペースを除去して正規化（"KB / ky" → "KB/ky"）
                    return re.sub(r'\s*/\s*', '/', m.group(1)).strip()
            break

    # 検査名固有パターンに一致するキーがあったが遺伝子型が取れなかった場合、
    # 汎用フォールバックで他の検査の遺伝子型を誤取得しないよう空を返す
    if matched_specific:
        return ""

    # 汎用パターン（フォールバック — 検査名が未知の場合のみ）
    m = re.search(r'\b([PN])/([PN])\b', result_text)
    if m:
        return f"{m.group(1)}/{m.group(2)}"

    generic_patterns = [
        r'(at/at|ay/at|ay/ay|a/a|aw/at)',
        r'(Bb|BB|bb)\b',
        r'(D/D|D/d|d/d)\b',
        r'(E/e|e/e|E/E|Em/E|Em/e)\b',
        r'(En/En|EM/EM)',
        r'(KB/KB|K/K|KB/ky|KB/kbr|ky/ky|kbr/ky)',
        r'(m/m|M/m|M/M)',
        r'(Cu/Cu|Cu/N|N/N)',
        r'(F/F|F/f|f/f)',
        r'(sp/sp|S/sp|S/S)',
        r'(BL/BL|BL/bs|bs/bs)',
    ]
    for p in generic_patterns:
        m = re.search(p, result_text, re.IGNORECASE)
        if m:
            return m.group(1)

    m = re.search(r'^([A-Za-z/\[\]\d\s]{1,30})\s*[-–]', result_text)
    if m:
        return m.group(1).strip()

    return ""


def parse_health_tests(text: str) -> list:
    """健康検査結果を解析"""
    results = []

    category_headers = [
        ("Musculoskeletal", r"Musculoskeletal"),
        ("Haemolymphatic", r"Haemolymphatic"),
        ("Nervous system / Neurologic", r"Nervous\s+system|Neurologic"),
        ("Metabolic", r"Metabolic"),
        ("Ophthalmologic", r"Ophthalmologic"),
        ("Cardiorespiratory", r"Cardiorespiratory"),
        ("Cardiovascular", r"Cardiovascular"),
        ("Dermatologic", r"Dermatologic"),
        ("Immunological", r"Immunological"),
        ("Respiratory", r"Respiratory"),
        ("Urogenital", r"Urogenital"),
    ]

    lines = text.split('\n')
    current_category = "Trait"
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        for cat_name, cat_pattern in category_headers:
            if re.search(cat_pattern, line, re.IGNORECASE):
                current_category = cat_name
                break

        if "Trait (Associated with Phenotype)" in line:
            current_category = "Trait"

        if re.search(r'NORMAL\s*\(N/N\)|CARRIER\s*\(P/N\)|POSITIVE\s*\(P/P\)', line, re.IGNORECASE):
            m = re.match(r'(.+?)\s+((?:NORMAL|CARRIER|POSITIVE)\s*\([PN]/[PN]\).+)', line, re.IGNORECASE)
            if m:
                test_name = m.group(1).strip()
                result_text = m.group(2).strip()
            else:
                test_name = lines[i-1].strip() if i > 0 else ""
                result_text = line.strip()

            test_name = sanitize_text(test_name)
            result_text = sanitize_text(result_text)
            test_name = re.sub(r'^[\s\uf0b7\u2022\u25cf]+', '', test_name)
            test_name = re.sub(r'\s+', ' ', test_name).strip()

            if test_name and len(test_name) > 2 and is_valid_health_test(test_name):
                status = classify_result(result_text)
                genotype = extract_genotype(result_text, test_name)
                jp_name = get_japanese_name(test_name)
                cat_jp = get_category_jp(current_category)

                result = TestResult(
                    category=cat_jp,
                    test_name=test_name,
                    genotype=genotype,
                    result_text=result_text,
                    status=status,
                    japanese_name=jp_name,
                )

                if current_category != "Trait":
                    results.append(result)

        i += 1

    # 重複除去（extract_text + extract_tables で同じ結果が2回取得される場合がある）
    seen = set()
    unique_results = []
    for r in results:
        key = (r.test_name, r.status, r.genotype)
        if key not in seen:
            seen.add(key)
            unique_results.append(r)

    return unique_results


def parse_trait_results_from_text(text: str) -> list:
    """形質（毛色）結果をテキストから解析"""
    results = []

    trait_items = [
        ("A Locus (Agouti)", r"A\s+Locus\s*\(Agouti\)"),
        ("B Locus (Brown)", r"B\s+Locus\s*[-–(]?\s*(?:Brown|Bd|Bs|Bc|Various)"),
        ("Chondrodysplasia (CDPA)", r"Chondrodysplasia\s*\(CDPA\)"),
        ("Curly Coat/Hair Variant 1", r"Curly\s+Coat"),
        ("D (Dilute) Locus", r"D\s*\(?Dilute\)?\s+Locus"),
        ("E Locus (Cream/Red/Yellow)", r"E\s+Locus\s*[-–]?\s*\(?Cream"),
        ("EM (MC1R) - Melanistic Mask", r"EM\s*\(MC1R\)|Melanistic\s+Mask"),
        ("Furnishings (RSPO2)", r"Furnishings\s*\(RSPO2\)"),
        ("K Locus (Dominant Black)", r"K\s+Locus\s*\(Dominant"),
        ("M Locus (Merle/Dapple)", r"M\s+Locus\s*\(Merle"),
        ("Pied", r"Pied\b"),
        ("Brown TYRP1", r"Brown\s+TYRP1"),
        ("Improper Coat", r"Improper\s+Coat"),
        ("Coat Length", r"Coat\s+Length"),
    ]

    lines = text.split('\n')

    for test_name, pattern in trait_items:
        for i, line in enumerate(lines):
            if re.search(pattern, line, re.IGNORECASE):
                result_text = line
                for j in range(1, 3):
                    if i + j < len(lines):
                        next_line = lines[i + j].strip()
                        # 他の形質・健康検査のヘッダー行は取り込まない
                        if next_line and not re.search(
                            r'^[A-Z]\s+Locus|^[ABDEKMS]\s*\(|^Breed|^Owner|^Microchip'
                            r'|^Pied|^Brown\s+TYRP|^Curly\s+Coat|^Furnish|^Chondro'
                            r'|^Improper|^Coat\s+Length|^EM\s*\(MC1R\)|^Melanistic'
                            r'|LOCUS\]|^\d+\.\s',
                            next_line, re.IGNORECASE
                        ):
                            result_text += " " + next_line
                        else:
                            break

                # 同一行に他の検査項目が連結されている場合、手前で切り取る
                # 例: "D/D - NORMAL...E座位 (エクステンション) E Locus..." → D Locusの部分のみ
                other_trait_boundary = re.compile(
                    r'(?:A座位|B座位|D座位|E座位|EM座位|K座位|M座位|パイド|ブラウン\s*TYRP|ファーニシング|'
                    r'巻き毛|軟骨異形成|Chondrodysplasia|Curly\s+Coat|Furnishings|'
                    r'(?<![A-Z])A\s+Locus|(?<![A-Z])B\s+Locus|(?<![A-Z])D\s*\(Dilute\)|(?<![A-Z])E\s+Locus|'
                    r'EM\s*\(MC1R\)|(?<![A-Z])K\s+Locus|(?<![A-Z])M\s+Locus|'
                    r'(?<![A-Za-z])Pied(?![A-Za-z])|Brown\s+TYRP1)',
                    re.IGNORECASE
                )
                # result_textの先頭（自分自身のマッチ）以降で、次の検査項目の開始位置を探す
                first_match = re.search(pattern, result_text, re.IGNORECASE)
                search_start = first_match.end() if first_match else 0
                boundary_match = other_trait_boundary.search(result_text, search_start)
                if boundary_match:
                    result_text = result_text[:boundary_match.start()].strip()

                result_clean = re.sub(pattern, '', result_text, flags=re.IGNORECASE).strip()
                result_clean = re.sub(r'^[\s\-–:]+', '', result_clean).strip()

                genotype = extract_genotype(result_text, test_name)
                status = classify_result(result_text)
                if status == "normal" and test_name != "Chondrodysplasia (CDPA)":
                    status = "trait"

                jp_name = get_japanese_name(test_name)

                results.append(TestResult(
                    category="形質（毛色・外見）",
                    test_name=test_name,
                    genotype=genotype,
                    result_text=result_clean if result_clean else result_text,
                    status=status,
                    japanese_name=jp_name,
                ))
                break

    # --- 後処理: 遺伝子型が空の項目を全文から補完 ---
    # PDF抽出でセル境界が崩れ、遺伝子型が隣の検査項目テキストに
    # 紛れ込んでいるケースに対応
    full_text = text  # 全文を使って再検索
    for r in results:
        if r.genotype:
            continue
        genotype_found = extract_genotype(full_text, r.test_name)
        if genotype_found:
            r.genotype = genotype_found

    # A Locus 特別処理: extract_text() でセル境界の at/at 等が
    # 失われるケースに対応。結果テキストのキーワードから推論する
    for r in results:
        if "a locus" in r.test_name.lower() and not r.genotype:
            # 全文から A Locus 遺伝子型を再検索
            m = re.search(r'\b(at/at|ay/at|ay/ay|a/a|aw/at|aw/aw|ay/aw)\b', full_text, re.IGNORECASE)
            if m:
                r.genotype = m.group(1)
            else:
                # 結果テキストのキーワードから遺伝子型を推論
                rt = r.result_text.upper() if r.result_text else ""
                ft = full_text.upper()
                # "TAN POINT" / "PHANTOM" → at/at
                if re.search(r'TAN\s*POINT|PHANTOM', rt) or re.search(r'A\s*LOCUS.*TAN\s*POINT|A\s*LOCUS.*PHANTOM', ft):
                    r.genotype = "at/at"
                # "SABLE" → ay/ay or ay/at
                elif re.search(r'SABLE', rt):
                    if re.search(r'CARRIER|CARRIES', rt):
                        r.genotype = "ay/at"
                    else:
                        r.genotype = "ay/ay"
                # "RECESSIVE BLACK" → a/a
                elif re.search(r'RECESSIVE\s+BLACK', rt):
                    r.genotype = "a/a"

    # K Locus 特別処理: 全文から KB/ky パターンを探す
    for r in results:
        if "k locus" in r.test_name.lower() and (not r.genotype or len(r.genotype) <= 3):
            m = re.search(r'KB\s*/\s*(ky|kbr)', full_text, re.IGNORECASE)
            if m:
                r.genotype = f"KB/{m.group(1)}"
            elif not r.genotype:
                # "ONE COPY DOMINANT BLACK" + "ky" or "kbr" のパターン
                m = re.search(r'ONE\s+COPY\s+DOMINANT\s+BLACK.*?\b(ky|kbr)\b', full_text, re.IGNORECASE)
                if m:
                    r.genotype = f"KB/{m.group(1)}"

    return results


def parse_heterozygosity(text: str) -> Optional[float]:
    """Orivet PDF テキストからヘテロ接合率（ゲノム多様性）を抽出。

    Orivet / Embark 等の DNA 検査が報告する『ヘテロ接合率』『遺伝的多様性』を
    柔軟に拾う。表記揺れに対応:
      - "Heterozygosity: 35.2%" / "Genetic Diversity 35.2 %"
      - "ヘテロ接合率: 35.2%" / "遺伝的多様性 35.2%"
      - "Heterozygosity Rate: 0.352"（小数 → % に換算）

    見つからなければ None。0-100 の % 値（float）で返す。
    """
    if not text:
        return None

    def _to_percent(num_str, has_percent):
        try:
            v = float(num_str)
        except (TypeError, ValueError):
            return None
        # 小数表記（0.373 など、% 記号なしで 1 以下）は % に換算
        if not has_percent and v <= 1.0:
            v *= 100.0
        if v < 0 or v > 100:
            return None
        return round(v, 1)

    # 1) Orivet の正準ラベル "Heterozygosity Score: 0.373" を最優先で拾う
    #    （prose 中の「28%」やレンジの「23.4%」を誤検出しないため）
    m = re.search(
        r"heterozygosity\s*score\s*[:：]?\s*([0-9]+(?:\.[0-9]+)?)\s*(%|％)?",
        text, re.IGNORECASE,
    )
    if m:
        v = _to_percent(m.group(1), bool(m.group(2)))
        if v is not None:
            return v

    # 2) 一般ラベル（英日両方）。値はパーセント（35.2%）か小数（0.352）。
    label = (
        r"(?:heterozygosity|genetic\s+diversity|genomic\s+diversity"
        r"|ヘテロ接合率|ヘテロ接合性|遺伝的多様性|ゲノム多様性)"
    )
    # ラベル直後に許容する任意の修飾:
    #   - 括弧注記        例: 遺伝的多様性（ヘテロ接合率）
    #   - 限定語          例: Heterozygosity Rate / Diversity Score / 多様性スコア
    qualifier = (
        r"(?:\s*[（(][^）)\n]{0,20}[）)])?"
        r"(?:\s*(?:rate|score|index|value|スコア|率|指数|値))?"
    )
    pattern = re.compile(
        label + qualifier + r"\s*[:：]?\s*([0-9]+(?:\.[0-9]+)?)\s*(%|％)?",
        re.IGNORECASE,
    )
    m = pattern.search(text)
    if not m:
        return None
    return _to_percent(m.group(1), bool(m.group(2)))


def parse_heterozygosity_range(text: str):
    """Orivet PDF の犬種別『Typical range 23.4% - 32.6%』を抽出。

    犬種ごとのヘテロ接合率の標準域。個体値がこの範囲の上か下かを示すために使う。
    見つからなければ None、見つかれば (low, high) のタプル（% 値）。
    """
    if not text:
        return None
    m = re.search(
        r"typical\s+range\s*[:：]?\s*"
        r"([0-9]+(?:\.[0-9]+)?)\s*(?:%|％)?\s*[-–~〜]\s*([0-9]+(?:\.[0-9]+)?)\s*(?:%|％)?",
        text, re.IGNORECASE,
    )
    if not m:
        # 日本語表記「標準域 23.4% 〜 32.6%」も試す
        m = re.search(
            r"(?:標準域|標準範囲|典型的範囲|典型値)\s*[:：]?\s*"
            r"([0-9]+(?:\.[0-9]+)?)\s*(?:%|％)?\s*[-–~〜]\s*([0-9]+(?:\.[0-9]+)?)\s*(?:%|％)?",
            text,
        )
    if not m:
        return None
    try:
        low = float(m.group(1))
        high = float(m.group(2))
    except (TypeError, ValueError):
        return None
    if not (0 <= low <= 100 and 0 <= high <= 100) or low > high:
        return None
    return (round(low, 1), round(high, 1))


def parse_pdf(pdf_path: str) -> Optional[DogProfile]:
    """PDFファイル1つを解析してDogProfileを返す"""
    basename = os.path.basename(pdf_path)
    if "DNAP" in basename.upper() or "DNA PROFILE" in basename.upper():
        text = extract_all_text(pdf_path)
        if "ISAG Profile" in text or "DNA Profile" in text:
            if "Health Tests Reported" not in text:
                print(f"  スキップ (DNAプロファイル): {basename}")
                return None

    if "見方" in basename or "説明" in basename:
        print(f"  スキップ (ガイド): {basename}")
        return None

    print(f"  解析中: {basename}")

    text = extract_all_text(pdf_path)

    if "Genetic Summary Report" not in text and "Health Tests Reported" not in text:
        print("  → Orivet Genetic Summary Report ではありません。スキップします。")
        return None

    info = parse_animal_details(text)
    if not info.get("pet_name") and not info.get("registered_name"):
        print("  → 動物情報を検出できませんでした。スキップします。")
        return None

    health_results = parse_health_tests(text)
    trait_results = parse_trait_results_from_text(text)

    dog = DogProfile(
        pet_name=info.get("pet_name", ""),
        registered_name=info.get("registered_name", ""),
        breed=info.get("breed", ""),
        sex=info.get("sex", ""),
        dob=info.get("dob", ""),
        colour=info.get("colour", ""),
        microchip=info.get("microchip", ""),
        case_number=info.get("case_number", ""),
        owner_name=info.get("owner_name", ""),
        test_date=info.get("test_date", ""),
        test_requested=info.get("test_requested", ""),
        approved_collection=info.get("approved_collection", ""),
        sample_type=info.get("sample_type", ""),
        health_results=health_results,
        trait_results=trait_results,
        source_file=basename,
        heterozygosity=parse_heterozygosity(text),
        heterozygosity_range=(lambda r: list(r) if r else None)(parse_heterozygosity_range(text)),
    )

    print(f"  → {dog.pet_name or dog.registered_name}: 健康{len(health_results)}項目, 形質{len(trait_results)}項目")
    return dog


def is_pedigree_pdf(text: str) -> bool:
    """PDFテキストが血統書かどうかを判定"""
    # Orivet遺伝子レポートは除外
    if "Genetic Summary Report" in text or "Health Tests Reported" in text:
        return False
    # 血統書キーワードの検出
    pedigree_keywords = [
        r'JKC-PT|ジャパンケネルクラブ|JAPAN KENNEL CLUB',
        r'ALAJ|Australian Labradoodle',
        r'AKC|AMERICAN KENNEL CLUB',
        r'PEDIGREE|血統書|血統証明',
        r'SIRE.*DAM|DAM.*SIRE',
        r'G\.?\s*SIRE|G\.?\s*DAM',
        r'犬\s*名.*犬\s*種|犬\s*種.*犬\s*名',
        r'Name of Dog',
    ]
    score = sum(1 for p in pedigree_keywords if re.search(p, text, re.IGNORECASE))
    return score >= 1


def parse_pedigree_pdf(pdf_path: str) -> Optional[Pedigree]:
    """血統書PDFからPedigreeデータを抽出"""
    if not HAS_PDFPLUMBER:
        return None

    basename = os.path.basename(pdf_path)
    print(f"  血統書PDF解析中: {basename}")

    try:
        text = extract_all_text(pdf_path)
    except Exception as e:
        print(f"  → PDF読み取りエラー: {e}")
        return None

    if not text or not is_pedigree_pdf(text):
        return None

    ped = parse_pedigree_text(text)
    if ped and ped.dog_name:
        ped.source_file = basename
        print(f"  → 血統書PDF検出: {ped.dog_name}")
        return ped
    else:
        print("  → 血統書データの解析に失敗しました")
        return None


# ████████████████████████████████████████████████████████████
# PART 2: 血統書 OCR + COI 算出
# ████████████████████████████████████████████████████████████

_OCR_DOMAIN_KEYWORDS = (
    # 構造ラベル: ヒット数を増やすほどスコア優位 → 質の高い OCR
    "PEDIGREE", "JKC", "SIRE", "DAM", "BREED", "COLOR", "BIRTH",
    "KENNEL", "JAPAN", "OWNER", "BREEDER", "REGIST", "CHAMP",
    "ジャパンケネルクラブ", "犬名", "犬種", "性別", "毛色",
    "生年月日", "登録番号", "父", "母", "祖父", "祖母", "ブリーダー",
)


def _score_ocr_text(text: str) -> float:
    """OCR 結果の品質を粗くスコア化（高いほど良好）。

    - 制御文字・記号過多は減点
    - 血統書ドメインキーワードの出現でブースト
    - 単語密度（連続英字 + カタカナ / 漢字）
    """
    if not text:
        return 0.0
    s = text
    n = len(s)
    if n == 0:
        return 0.0
    # ドメインキーワード: 各 +25 点
    kw_score = sum(25 for kw in _OCR_DOMAIN_KEYWORDS if kw in s.upper() or kw in s)
    # 英字 / かな / 漢字の比率: 0–1
    letters = sum(1 for ch in s if ch.isalnum() or ("぀" <= ch <= "ヿ") or ("一" <= ch <= "鿿"))
    letter_ratio = letters / n
    # 記号過多ペナルティ
    noise = sum(1 for ch in s if ch in "~^`'\"|\\<>*=#@$%")
    noise_ratio = noise / max(n, 1)
    return kw_score + 50 * letter_ratio - 80 * noise_ratio + min(n, 2000) / 50


def _adaptive_threshold_pil(gray):
    """numpy なしの局所平均ベース適応的二値化。

    画像を 16×16 のグリッドに分割し、各セルの平均輝度より少し下を閾値に。
    画像全体に強い陰影がある写真でも文字が消えにくい。
    """
    from PIL import Image as _Image
    w, h = gray.size
    grid = gray.resize((16, 16), _Image.LANCZOS)
    grid_pixels = list(grid.getdata())  # length 256
    # セルあたりの幅 / 高さ
    cw, ch = max(w // 16, 1), max(h // 16, 1)
    binarized = gray.copy()
    px = binarized.load()
    src = gray.load()
    for cy in range(16):
        for cx in range(16):
            mean = grid_pixels[cy * 16 + cx]
            thr = max(mean - 20, 60)
            x0 = cx * cw
            x1 = (cx + 1) * cw if cx < 15 else w
            y0 = cy * ch
            y1 = (cy + 1) * ch if cy < 15 else h
            for y in range(y0, y1):
                for x in range(x0, x1):
                    px[x, y] = 255 if src[x, y] > thr else 0
    return binarized


def _ocr_preprocess_variants(img):
    """1 枚の画像から OCR 用の派生バリアントを複数生成。

    異なる前処理それぞれで OCR にかけ、スコア最良を採用する。
    変換コストの低い順に並べる（早期終了可能なため）。
    """
    from PIL import ImageEnhance, ImageFilter, ImageOps
    # 共通: グレースケール + 軽い denoise
    gray = img.convert("L")
    gray = gray.filter(ImageFilter.MedianFilter(size=3))

    variants = []

    # v1: シンプルなグレースケール（多くのケースでこれが最良）
    variants.append(("gray", gray))

    # v2: コントラスト + シャープネス強化
    enhanced = ImageEnhance.Contrast(gray).enhance(2.2)
    enhanced = ImageEnhance.Sharpness(enhanced).enhance(2.2)
    variants.append(("enhanced", enhanced))

    # v3: 局所適応的二値化（陰影のある写真に強い）
    try:
        adaptive = _adaptive_threshold_pil(enhanced)
        variants.append(("adaptive", adaptive))
    except Exception:
        pass

    # v4: 自動コントラスト + 反転判定（白背景 / 黒背景の対応）
    try:
        autocontrast = ImageOps.autocontrast(gray, cutoff=2)
        variants.append(("autocontrast", autocontrast))
        # ヒストグラムが暗側に偏っていれば反転を追加（白黒反転スキャン対応）
        hist = autocontrast.histogram()
        dark = sum(hist[:128])
        light = sum(hist[128:])
        if dark > light * 2:
            variants.append(("inverted", ImageOps.invert(autocontrast)))
    except Exception:
        pass

    return variants


def _ocr_run_passes(img, lang_pref="eng+jpn"):
    """前処理バリアント × PSM モードを総当たりし、スコア最良の text を返す。

    早期終了: 十分高スコアな出力が得られたら以降をスキップ。
    """
    ocr_timeout = 90
    # PSM (Page Segmentation Mode):
    #   6 = uniform block of text (デフォルト、多くの blood line で良好)
    #   4 = single column of variable size
    #   11 = sparse text (キャプション・短い文字列が散らばっているケース)
    psm_configs = ['--psm 6 --oem 3', '--psm 4 --oem 3', '--psm 11 --oem 3']
    best_text = ""
    best_score = 0.0

    for variant_name, v_img in _ocr_preprocess_variants(img):
        for cfg in psm_configs:
            try:
                text = pytesseract.image_to_string(v_img, lang=lang_pref, config=cfg, timeout=ocr_timeout)
            except RuntimeError:
                continue
            if not text:
                continue
            score = _score_ocr_text(text)
            if score > best_score:
                best_score = score
                best_text = text
            # 十分なドメイン語数 + 長さがあれば早期終了
            if best_score >= 250 and len(best_text) > 400:
                return best_text
    return best_text


def try_ocr(image_path: str) -> str:
    """画像からテキストを抽出（Tesseract OCR）— 多パス前処理 + スコア選択。

    特長:
      - EXIF 回転自動補正
      - 4 種類の前処理バリアント × 3 種類の PSM モードを試行
      - OCR 出力をドメインキーワード / 文字密度 / ノイズ率でスコア化
      - 最良スコアの出力を採用し、結果テキストは `_clean_ocr_text` で補正
      - 横向き写真にも対応するため、初回スコアが低い場合は 90°/270°
        回転版も試行
    """
    if not HAS_OCR:
        print("  pytesseract が未インストールです。")
        print("  pip install pytesseract Pillow")
        print("  + Tesseract OCR本体: sudo apt install tesseract-ocr tesseract-ocr-jpn")
        return ""
    try:
        from PIL import ImageOps
        img = Image.open(image_path)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")

        # EXIF 回転補正
        try:
            img = ImageOps.exif_transpose(img)
        except Exception:
            pass

        # リサイズ（Tesseract の最適解像度域へ）
        max_dim = 2200
        if max(img.size) > max_dim:
            ratio = max_dim / max(img.size)
            img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)

        # 1st pass: そのまま
        text = _ocr_run_passes(img, lang_pref="eng+jpn")
        score = _score_ocr_text(text)

        # スコアが低い場合は回転を試す（スマホ撮影で横向きになりがち）
        if score < 120:
            for angle in (90, 270, 180):
                try:
                    rot = img.rotate(angle, expand=True)
                except Exception:
                    continue
                t2 = _ocr_run_passes(rot, lang_pref="eng+jpn")
                s2 = _score_ocr_text(t2)
                if s2 > score:
                    text, score = t2, s2
                if score >= 250:
                    break

        # 日本語優先パスも 1 回だけ追加で試す（カナ・漢字主体の血統書向け）
        if score < 200:
            t3 = _ocr_run_passes(img, lang_pref="jpn+eng")
            s3 = _score_ocr_text(t3)
            if s3 > score:
                text, score = t3, s3

        return text
    except Exception as e:
        print(f"  OCRエラー: {e}")
        return ""


def detect_pedigree_format(text: str) -> str:
    """血統書のフォーマットを自動判定"""
    text = _clean_ocr_text(text)
    if re.search(r'JKC-PT|ジャパンケネルクラブ|JAPAN KENNEL CLUB', text, re.IGNORECASE):
        return "jkc"
    if re.search(r'ALAJ|Australian Labradoodle|ラブラドゥードル', text, re.IGNORECASE):
        return "alaj"
    if re.search(r'AKC|AMERICAN KENNEL CLUB', text, re.IGNORECASE):
        return "akc"
    if re.search(r'KC\b|THE KENNEL CLUB', text, re.IGNORECASE):
        return "kc"
    if re.search(r'SIRE|DAM|G\.SIRE|G\.DAM|PEDIGREE|血統書', text, re.IGNORECASE):
        return "generic"
    return "generic"


def _extract_basic_info(text: str, ped: Pedigree):
    """共通の基本情報を抽出"""
    if re.search(r'\bMALE\b|オス|牡|♂|性\s*別\s*Male', text, re.IGNORECASE):
        ped.sex = "MALE"
    elif re.search(r'\bFEMALE\b|メス|牝|♀|性\s*別\s*Female', text, re.IGNORECASE):
        ped.sex = "FEMALE"

    m = re.search(r'(\d{4}年\s*\d{1,2}月\s*\d{1,2}日)', text)
    if m:
        ped.dob = m.group(1)
    else:
        m = re.search(r'(?:生年月日|Date of Birth|DOB|D\.O\.B)\s*[:\s]*(\d{4}[/\-]\d{1,2}[/\-]\d{1,2})', text, re.IGNORECASE)
        if m:
            ped.dob = m.group(1)
        else:
            m = re.search(r'(?:生\s*年\s*月\s*日)\s*(\d{4}/\d{1,2}/\d{1,2})', text)
            if m:
                ped.dob = m.group(1)

    m = re.search(r'(?:毛\s*色|色|Color|Colour)\s*[:\s]*([A-Za-z\s]+?)(?:\n|$)', text, re.IGNORECASE)
    if m:
        ped.color = m.group(1).strip()

    m = re.search(r'(?:マイクロチップ|Microchip|MC|ID)\s*[番号:\s]*(\d{10,15})', text, re.IGNORECASE)
    if m:
        ped.microchip = m.group(1)

    m = re.search(r'(?:所\s*有\s*者|Owner)\s*[:\s]*(.+?)(?:\n|$)', text, re.IGNORECASE)
    if m:
        ped.owner = m.group(1).strip()

    m = re.search(r'(?:繁殖者|Breeder)\s*[:\s]*(.+?)(?:\n|$)', text, re.IGNORECASE)
    if m:
        ped.breeder = m.group(1).strip()

    m = re.search(r'(?:犬\s*種|Breed)\s*[:\s]*(.+?)(?:\n|$)', text, re.IGNORECASE)
    if m:
        ped.breed = m.group(1).strip()


def _extract_jkc_dog_name_from_line(text_after: str) -> tuple:
    """JKCラベル直後のテキストから犬名・登録番号・毛色・DNA番号を抽出"""
    name = ""
    registration = ""
    color = ""
    dna = ""
    lines = [l.strip() for l in text_after.split('\n') if l.strip()]
    for line in lines:
        # ラベル行（SIRE/DAM等）自体はスキップ
        if re.match(r'^(?:SIRE|DAM|G\.\s*(?:G\.\s*)?(?:SIRE|DAM)|父|母|祖父|祖母|曾祖父|曾祖母)\b', line, re.IGNORECASE):
            continue
        # CH/タイトル + 犬名行
        if not name:
            # タイトル(CH/xx.xx)を除いた犬名を取得
            cleaned = re.sub(r'^(?:CH/[\d.]+\s*,?\s*)*', '', line).strip()
            # "SMASH JP ..." のようなケネル名パターン
            m_name = re.search(r'([A-Z][A-Z\s]+(?:JP|OF|DE|VAN|VON)\s+[A-Z][A-Z\s]+)', cleaned)
            if m_name:
                name = m_name.group(1).strip()
                continue
            # 英字3文字以上で犬名とみなす
            cleaned = re.sub(r'\s*JKC-PT.*$', '', cleaned).strip()
            cleaned = re.sub(r'\s*CD\d*\s*$', '', cleaned).strip()
            if len(cleaned) >= 3 and re.search(r'[A-Z]', cleaned):
                name = cleaned
                continue
        # JKC-PT 登録番号
        m_reg = re.search(r'(JKC-PT\s*-?\s*\d+/\d+)', line)
        if m_reg:
            registration = m_reg.group(1)
        # DNA番号
        m_dna = re.search(r'(DNA\s*JP\d+/\d+)', line)
        if m_dna:
            dna = m_dna.group(1)
        # 毛色
        m_color = re.search(r'\b(BLK|BLACK|WH|WHITE|BR|BROWN|RED|APR|APRICOT|CR|CREAM|SV|SILVER|BL|BLUE|CAFE|GOLD|FAWN|SABLE|BRINDLE|MERLE|PARTI|BEIGE)\b', line, re.IGNORECASE)
        if m_color and not color:
            color = m_color.group(1).upper()
        # 4行で十分
        if name and registration:
            break
    return name, registration, color, dna


def _parse_jkc_ancestors(text: str, ped: Pedigree):
    """JKC形式の祖先名を抽出（番号ベース + ラベルベース）"""
    # --- 方法1: 番号ベース ---
    lines = text.split('\n')
    ancestors = {}
    for line in lines:
        m = re.match(r'\s*(\d{1,2})\s*[\|\{(]?\s*(.+)', line)
        if m:
            num = int(m.group(1))
            if 1 <= num <= 14:
                name_text = m.group(2).strip()
                # ラベル部分を除去（"G.SIRE 祖父" 等）
                name_text = re.sub(r'^(?:G\.?\s*G\.?\s*)?(?:SIRE|DAM)\s*', '', name_text, flags=re.IGNORECASE).strip()
                name_text = re.sub(r'^(?:父|母|祖父|祖母|曾祖父|曾祖母)\s*', '', name_text).strip()
                name_text = re.sub(r'\s*JKC-PT.*$', '', name_text).strip()
                name_text = re.sub(r'\s*CH/\d+.*$', '', name_text).strip()
                name_text = re.sub(r'\s*CD\d*\s*$', '', name_text).strip()
                if len(name_text) > 2:
                    ancestors[num] = name_text

    pos_map = {
        1: "sire", 2: "dam", 3: "ss", 4: "sd", 5: "ds", 6: "dd",
        7: "sss", 8: "ssd", 9: "sds", 10: "sdd",
        11: "dss", 12: "dsd", 13: "dds", 14: "ddd"
    }
    for num, name in ancestors.items():
        if num in pos_map:
            setattr(ped, pos_map[num], Ancestor(position=pos_map[num], name=name))

    # --- 方法2: ラベルベース（番号ベースで取れなかったものを補完）---
    # JKC血統書のラベルパターン: "G.G.SIRE 曾祖父" / "G.SIRE 祖父" / "SIRE 父"

    # G.G.SIRE/G.G.DAM の位置を順番に割り当てるためのカウンター
    gg_sire_slots = ['sss', 'sds', 'dss', 'dds']
    gg_dam_slots = ['ssd', 'sdd', 'dsd', 'ddd']
    g_sire_slots = ['ss', 'ds']
    g_dam_slots = ['sd', 'dd']
    gg_sire_idx = 0
    gg_dam_idx = 0
    g_sire_idx = 0
    g_dam_idx = 0

    # G.G.SIRE / G.G.DAM を検索
    for m in re.finditer(r'G\.?\s*G\.?\s*SIRE\s*(?:曾祖父)?', text, re.IGNORECASE):
        pos_after = m.end()
        remaining = text[pos_after:pos_after + 300]
        name, reg, color, dna = _extract_jkc_dog_name_from_line(remaining)
        if name and gg_sire_idx < len(gg_sire_slots):
            slot = gg_sire_slots[gg_sire_idx]
            if not getattr(ped, slot):
                setattr(ped, slot, Ancestor(position=slot, name=name, registration=reg, color=color, dna_number=dna))
            gg_sire_idx += 1

    for m in re.finditer(r'G\.?\s*G\.?\s*DAM\s*(?:曾祖母)?', text, re.IGNORECASE):
        pos_after = m.end()
        remaining = text[pos_after:pos_after + 300]
        name, reg, color, dna = _extract_jkc_dog_name_from_line(remaining)
        if name and gg_dam_idx < len(gg_dam_slots):
            slot = gg_dam_slots[gg_dam_idx]
            if not getattr(ped, slot):
                setattr(ped, slot, Ancestor(position=slot, name=name, registration=reg, color=color, dna_number=dna))
            gg_dam_idx += 1

    # G.SIRE / G.DAM（G.G.を除外）
    for m in re.finditer(r'G\.?\s*SIRE\s*(?:祖父)?', text, re.IGNORECASE):
        # G.G.SIREにマッチしていないか確認
        start = m.start()
        prefix = text[max(0, start - 3):start]
        if re.search(r'G\.?$', prefix):
            continue
        pos_after = m.end()
        remaining = text[pos_after:pos_after + 300]
        name, reg, color, dna = _extract_jkc_dog_name_from_line(remaining)
        if name and g_sire_idx < len(g_sire_slots):
            slot = g_sire_slots[g_sire_idx]
            if not getattr(ped, slot):
                setattr(ped, slot, Ancestor(position=slot, name=name, registration=reg, color=color, dna_number=dna))
            g_sire_idx += 1

    for m in re.finditer(r'G\.?\s*DAM\s*(?:祖母)?', text, re.IGNORECASE):
        start = m.start()
        prefix = text[max(0, start - 3):start]
        if re.search(r'G\.?$', prefix):
            continue
        pos_after = m.end()
        remaining = text[pos_after:pos_after + 300]
        name, reg, color, dna = _extract_jkc_dog_name_from_line(remaining)
        if name and g_dam_idx < len(g_dam_slots):
            slot = g_dam_slots[g_dam_idx]
            if not getattr(ped, slot):
                setattr(ped, slot, Ancestor(position=slot, name=name, registration=reg, color=color, dna_number=dna))
            g_dam_idx += 1

    # SIRE / DAM（G.SIRE/G.DAMを除外）
    if not ped.sire:
        m = re.search(r'(?<![G.])\bSIRE\s*(?:父)?', text, re.IGNORECASE)
        if m:
            remaining = text[m.end():m.end() + 300]
            name, reg, color, dna = _extract_jkc_dog_name_from_line(remaining)
            if name:
                ped.sire = Ancestor(position="sire", name=name, registration=reg, color=color, dna_number=dna)

    if not ped.dam:
        m = re.search(r'(?<![G.])\bDAM\s*(?:母)?', text, re.IGNORECASE)
        if m:
            remaining = text[m.end():m.end() + 300]
            name, reg, color, dna = _extract_jkc_dog_name_from_line(remaining)
            if name:
                ped.dam = Ancestor(position="dam", name=name, registration=reg, color=color, dna_number=dna)


def _parse_labeled_ancestors(text: str, ped: Pedigree):
    """ラベルベースの祖先抽出（SIRE/DAM/G.SIRE/G.DAM形式 — ALAJ等）"""

    def get_name_after_label(text: str, label_end: int, max_chars: int = 200) -> str:
        remaining = text[label_end:label_end + max_chars]
        lines = [l.strip() for l in remaining.split('\n') if l.strip()]
        for line in lines:
            if re.match(r'^(?:SIRE|DAM|G\.?SIRE|G\.?DAM|GG|父犬|母犬|祖父|祖母|曾祖|犬種|性別|サイズ|毛色|生年月日|所有者)', line, re.IGNORECASE):
                continue
            cleaned = re.sub(r'^\d+\s*', '', line).strip()
            if len(cleaned) >= 3 and not re.match(r'^\d{4}[/\-]', cleaned):
                return cleaned
        return ""

    sire_match = re.search(r'(?:^|\n)\s*SIRE\b', text, re.IGNORECASE)
    dam_match = re.search(r'(?:^|\n)\s*DAM\b', text, re.IGNORECASE)

    gsire_positions = [(m.start(), m.end()) for m in re.finditer(r'G\.?SIRE\b|父方祖父|祖父犬', text, re.IGNORECASE)]
    gdam_positions = [(m.start(), m.end()) for m in re.finditer(r'G\.?DAM\b|父方祖母|祖母犬', text, re.IGNORECASE)]

    if sire_match:
        name = get_name_after_label(text, sire_match.end())
        if name:
            ped.sire = Ancestor(position="sire", name=name)

    if dam_match:
        name = get_name_after_label(text, dam_match.end())
        if name:
            ped.dam = Ancestor(position="dam", name=name)

    dam_pos = dam_match.start() if dam_match else len(text) // 2

    sire_gsires = [p for p in gsire_positions if p[0] < dam_pos]
    sire_gdams = [p for p in gdam_positions if p[0] < dam_pos]
    dam_gsires = [p for p in gsire_positions if p[0] >= dam_pos]
    dam_gdams = [p for p in gdam_positions if p[0] >= dam_pos]

    if sire_gsires:
        name = get_name_after_label(text, sire_gsires[0][1])
        if name:
            ped.ss = Ancestor(position="ss", name=name)
    if sire_gdams:
        name = get_name_after_label(text, sire_gdams[0][1])
        if name:
            ped.sd = Ancestor(position="sd", name=name)
    if dam_gsires:
        name = get_name_after_label(text, dam_gsires[0][1])
        if name:
            ped.ds = Ancestor(position="ds", name=name)
    if dam_gdams:
        name = get_name_after_label(text, dam_gdams[0][1])
        if name:
            ped.dd = Ancestor(position="dd", name=name)

    gg_sire_positions = [(m.start(), m.end()) for m in re.finditer(r'GG\.?SIRE\b|G\.G\.SIRE|曾祖父', text, re.IGNORECASE)]
    gg_dam_positions = [(m.start(), m.end()) for m in re.finditer(r'GG\.?DAM\b|G\.G\.DAM|曾祖母', text, re.IGNORECASE)]

    gg_sire_fields = ["sss", "sds", "dss", "dds"]
    gg_dam_fields = ["ssd", "sdd", "dsd", "ddd"]

    for idx, (start, end) in enumerate(gg_sire_positions):
        if idx < len(gg_sire_fields):
            name = get_name_after_label(text, end)
            if name:
                setattr(ped, gg_sire_fields[idx], Ancestor(position=gg_sire_fields[idx], name=name))

    for idx, (start, end) in enumerate(gg_dam_positions):
        if idx < len(gg_dam_fields):
            name = get_name_after_label(text, end)
            if name:
                setattr(ped, gg_dam_fields[idx], Ancestor(position=gg_dam_fields[idx], name=name))


def parse_jkc_pedigree_text(text: str) -> Optional[Pedigree]:
    """OCRテキストから血統書を解析（全フォーマット自動対応）"""
    return parse_pedigree_text(text)


# OCR ラベル誤認識: 大文字ラベル系
_OCR_LABEL_FIXES_UPPER = {
    # KENNEL CLUB 系
    'KENNE1': 'KENNEL', 'KENNE!': 'KENNEL', 'KENNEI': 'KENNEL', 'KENN3L': 'KENNEL',
    'K3NNEL': 'KENNEL', 'KENNF1': 'KENNEL',
    'C1UB': 'CLUB', 'CIUB': 'CLUB', 'CLU8': 'CLUB', 'CL[J]B': 'CLUB', 'C|UB': 'CLUB',
    # JAPAN
    'J@PAN': 'JAPAN', 'JAP@N': 'JAPAN', 'J4PAN': 'JAPAN', 'JAP4N': 'JAPAN',
    'JAFAN': 'JAPAN', 'JAPAH': 'JAPAN', '|APAN': 'JAPAN',
    # POODLE
    'P00DLE': 'POODLE', 'P0ODLE': 'POODLE', 'POOD1E': 'POODLE', 'POODIE': 'POODLE',
    'PO0DLE': 'POODLE',
    # MALE / FEMALE
    'MA1E': 'MALE', 'MAIE': 'MALE', 'MAL3': 'MALE',
    'FEMA1E': 'FEMALE', 'FEMAIE': 'FEMALE', 'FEMAL3': 'FEMALE', 'FEM4LE': 'FEMALE',
    # PEDIGREE
    'PEDI6REE': 'PEDIGREE', 'PEDIGR3E': 'PEDIGREE', 'PEDIGR EE': 'PEDIGREE',
    'PED1GREE': 'PEDIGREE', 'PEDIGR££': 'PEDIGREE', 'PEDlGREE': 'PEDIGREE',
    # SIRE / DAM
    'S1RE': 'SIRE', 'SlRE': 'SIRE', 'S|RE': 'SIRE', 'SI RE': 'SIRE',
    'DAlVl': 'DAM', 'DAlVI': 'DAM', 'D4M': 'DAM',
    # BREED / COLOR / BIRTH / OWNER / BREEDER
    'BR33D': 'BREED', 'BRE3D': 'BREED', 'BR3ED': 'BREED', 'BREEED': 'BREED',
    'C0LOR': 'COLOR', 'COL0R': 'COLOR', 'COLDR': 'COLOR',
    'B1RTH': 'BIRTH', 'BIRTII': 'BIRTH', 'B|RTH': 'BIRTH',
    'OWN3R': 'OWNER', 'OVVNER': 'OWNER',
    'BR33DER': 'BREEDER', 'BREE0ER': 'BREEDER',
    # 登録番号系
    'REGI5T': 'REGIST', 'REG1ST': 'REGIST', 'R3GIST': 'REGIST',
    'JKC PT': 'JKC-PT',
    # チャンピオン称号
    'CHAMP10N': 'CHAMPION', 'CH4MPION': 'CHAMPION', 'CHAMPlON': 'CHAMPION',
    'INT CH': 'INT.CH',
}

# 混在ラベル / 小文字混じり
_OCR_LABEL_FIXES_MIXED = {
    'Nam3': 'Name', 'Narne': 'Name', 'NarrIe': 'Name', 'Naine': 'Name',
    'D0g': 'Dog', 'D09': 'Dog',
    'Br33d': 'Breed', 'Bre3d': 'Breed', 'Brced': 'Breed',
    'C0lor': 'Color', 'Col0r': 'Color', 'Colcr': 'Color',
    'Bi rth': 'Birth', 'Birth date': 'Birth Date', 'Birthdate': 'Birth Date',
    'Sex:': 'Sex:', 'S ex': 'Sex',
    'Sire :': 'Sire:', 'Dam :': 'Dam:',
    'Reg.No': 'Reg. No', 'RegNo': 'Reg. No',
}

# 日本語ラベルの誤認識（OCR が漢字を分解 / 誤判別しがち）
_OCR_LABEL_FIXES_JA = {
    'ジャパン ケネル クラブ': 'ジャパンケネルクラブ',
    'ジヤパン': 'ジャパン',
    'ケンネル': 'ケネル',
    'クラフ': 'クラブ', 'クラフ ': 'クラブ ',
    '犬 名': '犬名',
    '犬 種': '犬種',
    '性 別': '性別',
    '毛 色': '毛色',
    '生 年 月 日': '生年月日', '生年 月日': '生年月日',
    '登録 番号': '登録番号', '登 録番号': '登録番号',
    'ブ リーダー': 'ブリーダー',
    '父 ': '父 ', '母 ': '母 ',
}


def _normalize_ocr_unicode(text: str) -> str:
    """OCR 出力の Unicode 揺らぎを正規化。

    - 全角英数 → 半角（ラベル抽出の正規表現が機能するため）
    - 制御文字・ゼロ幅文字を除去
    - 連続空白を圧縮、行末空白を削除
    - 全角コロン → 半角コロン（ラベル直後の値抽出に必要）
    """
    if not text:
        return text
    # ゼロ幅・BOM・制御文字（改行/タブは保持）
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    # Unicode 不可視文字（ZWSP/ZWJ/BOM 等）
    text = re.sub('[\u200b-\u200f\u202a-\u202e\u2060\ufeff]', '', text)
    # 全角 ASCII → 半角（ASCII レンジ + 全角スペース）
    out_chars = []
    for ch in text:
        code = ord(ch)
        if 0xff01 <= code <= 0xff5e:
            out_chars.append(chr(code - 0xfee0))
        elif code == 0x3000:
            out_chars.append(' ')
        else:
            out_chars.append(ch)
    text = ''.join(out_chars)
    # 全角コロン・括弧
    text = text.replace('：', ':').replace('，', ',').replace('．', '.')
    # 行末スペース、連続スペース
    lines = []
    for line in text.split('\n'):
        line = re.sub(r'[ \t]+', ' ', line).rstrip()
        lines.append(line)
    text = '\n'.join(lines)
    # 連続空行を 1 行に
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text


def _clean_ocr_text(text: str) -> str:
    """OCR 出力をパース可能な形に整える。

    1. Unicode 正規化（全角→半角・制御文字除去・空白圧縮）
    2. 大文字ラベルの典型的な誤認識を辞書置換（KENNE1→KENNEL 等）
    3. 大小混在ラベルの誤認識
    4. 日本語ラベルの分かち書き誤認識（"犬 名"→"犬名" 等）
    5. 空白挿入耐性パターン（"P E D I G R E E"→"PEDIGREE" 等）

    既存呼び出し側との互換性を維持するため、入力空文字列はそのまま返す。
    """
    if not text:
        return text
    text = _normalize_ocr_unicode(text)
    # 全置換
    for wrong, right in _OCR_LABEL_FIXES_UPPER.items():
        text = text.replace(wrong, right)
    for wrong, right in _OCR_LABEL_FIXES_MIXED.items():
        text = text.replace(wrong, right)
    for wrong, right in _OCR_LABEL_FIXES_JA.items():
        text = text.replace(wrong, right)
    # 空白挿入耐性: ラベル間にスペースが混入したケース
    # "P E D I G R E E" 等の単語間スペースを除去
    text = re.sub(r'\bP\s*E\s*D\s*I\s*G\s*R\s*E\s*E\b', 'PEDIGREE', text, flags=re.IGNORECASE)
    text = re.sub(r'\bS\s*I\s*R\s*E\b', 'SIRE', text)
    text = re.sub(r'\bD\s*A\s*M\b', 'DAM', text)
    text = re.sub(r'\bK\s*E\s*N\s*N\s*E\s*L\b', 'KENNEL', text)
    text = re.sub(r'\bJ\s*A\s*P\s*A\s*N\b', 'JAPAN', text)
    text = re.sub(r'\bJ\s*K\s*C\b', 'JKC', text)
    # JKC 登録番号: "JKC - PT - 12345 / 67" のような揺らぎを正規化
    text = re.sub(r'JKC\s*-?\s*PT\s*-?\s*(\d+)\s*/\s*(\d+)', r'JKC-PT-\1/\2', text)
    return text


def parse_pedigree_text(text: str) -> Optional[Pedigree]:
    """OCRテキストから血統書を解析（全フォーマット自動対応）"""
    if not text:
        return None

    text = _clean_ocr_text(text)

    ped = Pedigree()
    fmt = detect_pedigree_format(text)

    _extract_basic_info(text, ped)

    if fmt == "jkc":
        # 犬名抽出: "Name of Dog" / "犬名" の後
        m = re.search(r'(?:Name of Dog|犬\s*名)\s*\n?\s*(.+?)(?:\n|$)', text)
        if m:
            candidate = m.group(1).strip()
            # "Name of Dog" 自体の残りやラベルだけの場合はスキップして次の行を探す
            if len(candidate) < 3 or re.match(r'^(?:Name|犬名|犬\s*名)', candidate, re.IGNORECASE):
                # 次の非空行を犬名として取る
                after = text[m.end():]
                for line in after.split('\n'):
                    line = line.strip()
                    if line and len(line) >= 3 and not re.match(r'^(?:Breed|犬\s*種|登録|スマ)', line, re.IGNORECASE):
                        # カタカナ読み行はスキップ
                        if re.match(r'^[ァ-ヴー\s゛゜]+$', line):
                            continue
                        ped.dog_name = line
                        break
            else:
                ped.dog_name = candidate
        # 犬名が取れなかった場合、大文字ケネル名パターンで探す
        if not ped.dog_name:
            m_name = re.search(r'([A-Z][A-Z\s]+(?:JP|OF|DE|VAN|VON)\s+[A-Z][A-Z\s]+)\s*\n', text)
            if m_name:
                ped.dog_name = m_name.group(1).strip()

        m = re.search(r'(JKC-PT\s*-?\s*\d+/\d+)', text)
        if m:
            ped.registration = m.group(1)
        _parse_jkc_ancestors(text, ped)

    elif fmt == "alaj":
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        if lines:
            for i, line in enumerate(lines):
                if re.search(r'PEDIGREE|血統書', line, re.IGNORECASE):
                    # 犬名がPEDIGREEと同じ行にある場合
                    name_on_line = re.sub(r'\s*PEDIGREE\s*', '', line, flags=re.IGNORECASE).strip()
                    name_on_line = re.sub(r'\s*血統書\s*', '', name_on_line).strip()
                    if len(name_on_line) > 3:
                        ped.dog_name = name_on_line
                    elif i > 0:
                        ped.dog_name = lines[i-1].strip()
                    break
            if not ped.dog_name and lines:
                ped.dog_name = lines[0].strip()

        m = re.search(r'(?:登\s*録\s*番\s*号|Registration)\s*[:\s]*([A-Z]{2,}\d+)', text, re.IGNORECASE)
        if m:
            ped.registration = m.group(1)

        _parse_labeled_ancestors(text, ped)

    else:
        m = re.search(r'(?:Name of Dog|犬名|Dog Name|名前)\s*[:\s]*\n?\s*(.+?)(?:\n|$)', text, re.IGNORECASE)
        if m:
            ped.dog_name = m.group(1).strip()
        else:
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            if lines:
                ped.dog_name = lines[0]

        m = re.search(r'(?:登録番号|Registration|Reg\.?\s*No\.?)\s*[:\s]*([A-Z0-9\-/]+)', text, re.IGNORECASE)
        if m:
            ped.registration = m.group(1)

        _parse_labeled_ancestors(text, ped)
        if not ped.sire:
            _parse_jkc_ancestors(text, ped)

    if not ped.dog_name:
        # フォールバック：ラベル行やOCRノイズを除外して犬名候補を探す
        skip_patterns = re.compile(
            r'^(?:JAPAN|JKC|ALAJ|AKC|KENNEL|CLUB|PEDIGREE|BREED|SIRE|DAM|Name|犬名|犬種|血統|登録|'
            r'Date|所有者|繁殖者|Owner|Breeder|Microchip|マイクロ|性別|毛色|Color|生年月日)',
            re.IGNORECASE
        )
        lines = [l.strip() for l in text.split('\n') if l.strip() and len(l.strip()) > 5]
        for line in lines:
            if not skip_patterns.search(line):
                ped.dog_name = line
                break

    return ped


def calc_coi_3gen(ped: Pedigree) -> dict:
    """3世代の血統データからCOIを算出"""
    sire_ancestors = []
    dam_ancestors = []

    def _normalize_name(name):
        """スペースの揺れを正規化して比較精度を向上"""
        return re.sub(r'\s+', ' ', name.strip().upper())

    def add_if_exists(lst, ancestor, gen):
        if ancestor and ancestor.name:
            lst.append({"name": _normalize_name(ancestor.name), "gen": gen})

    add_if_exists(sire_ancestors, ped.sire, 1)
    add_if_exists(sire_ancestors, ped.ss, 2)
    add_if_exists(sire_ancestors, ped.sd, 2)
    add_if_exists(sire_ancestors, ped.sss, 3)
    add_if_exists(sire_ancestors, ped.ssd, 3)
    add_if_exists(sire_ancestors, ped.sds, 3)
    add_if_exists(sire_ancestors, ped.sdd, 3)

    add_if_exists(dam_ancestors, ped.dam, 1)
    add_if_exists(dam_ancestors, ped.ds, 2)
    add_if_exists(dam_ancestors, ped.dd, 2)
    add_if_exists(dam_ancestors, ped.dss, 3)
    add_if_exists(dam_ancestors, ped.dsd, 3)
    add_if_exists(dam_ancestors, ped.dds, 3)
    add_if_exists(dam_ancestors, ped.ddd, 3)

    coi = 0.0
    common = []
    for sa in sire_ancestors:
        for da in dam_ancestors:
            if sa["name"] == da["name"]:
                contribution = 0.5 ** (sa["gen"] + da["gen"] + 1)
                coi += contribution
                common.append({
                    "name": sa["name"],
                    "sire_gen": sa["gen"],
                    "dam_gen": da["gen"],
                    "contribution": contribution
                })

    return {
        "coi": coi,
        "coi_pct": coi * 100,
        "common_ancestors": common,
        "sire_count": len(sire_ancestors),
        "dam_count": len(dam_ancestors),
    }


def calc_coi_cross(sire_ped: Pedigree, dam_ped: Pedigree) -> dict:
    """2頭の血統書から交配時のCOIを算出（子犬のCOI予測）"""
    sire_all = []
    dam_all = []

    def collect(lst, ped, base_gen=0):
        ancestors_data = [
            (ped.sire, 1), (ped.dam, 1),
            (ped.ss, 2), (ped.sd, 2), (ped.ds, 2), (ped.dd, 2),
            (ped.sss, 3), (ped.ssd, 3), (ped.sds, 3), (ped.sdd, 3),
            (ped.dss, 3), (ped.dsd, 3), (ped.dds, 3), (ped.ddd, 3),
        ]
        def _norm(n):
            return re.sub(r'\s+', ' ', n.strip().upper())
        if ped.dog_name:
            lst.append({"name": _norm(ped.dog_name), "gen": base_gen})
        for anc, gen in ancestors_data:
            if anc and anc.name:
                lst.append({"name": _norm(anc.name), "gen": base_gen + gen})

    collect(sire_all, sire_ped, 0)
    collect(dam_all, dam_ped, 0)

    coi = 0.0
    common = []
    seen = set()
    for sa in sire_all:
        for da in dam_all:
            if sa["name"] == da["name"]:
                key = f"{sa['name']}_{sa['gen']}_{da['gen']}"
                if key not in seen:
                    seen.add(key)
                    contribution = 0.5 ** (sa["gen"] + da["gen"] + 1)
                    coi += contribution
                    common.append({
                        "name": sa["name"],
                        "sire_gen": sa["gen"],
                        "dam_gen": da["gen"],
                        "contribution": contribution
                    })

    return {
        "coi": coi,
        "coi_pct": coi * 100,
        "common_ancestors": common,
    }


# ████████████████████████████████████████████████████████████
# PART 3: 統合HTML出力
# ████████████████████████████████████████████████████████████

def sanitize_text(text: str) -> str:
    """PDF由来の文字化け・制御文字を修正"""
    if not text:
        return text
    # 制御文字を除去
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    # よくある文字化けパターンを修正
    text = text.replace('\ufffd', '')      # 置換文字(�)を除去
    text = re.sub(r'(?<=[a-z])\s*\ufb00\s*(?=[a-z])', 'ff', text)  # ﬀ → ff
    text = re.sub(r'(?<=[a-z])\s*\ufb01\s*(?=[a-z])', 'fi', text)  # ﬁ → fi
    text = re.sub(r'(?<=[a-z])\s*\ufb02\s*(?=[a-z])', 'fl', text)  # ﬂ → fl
    text = re.sub(r'(?<=[a-z])\s*\ufb03\s*(?=[a-z])', 'ffi', text) # ﬃ → ffi
    text = re.sub(r'(?<=[a-z])\s*\ufb04\s*(?=[a-z])', 'ffl', text) # ﬄ → ffl
    return text


def sanitize_for_excel(text: str) -> str:
    """Excel出力用: 制御文字除去 + CSV/Excel formula injection 対策。

    OWASP CSV Injection: 先頭が =/+/-/@ のセルは Excel/LibreOffice で
    式として評価され、外部URL取得や任意関数実行のリスクがある。
    先頭に ' をプレフィックスして文字列扱いを強制する（[BUG-006]）。
    """
    text = sanitize_text(text)
    if isinstance(text, str) and text and text[0] in ('=', '+', '-', '@'):
        text = "'" + text
    return text


def _h(text) -> str:
    """HTMLエスケープ（XSS対策）"""
    if text is None:
        return ""
    s = str(text)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&#x27;")


def status_badge(status: str, text: str) -> str:
    return f'<span class="status {status}">{_h(text)}</span>'


def generate_unified_html(dogs: list, pedigrees: list, output_path: str):
    """遺伝子検査 + 血統書 + COIの統合HTMLレポートを生成"""

    now_str = datetime.now().strftime('%Y年%m月%d日')

    # Count totals
    total_normal = sum(len([r for r in d.health_results if r.status == "normal"]) for d in dogs)
    total_carrier = sum(len([r for r in d.health_results if r.status == "carrier"]) for d in dogs)
    total_positive = sum(len([r for r in d.health_results if r.status == "positive"]) for d in dogs)

    # 重症度別の陽性・キャリア集計（理解できるコンセプト）
    def _disease_severity_for_result(r):
        d = get_disease_detail(r.test_name)
        return get_disease_severity(d) if d else None

    high_risk_positive = 0
    high_risk_carrier = 0
    for d in dogs:
        for r in d.health_results:
            sev = _disease_severity_for_result(r)
            if sev == "high":
                if r.status == "positive":
                    high_risk_positive += 1
                elif r.status == "carrier":
                    high_risk_carrier += 1

    has_orivet = len(dogs) > 0
    has_pedigree = len(pedigrees) > 0

    # ── Dog tabs (Orivet) ──
    tab_buttons = ""
    tab_contents = ""
    sex_i18n_en = {}  # per-dog sex translations for JS
    het_i18n_ja = {}  # ヘテロ接合率パネルの日本語ラベル（EN→JA トグル復元用）

    for idx, dog in enumerate(dogs):
        name = dog.pet_name or dog.registered_name or f"犬{idx+1}"
        safe_id = re.sub(r'[^a-zA-Z0-9]', '_', name.lower())

        tab_buttons += f'    <div class="tab" onclick="showTab(\'{safe_id}\')">{_h(name)}</div>\n'

        sex_class = "male" if "male" in dog.sex.lower() else "female"
        sex_label = "オス" if "male" in dog.sex.lower() else "メス"
        sex_label_en = "Male" if "male" in dog.sex.lower() else "Female"
        sex_i18n_en[f"sex_{safe_id}"] = sex_label_en

        health_rows = ""
        for r in dog.health_results:
            display_name = _h(r.japanese_name if r.japanese_name else r.test_name)
            badge = status_badge(r.status, r.genotype if r.genotype else r.status.upper())
            annotation = get_health_annotation(r.test_name, r.genotype, r.status)
            annotation_html = f'<div style="margin-top:4px;padding:6px 8px;background:#fef3c7;border-left:3px solid #f59e0b;border-radius:4px;font-size:0.85em;color:#374151;">{_h(annotation)}</div>' if annotation else ''
            # ↓ 「理解できること」コンセプト: 詳細解説 + 参考リンク
            detail = get_disease_detail(r.test_name)
            detail_html = render_detail_html(detail) if detail else ""
            # 重症度バッジ（KB エントリがある場合のみ表示）
            severity_html = ""
            if detail:
                sev = get_disease_severity(detail)
                meta = SEVERITY_LABELS.get(sev, {})
                if meta:
                    severity_html = (
                        f'<span class="severity-badge" style="background:{meta["bg"]};color:{meta["color"]};'
                        f'display:inline-block;margin-left:6px;padding:1px 8px;border-radius:10px;'
                        f'font-size:0.72em;font-weight:700;vertical-align:middle;">'
                        f'{meta["emoji"]} {meta["label"]}</span>'
                    )
            health_rows += f"""        <tr>
          <td>{_h(r.category)}</td>
          <td>{display_name}{severity_html}<br><small style="color:#6b7280">{_h(r.test_name)}</small>{annotation_html}{detail_html}</td>
          <td>{badge}</td>
          <td style="font-size:0.85em">{_h(r.result_text[:120])}</td>
        </tr>\n"""

        trait_rows = ""
        for r in dog.trait_results:
            display_name = _h(r.japanese_name if r.japanese_name else r.test_name)
            badge = status_badge("trait", r.genotype if r.genotype else "—")
            annotation = get_trait_annotation(r.test_name, r.genotype)
            annotation_html = f'<div style="margin-top:4px;padding:6px 8px;background:#f0f4ff;border-left:3px solid #667eea;border-radius:4px;font-size:0.85em;color:#374151;">{_h(annotation)}</div>' if annotation else ''
            # ↓ 形質の詳細解説 + 参考リンク
            t_detail = get_trait_detail(r.test_name)
            t_detail_html = render_detail_html(t_detail) if t_detail else ""
            # 重症度バッジ（致死ホモ・ダブルマール等の危険形質のみ — 明示 severity がある場合のみ表示）
            trait_severity_html = ""
            if t_detail and t_detail.get("severity") in ("high", "medium", "low"):
                sev = t_detail["severity"]
                meta = SEVERITY_LABELS.get(sev, {})
                if meta:
                    trait_severity_html = (
                        f'<span class="severity-badge" style="background:{meta["bg"]};color:{meta["color"]};'
                        f'display:inline-block;margin-left:6px;padding:1px 8px;border-radius:10px;'
                        f'font-size:0.72em;font-weight:700;vertical-align:middle;">'
                        f'{meta["emoji"]} {meta["label"]}</span>'
                    )
            trait_rows += f"""        <tr>
          <td>{display_name}{trait_severity_html}<br><small style="color:#6b7280">{_h(r.test_name)}</small>{t_detail_html}</td>
          <td>{badge}</td>
          <td style="font-size:0.85em">{_h(r.result_text[:150])}{annotation_html}</td>
        </tr>\n"""

        # ── ヘテロ接合率（ゲノム多様性）パネル ──
        # Orivet PDF に Heterozygosity Score が含まれる場合のみ表示。
        # 血統ベース COI とは別指標であることを明示する。
        hetero_block = ""
        if dog.heterozygosity is not None:
            het = dog.heterozygosity
            rng = dog.heterozygosity_range
            range_html = ""
            if rng and len(rng) == 2:
                low, high = rng[0], rng[1]
                if het < low:
                    judge_ja = "標準域より低い（多様性やや低め）"
                    judge_en = "Below the breed range (diversity slightly low)"
                    jcolor = "#92400e"
                elif het > high:
                    judge_ja = "標準域より高い（多様性良好）"
                    judge_en = "Above the breed range (good diversity)"
                    jcolor = "#166534"
                else:
                    judge_ja = "標準域内"
                    judge_en = "Within the breed range"
                    jcolor = "#0e7490"
                range_html = (
                    f'<div style="margin-top:8px;font-size:0.9em;">'
                    f'<span style="color:{jcolor};font-weight:700;" data-i18n="het_judge_{safe_id}">{judge_ja}</span>'
                    f'<span style="color:#6b7280;"> — <span data-i18n="het_range_label">犬種標準域</span> {low}%–{high}%</span>'
                    f'</div>'
                )
                sex_i18n_en[f"het_judge_{safe_id}"] = judge_en
                het_i18n_ja[f"het_judge_{safe_id}"] = judge_ja
            hetero_block = f"""
      <div style="margin:14px 0;padding:14px 18px;background:#ecfeff;border:1px solid #a5f3fc;border-radius:10px;">
        <div style="font-weight:700;color:#0e7490;margin-bottom:4px;">
          🔬 <span data-i18n="het_title">ヘテロ接合率（ゲノム多様性）</span>:
          <span style="font-size:1.15em;">{het}%</span>
        </div>
        {range_html}
        <div style="margin-top:8px;font-size:0.82em;color:#6b7280;line-height:1.6;">
          <span data-i18n="het_note">⚠️ これは Orivet の DNA 検査が実測した値です。本アプリの繁殖シミュレーターが算出する『血統ベース COI』とは別指標のため、数値を直接比較しないでください（ヘテロ接合率は高いほど遺伝的多様性が高く、一般に良好）。</span>
        </div>
      </div>"""
            sex_i18n_en["het_title"] = "Heterozygosity (genomic diversity)"
            sex_i18n_en["het_range_label"] = "breed typical range"
            sex_i18n_en["het_note"] = (
                "⚠️ This is the value measured by Orivet's DNA test. It is a different metric "
                "from the pedigree-based COI calculated by this app's breeding simulator, so do "
                "not compare the numbers directly (higher heterozygosity means greater genetic "
                "diversity, generally favorable)."
            )
            het_i18n_ja["het_title"] = "ヘテロ接合率（ゲノム多様性）"
            het_i18n_ja["het_range_label"] = "犬種標準域"
            het_i18n_ja["het_note"] = (
                "⚠️ これは Orivet の DNA 検査が実測した値です。本アプリの繁殖シミュレーターが算出する"
                "『血統ベース COI』とは別指標のため、数値を直接比較しないでください"
                "（ヘテロ接合率は高いほど遺伝的多様性が高く、一般に良好）。"
            )

        tab_contents += f"""
  <div id="{safe_id}" class="tab-content">
    <div class="dog-card">
      <div class="dog-header">
        <div>
          <div class="dog-name">{_h(name)}</div>
          <div class="dog-reg">{_h(dog.registered_name)} — {_h(dog.case_number)}</div>
        </div>
        <div class="dog-meta">
          <span class="meta-tag {sex_class}"><span data-i18n="sex_{safe_id}">{_h(sex_label)}</span> ({_h(dog.sex)})</span>
          <span class="meta-tag">{_h(dog.breed)}</span>
          <span class="meta-tag">{_h(dog.dob)}</span>
          {'<span class="meta-tag">MC: ' + _h(dog.microchip) + '</span>' if dog.microchip else ''}
          {'<span class="meta-tag"><span data-i18n="lbl_colour">毛色</span>: ' + _h(dog.colour) + '</span>' if dog.colour else ''}
        </div>
      </div>
{hetero_block}
      <h3 class="section-title"><span data-i18n="health_results">健康検査結果</span> ({len(dog.health_results)}<span data-i18n="items_suffix">項目</span>)</h3>
      <table class="results-table">
        <tr><th data-i18n="th_category">カテゴリー</th><th data-i18n="th_test">検査項目</th><th data-i18n="th_result">結果</th><th data-i18n="th_detail">詳細</th></tr>
{health_rows}
      </table>

      <h3 class="section-title"><span data-i18n="trait_results">毛色・形質検査結果</span> ({len(dog.trait_results)}<span data-i18n="items_suffix">項目</span>)</h3>
      <table class="results-table">
        <tr><th data-i18n="th_test">検査項目</th><th data-i18n="th_genotype">遺伝子型</th><th data-i18n="th_detail">詳細</th></tr>
{trait_rows}
      </table>
    </div>
  </div>"""

    # ── Comparison table ──
    compare_tab_button = ""
    compare_tab_html = ""
    if len(dogs) > 1:
        all_tests = {}
        for dog in dogs:
            for r in dog.health_results:
                key = r.test_name
                if key not in all_tests:
                    all_tests[key] = {"jp": r.japanese_name or r.test_name, "name": r.test_name}

        compare_header = '<th data-i18n="th_test">検査項目</th>'
        for dog in dogs:
            compare_header += f"<th>{_h(dog.pet_name or dog.registered_name)}</th>"

        compare_health_rows = ""
        for test_key, test_info in all_tests.items():
            row = f"<td>{test_info['jp']}</td>"
            for dog in dogs:
                found = False
                for r in dog.health_results:
                    if r.test_name == test_key:
                        badge = status_badge(r.status, r.genotype if r.genotype else r.status.upper())
                        row += f"<td>{badge}</td>"
                        found = True
                        break
                if not found:
                    row += "<td>—</td>"
            compare_health_rows += f"        <tr>{row}</tr>\n"

        compare_tab_button = '<div class="tab" onclick="showTab(\'compare\')"><span data-i18n="tab_compare">比較表</span></div>'
        compare_tab_html = f"""<div id="compare" class="tab-content">
    <div class="dog-card">
      <h2 class="section-title" data-i18n="health_compare">健康検査 比較表</h2>
      <div style="overflow-x:auto;">
      <table class="compare-table">
        <tr>{compare_header}</tr>
{compare_health_rows}
      </table>
      </div>
    </div>
  </div>"""

    # ── Alerts ──
    alerts_html = ""
    for dog in dogs:
        for r in dog.health_results:
            name = _h(dog.pet_name or dog.registered_name)
            display_name = _h(r.japanese_name if r.japanese_name else r.test_name)
            if r.status == "positive":
                alerts_html += f"""      <div class="breed-warn danger">
        <div class="warn-title">{display_name} — <span data-i18n="lbl_positive">ポジティブ</span> (P/P): {name}</div>
        <p data-i18n="alert_positive">変異が2コピー検出されました。発症リスクがあります。獣医師にご相談の上、適切なケアをお願いいたします。</p>
        <p><small><span data-i18n="lbl_original">原文</span>: {_h(r.result_text[:200])}</small></p>
      </div>\n"""
            elif r.status == "carrier":
                alerts_html += f"""      <div class="breed-warn">
        <div class="warn-title">{display_name} — <span data-i18n="lbl_carrier">キャリア</span> (P/N): {name}</div>
        <p data-i18n="alert_carrier">変異が1コピー検出されました。キャリアまたはポジティブの個体との交配で発症する子犬が生まれる可能性があります。</p>
      </div>\n"""

    if not alerts_html and has_orivet:
        alerts_html = '<div class="breed-warn" style="background:#dcfce7;border-color:#86efac;"><div class="warn-title" style="color:#166534;" data-i18n="all_clear">全頭クリア</div><p data-i18n="all_clear_desc">全ての健康検査項目でノーマル（変異なし）でした。</p></div>'

    # ── Overview table rows ──
    overview_table_rows = ""
    for d in dogs:
        overview_table_rows += f"<tr><td><strong>{_h(d.pet_name)}</strong></td><td>{_h(d.registered_name)}</td><td>{_h(d.breed)}</td><td>{_h(d.sex)}</td><td>{_h(d.dob)}</td><td>{_h(d.case_number)}</td></tr>\n"

    # ── Pedigree section ──
    pedigree_tab_button = ""
    pedigree_tab_html = ""
    if has_pedigree:
        pedigree_tab_button = '<div class="tab" onclick="showTab(\'pedigree\')"><span data-i18n="tab_pedigree">血統書 + COI</span></div>'

        ped_parts = []
        for ped in pedigrees:
            coi_result = calc_coi_3gen(ped)

            ancestors_html = ""
            for label, anc in ped.all_ancestors():
                if anc:
                    _color_map = {
                        "BLK": "#1a1a1a", "BLACK": "#1a1a1a",
                        "WH": "#f5f5f5", "WHITE": "#f5f5f5",
                        "BR": "#8B4513", "BROWN": "#8B4513", "CHOCO": "#5C3317",
                        "RED": "#CD5C5C",
                        "APR": "#FBCEB1", "APRICOT": "#FBCEB1",
                        "CR": "#FFF8DC", "CREAM": "#FFF8DC",
                        "SV": "#C0C0C0", "SILVER": "#C0C0C0",
                        "BL": "#4a6fa5", "BLUE": "#4a6fa5",
                        "CAFE": "#A67B5B", "CAFE AU LAIT": "#A67B5B",
                        "GR": "#DAA520", "GOLD": "#DAA520", "GOLDEN": "#DAA520",
                        "F": "#D2B48C", "FAWN": "#D2B48C",
                        "SBL": "#D2691E", "SABLE": "#D2691E",
                        "BRN": "#CD853F", "BRINDLE": "#8B7355",
                        "MERLE": "#9FB6CD",
                        "PARTI": "#ffffff",
                        "TAN": "#D2B48C",
                        "BEIGE": "#F5F5DC",
                    }
                    _c = (anc.color or "").strip().upper()
                    color_dot = _color_map.get(_c, "#9ca3af")
                    # 部分一致フォールバック
                    if color_dot == "#9ca3af" and _c:
                        for k, v in _color_map.items():
                            if k in _c or _c in k:
                                color_dot = v
                                break
                    ancestors_html += f"""<tr>
                        <td>{_h(label)}</td>
                        <td style="font-weight:700;">{_h(anc.name)}</td>
                        <td>{_h(anc.registration)}</td>
                        <td><span style="display:inline-block;width:16px;height:16px;border-radius:50%;background:{color_dot};border:1px solid #bbb;vertical-align:middle;margin-right:6px;"></span>{_h(anc.color)}</td>
                        <td style="font-size:0.8em;">{_h(anc.titles)}</td>
                        <td style="font-size:0.8em;">{_h(anc.dna_number)}</td>
                    </tr>"""

            coi_color = '#22c55e' if coi_result['coi_pct'] < 6.25 else '#eab308' if coi_result['coi_pct'] < 12.5 else '#ef4444'
            common_text = ""
            if coi_result['common_ancestors']:
                names = ", ".join([_h(c['name']) for c in coi_result['common_ancestors']])
                common_text = f'<p><span data-i18n="lbl_common_ancestors">共通祖先</span>: {names}</p>'
            else:
                common_text = '<p data-i18n="no_common_ancestors">3世代以内に共通祖先は検出されませんでした。</p>'

            ped_parts.append(f"""
        <div class="dog-card">
            <h2 class="section-title">{_h(ped.dog_name)}</h2>
            <div class="info-grid">
                <div><strong data-i18n="lbl_breed">犬種</strong>: {_h(ped.breed)}</div>
                <div><strong data-i18n="lbl_reg_no">登録番号</strong>: {_h(ped.registration)}</div>
                <div><strong data-i18n="lbl_sex">性別</strong>: {_h(ped.sex)}</div>
                <div><strong data-i18n="lbl_dob">生年月日</strong>: {_h(ped.dob)}</div>
                <div><strong data-i18n="lbl_colour">毛色</strong>: {_h(ped.color)}</div>
                <div><strong data-i18n="lbl_microchip">マイクロチップ</strong>: {_h(ped.microchip)}</div>
                <div><strong data-i18n="lbl_breeder">ブリーダー</strong>: {_h(ped.breeder)}</div>
                <div><strong data-i18n="lbl_owner">オーナー</strong>: {_h(ped.owner)}</div>
            </div>

            <h3 class="section-title" data-i18n="ped_3gen">3世代血統表</h3>
            <table class="results-table">
                <tr><th data-i18n="th_position">位置</th><th data-i18n="th_dog_name">犬名</th><th data-i18n="th_reg_no">登録番号</th><th data-i18n="th_colour">毛色</th><th data-i18n="th_title">タイトル</th><th data-i18n="th_dna">DNA番号</th></tr>
                {ancestors_html}
            </table>

            <h3 class="section-title" data-i18n="coi_individual">近親交配係数 (COI) — 個体分析</h3>
            <div style="text-align:center;margin:20px 0;">
                <div style="font-size:3em;font-weight:800;color:{coi_color};">{coi_result['coi_pct']:.2f}%</div>
                <div style="color:#6b7280;" data-i18n="coi_wright">Wright's COI (3世代)</div>
            </div>
            {common_text}
        </div>""")

        # Cross COI if multiple pedigrees
        cross_html = ""
        if len(pedigrees) >= 2:
            cross_result = calc_coi_cross(pedigrees[0], pedigrees[1])
            cross_color = '#22c55e' if cross_result['coi_pct'] < 6.25 else '#eab308' if cross_result['coi_pct'] < 12.5 else '#ef4444'
            cross_common = ""
            if cross_result['common_ancestors']:
                items = ""
                for c in cross_result['common_ancestors']:
                    items += f"<li>{_h(c['name'])} (<span data-i18n='lbl_sire_side'>父方</span>{c['sire_gen']}<span data-i18n='lbl_gen'>世代</span> / <span data-i18n='lbl_dam_side'>母方</span>{c['dam_gen']}<span data-i18n='lbl_gen'>世代</span> → <span data-i18n='lbl_contribution'>寄与</span>: {c['contribution']*100:.3f}%)</li>"
                cross_common = f'<h3 data-i18n="lbl_common_ancestors">共通祖先</h3><ul>{items}</ul>'
            else:
                cross_common = '<p data-i18n="no_common_detected">共通祖先は検出されませんでした。</p>'

            cross_html = f"""
        <div class="dog-card">
            <h2 class="section-title"><span data-i18n="cross_coi_title">交配COI予測</span>: {_h(pedigrees[0].dog_name)} × {_h(pedigrees[1].dog_name)}</h2>
            <div style="text-align:center;margin:20px 0;">
                <div style="font-size:3em;font-weight:800;color:{cross_color};">{cross_result['coi_pct']:.2f}%</div>
                <div style="color:#6b7280;" data-i18n="expected_puppy_coi">予想される子犬のCOI</div>
            </div>
            {cross_common}
        </div>"""

        pedigree_tab_html = f"""<div id="pedigree" class="tab-content">
    {"".join(ped_parts)}
    {cross_html}
  </div>"""

    # ── Summary card counts ──
    summary_html = ""
    # 高リスク陽性カードのスタイル（陽性件数の有無で強調度を変える）
    high_risk_emphasis = ""
    if high_risk_positive > 0:
        high_risk_emphasis = "box-shadow:0 0 0 3px rgba(220,38,38,0.4),0 2px 8px rgba(0,0,0,0.06);"
    if has_orivet:
        summary_html = f"""  <div class="summary-row">
    <div class="summary-card"><div class="num blue">{len(dogs)}</div><div class="label" data-i18n="sum_tested">検査頭数</div></div>
    <div class="summary-card"><div class="num green">{total_normal}</div><div class="label" data-i18n="sum_normal">ノーマル項目</div></div>
    <div class="summary-card"><div class="num yellow">{total_carrier}</div><div class="label" data-i18n="sum_carrier">キャリア項目</div></div>
    <div class="summary-card"><div class="num red">{total_positive}</div><div class="label" data-i18n="sum_positive">ポジティブ (要注意)</div></div>
    <div class="summary-card" style="{high_risk_emphasis}"><div class="num red" style="font-size:1.7em;">🚨 {high_risk_positive}</div><div class="label" data-i18n="sum_high_risk_pos" style="font-weight:600;">高リスク疾患の陽性</div></div>
    {'<div class="summary-card"><div class="num" style="color:#4a1a7a;">' + str(len(pedigrees)) + '</div><div class="label" data-i18n="sum_pedigree">血統書データ</div></div>' if has_pedigree else ''}
  </div>"""
    elif has_pedigree:
        summary_html = f"""  <div class="summary-row">
    <div class="summary-card"><div class="num" style="color:#4a1a7a;">{len(pedigrees)}</div><div class="label" data-i18n="sum_pedigree">血統書データ</div></div>
  </div>"""

    # ── Subtitle ──
    features = []
    features_en = []
    if has_orivet:
        features.append("遺伝子検査")
        features_en.append("Genetic Test")
    if has_pedigree:
        features.append("血統書")
        features.append("COI算出")
        features_en.append("Pedigree")
        features_en.append("COI Calculation")
    subtitle = " + ".join(features)
    subtitle_en = " + ".join(features_en)

    # ── Pre-build conditional sections (avoid backslash in f-string) ──
    overview_tab_button = '<div class="tab active" onclick="showTab(\'overview\')"><span data-i18n="tab_overview">全体サマリー</span></div>' if has_orivet else ''

    if has_orivet:
        overview_table_html = '<div class="dog-card"><h2 class="section-title" data-i18n="test_subjects">検査対象一覧</h2><table class="results-table"><tr><th data-i18n="th_pet_name">ペット名</th><th data-i18n="th_reg_name">登録名</th><th data-i18n="lbl_breed">犬種</th><th data-i18n="lbl_sex">性別</th><th data-i18n="lbl_dob">生年月日</th><th data-i18n="th_case_no">ケース番号</th></tr>' + overview_table_rows + '</table></div>'
        alerts_section = '<div class="dog-card"><h2 class="section-title" data-i18n="alerts_title">要注意事項</h2>' + alerts_html + '</div>'
        overview_section = '<!-- Overview Tab -->\n  <div id="overview" class="tab-content active">\n  ' + overview_table_html + '\n  ' + alerts_section + '\n  </div>'
    else:
        overview_section = ''

    auto_activate_js = '' if has_orivet else "document.querySelector('.tab')?.click();"

    # ── Full HTML ──
    html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Orivet Genetics Report</title>
<style>
:root {{ --pink:#e6007e; --purple:#4a1a7a; --green:#22c55e; --yellow:#eab308; --red:#ef4444; --gray:#6b7280; --bg:#f8f9fa; }}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI','Hiragino Sans','Meiryo',sans-serif; background:var(--bg); color:#1f2937; line-height:1.6; }}
.container {{ max-width:1200px; margin:0 auto; padding:20px; }}
header {{ background:linear-gradient(135deg,var(--purple),var(--pink)); color:white; padding:30px 0; margin-bottom:30px; border-radius:0 0 20px 20px; }}
header .container {{ display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px; }}
header h1 {{ font-size:1.8em; }}
header p {{ opacity:0.9; font-size:0.95em; }}
.badge {{ display:inline-block; background:rgba(255,255,255,0.2); padding:4px 12px; border-radius:20px; font-size:0.85em; }}
.summary-row {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(140px,1fr)); gap:15px; margin-bottom:30px; }}
.summary-card {{ background:white; border-radius:12px; padding:20px; text-align:center; box-shadow:0 2px 8px rgba(0,0,0,0.06); }}
.summary-card .num {{ font-size:2em; font-weight:700; }}
.summary-card .label {{ font-size:0.85em; color:var(--gray); margin-top:4px; }}
.num.green {{ color:var(--green); }} .num.yellow {{ color:var(--yellow); }} .num.red {{ color:var(--red); }} .num.blue {{ color:#3b82f6; }}
.tabs {{ display:flex; gap:8px; margin-bottom:20px; flex-wrap:wrap; }}
.tab {{ padding:10px 20px; border-radius:10px 10px 0 0; cursor:pointer; font-weight:600; border:2px solid #e5e7eb; border-bottom:none; background:white; transition:all 0.2s; font-size:0.95em; }}
.tab:hover {{ background:#f3e8ff; }}
.tab.active {{ background:var(--purple); color:white; border-color:var(--purple); }}
.tab-content {{ display:none; }} .tab-content.active {{ display:block; }}
.dog-card {{ background:white; border-radius:16px; padding:24px; margin-bottom:24px; box-shadow:0 2px 12px rgba(0,0,0,0.06); }}
.dog-header {{ display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:15px; margin-bottom:20px; padding-bottom:16px; border-bottom:2px solid #f3f4f6; }}
.dog-name {{ font-size:1.4em; font-weight:700; color:var(--purple); }}
.dog-reg {{ font-size:0.85em; color:var(--gray); }}
.dog-meta {{ display:flex; gap:12px; flex-wrap:wrap; }}
.meta-tag {{ background:#f3f4f6; padding:4px 12px; border-radius:20px; font-size:0.82em; white-space:nowrap; }}
.meta-tag.male {{ background:#dbeafe; color:#1e40af; }} .meta-tag.female {{ background:#fce7f3; color:#be185d; }}
.section-title {{ font-size:1.1em; font-weight:700; color:var(--purple); margin:20px 0 12px; padding-left:10px; border-left:4px solid var(--pink); }}
.results-table {{ width:100%; border-collapse:separate; border-spacing:0; font-size:0.88em; }}
.results-table th {{ background:var(--purple); color:white; padding:10px 12px; text-align:left; font-weight:600; }}
.results-table th:first-child {{ border-radius:8px 0 0 0; }} .results-table th:last-child {{ border-radius:0 8px 0 0; }}
.results-table td {{ padding:10px 12px; border-bottom:1px solid #f3f4f6; vertical-align:top; }}
.results-table tr:hover td {{ background:#faf5ff; }}
.compare-table {{ width:100%; border-collapse:separate; border-spacing:0; font-size:0.85em; }}
.compare-table th {{ background:var(--purple); color:white; padding:10px; text-align:center; position:sticky; top:0; }}
.compare-table th:first-child {{ text-align:left; min-width:200px; }}
.compare-table td {{ padding:8px 10px; border-bottom:1px solid #f3f4f6; text-align:center; }}
.compare-table td:first-child {{ text-align:left; font-weight:500; }}
.compare-table tr:hover td {{ background:#faf5ff; }}
.status {{ display:inline-block; padding:3px 10px; border-radius:12px; font-size:0.82em; font-weight:600; white-space:nowrap; }}
.status.normal {{ background:#dcfce7; color:#166534; }}
.status.carrier {{ background:#fef3c7; color:#92400e; }}
.status.positive {{ background:#fee2e2; color:#991b1b; }}
.status.trait {{ background:#e0e7ff; color:#3730a3; }}
.legend {{ display:flex; gap:16px; flex-wrap:wrap; margin:16px 0; padding:12px 16px; background:white; border-radius:10px; font-size:0.85em; }}
.legend-item {{ display:flex; align-items:center; gap:6px; }}
.legend-dot {{ width:14px; height:14px; border-radius:4px; }}
.info-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:16px; font-size:0.9em; }}
.breed-warn {{ background:#fff7ed; border:1px solid #fed7aa; border-radius:10px; padding:16px; margin:12px 0; }}
.breed-warn.danger {{ background:#fef2f2; border-color:#fecaca; }}
.breed-warn .warn-title {{ font-weight:700; color:#c2410c; margin-bottom:4px; }}
.breed-warn.danger .warn-title {{ color:#dc2626; }}
.print-btn {{ background:var(--purple); color:white; border:none; padding:10px 20px; border-radius:8px; cursor:pointer; font-weight:600; }}
.print-btn:hover {{ background:var(--pink); }}
/* 詳細解説（理解できることコンセプト）— 折りたためる knowledge base 表示 */
details.kb-detail {{
  margin-top:8px; padding:6px 10px; background:#f8fafc;
  border:1px solid #e2e8f0; border-radius:6px; font-size:0.85em;
}}
details.kb-detail > summary {{
  cursor:pointer; font-weight:600; color:#5b21b6; padding:4px 0;
  user-select:none; list-style:none; outline:none;
}}
details.kb-detail > summary::-webkit-details-marker {{ display:none; }}
details.kb-detail > summary::before {{
  content:"▶"; display:inline-block; margin-right:6px; font-size:0.7em;
  transition:transform 0.15s;
}}
details.kb-detail[open] > summary::before {{ transform:rotate(90deg); }}
details.kb-detail .kb-body {{
  padding:10px 4px 4px; line-height:1.6; color:#1f2937;
}}
details.kb-detail .kb-title {{
  font-size:1.05em; color:#5b21b6; margin-bottom:10px; padding-bottom:6px;
  border-bottom:1px dashed #ddd6fe;
}}
details.kb-detail .kb-section {{ margin-bottom:10px; }}
details.kb-detail .kb-section strong {{ display:block; margin-bottom:3px; color:#4a1a7a; font-size:0.95em; }}
details.kb-detail .kb-refs a.kb-ref-link {{
  display:inline-block; margin:2px 6px 2px 0; padding:4px 10px;
  background:#ede9fe; color:#5b21b6; text-decoration:none;
  border-radius:14px; font-size:0.85em;
}}
details.kb-detail .kb-refs a.kb-ref-link:hover {{ background:#ddd6fe; }}
@media (max-width:768px) {{ header h1 {{ font-size:1.3em; }} .dog-header {{ flex-direction:column; }} .compare-table,.results-table {{ display:block; overflow-x:auto; }} .info-grid {{ grid-template-columns:1fr; }} }}
/* スマートフォン UX 集中改善 (PR #51) */
@media (max-width:480px) {{
  .container {{ padding:12px; }}
  header {{ padding:18px 0; }}
  header h1 {{ font-size:1.15em; }}
  .summary-row {{ grid-template-columns:repeat(2, 1fr); gap:8px; }}
  .summary-card {{ padding:12px; }}
  .summary-card .num {{ font-size:1.4em; }}
  .summary-card .label {{ font-size:0.75em; line-height:1.2; }}
  .tab {{ padding:10px 14px; min-height:44px; font-size:0.85em; }}
  .dog-card {{ padding:14px; border-radius:10px; }}
  .results-table th, .results-table td {{ padding:8px 6px; font-size:0.82em; }}
  details.kb-detail {{ font-size:0.82em; padding:6px 8px; }}
  details.kb-detail .kb-refs a.kb-ref-link {{ padding:6px 12px; min-height:36px; display:inline-block; }}
  body {{ padding-bottom:env(safe-area-inset-bottom); }}
}}
@media print {{ .tabs,.print-btn,header {{ display:none!important; }} .tab-content {{ display:block!important; page-break-inside:avoid; }} .dog-card {{ break-inside:avoid; }} }}
</style>
</head>
<body>
<header>
  <div class="container">
    <div>
      <h1>Orivet Genetics Report</h1>
      <p><span data-i18n="report_subtitle">{subtitle}</span> &nbsp;|&nbsp; <span data-i18n="generated_on">生成日</span>: {now_str}</p>
      <p><span class="badge">Orivet Genetics</span> <span class="badge" data-i18n="badge_pedigree">JKC血統書</span> <span class="badge">Wright's COI</span></p>
    </div>
    <button class="print-btn" onclick="window.print()" data-i18n="btn_print">印刷</button>
  </div>
</header>
<div class="container">
{summary_html}
  <div class="legend">
    <div class="legend-item"><div class="legend-dot" style="background:#dcfce7;border:1px solid #166534;"></div><span data-i18n="leg_normal">ノーマル (N/N)</span></div>
    <div class="legend-item"><div class="legend-dot" style="background:#fef3c7;border:1px solid #92400e;"></div><span data-i18n="leg_carrier">キャリア (P/N)</span></div>
    <div class="legend-item"><div class="legend-dot" style="background:#fee2e2;border:1px solid #991b1b;"></div><span data-i18n="leg_positive">ポジティブ (P/P)</span></div>
    <div class="legend-item"><div class="legend-dot" style="background:#e0e7ff;border:1px solid #3730a3;"></div><span data-i18n="leg_trait">形質 (Trait)</span></div>
  </div>

  <div class="tabs">
    {overview_tab_button}
{tab_buttons}
    {compare_tab_button}
    {pedigree_tab_button}
  </div>

  {overview_section}

  <!-- Individual Dog Tabs -->
{tab_contents}

  <!-- Compare Tab -->
  {compare_tab_html}

  <!-- Pedigree + COI Tab -->
  {pedigree_tab_html}
</div>
<script>
function showTab(id) {{
  document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(el => el.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  event.target.classList.add('active');
}}

// ── Language toggle ──
var REPORT_I18N = {{
  en: Object.assign({{
    report_subtitle: "{subtitle_en}",
    generated_on: "Generated",
    badge_pedigree: "Pedigree",
    btn_print: "Print",
    sum_tested: "Dogs Tested",
    sum_normal: "Normal",
    sum_carrier: "Carrier",
    sum_positive: "Positive (Attention)",
    sum_pedigree: "Pedigree Data",
    leg_normal: "Normal (N/N)",
    leg_carrier: "Carrier (P/N)",
    leg_positive: "Positive (P/P)",
    leg_trait: "Trait",
    tab_overview: "Summary",
    tab_compare: "Comparison",
    tab_pedigree: "Pedigree + COI",
    test_subjects: "Test Subjects",
    th_pet_name: "Pet Name",
    th_reg_name: "Registered Name",
    lbl_breed: "Breed",
    lbl_sex: "Sex",
    lbl_dob: "Date of Birth",
    th_case_no: "Case No.",
    alerts_title: "Alerts",
    health_results: "Health Test Results",
    items_suffix: " items",
    th_category: "Category",
    th_test: "Test",
    th_result: "Result",
    th_detail: "Details",
    trait_results: "Coat Color & Trait Results",
    th_genotype: "Genotype",
    lbl_colour: "Color",
    lbl_positive: "Positive",
    lbl_carrier: "Carrier",
    alert_positive: "Two copies of the mutation detected. There is a risk of developing the condition. Please consult your veterinarian for appropriate care.",
    alert_carrier: "One copy of the mutation detected. Breeding with a carrier or positive individual may produce affected puppies.",
    lbl_original: "Original",
    all_clear: "All Clear",
    all_clear_desc: "All health test results are normal (no mutations detected).",
    health_compare: "Health Test Comparison",
    lbl_reg_no: "Registration No.",
    lbl_microchip: "Microchip",
    lbl_breeder: "Breeder",
    lbl_owner: "Owner",
    ped_3gen: "3-Generation Pedigree",
    th_position: "Position",
    th_dog_name: "Dog Name",
    th_reg_no: "Reg. No.",
    th_colour: "Color",
    th_title: "Titles",
    th_dna: "DNA No.",
    coi_individual: "Coefficient of Inbreeding (COI) — Individual",
    coi_wright: "Wright's COI (3 generations)",
    lbl_common_ancestors: "Common Ancestors",
    no_common_ancestors: "No common ancestors detected within 3 generations.",
    no_common_detected: "No common ancestors detected.",
    cross_coi_title: "Breeding COI Prediction",
    expected_puppy_coi: "Expected Puppy COI",
    lbl_sire_side: "Sire",
    lbl_dam_side: "Dam",
    lbl_gen: " gen",
    lbl_contribution: "Contribution"
  }}, {json.dumps(sex_i18n_en, ensure_ascii=False)}),
  ja: {{
    report_subtitle: "{subtitle}",
    generated_on: "生成日",
    badge_pedigree: "JKC血統書",
    btn_print: "印刷",
    sum_tested: "検査頭数",
    sum_normal: "ノーマル項目",
    sum_carrier: "キャリア項目",
    sum_positive: "ポジティブ (要注意)",
    sum_pedigree: "血統書データ",
    leg_normal: "ノーマル (N/N)",
    leg_carrier: "キャリア (P/N)",
    leg_positive: "ポジティブ (P/P)",
    leg_trait: "形質 (Trait)",
    tab_overview: "全体サマリー",
    tab_compare: "比較表",
    tab_pedigree: "血統書 + COI",
    test_subjects: "検査対象一覧",
    th_pet_name: "ペット名",
    th_reg_name: "登録名",
    lbl_breed: "犬種",
    lbl_sex: "性別",
    lbl_dob: "生年月日",
    th_case_no: "ケース番号",
    alerts_title: "要注意事項",
    health_results: "健康検査結果",
    items_suffix: "項目",
    th_category: "カテゴリー",
    th_test: "検査項目",
    th_result: "結果",
    th_detail: "詳細",
    trait_results: "毛色・形質検査結果",
    th_genotype: "遺伝子型",
    lbl_colour: "毛色",
    lbl_positive: "ポジティブ",
    lbl_carrier: "キャリア",
    alert_positive: "変異が2コピー検出されました。発症リスクがあります。獣医師にご相談の上、適切なケアをお願いいたします。",
    alert_carrier: "変異が1コピー検出されました。キャリアまたはポジティブの個体との交配で発症する子犬が生まれる可能性があります。",
    lbl_original: "原文",
    all_clear: "全頭クリア",
    all_clear_desc: "全ての健康検査項目でノーマル（変異なし）でした。",
    health_compare: "健康検査 比較表",
    lbl_reg_no: "登録番号",
    lbl_microchip: "マイクロチップ",
    lbl_breeder: "ブリーダー",
    lbl_owner: "オーナー",
    ped_3gen: "3世代血統表",
    th_position: "位置",
    th_dog_name: "犬名",
    th_reg_no: "登録番号",
    th_colour: "毛色",
    th_title: "タイトル",
    th_dna: "DNA番号",
    coi_individual: "近親交配係数 (COI) — 個体分析",
    coi_wright: "Wright's COI (3世代)",
    lbl_common_ancestors: "共通祖先",
    no_common_ancestors: "3世代以内に共通祖先は検出されませんでした。",
    no_common_detected: "共通祖先は検出されませんでした。",
    cross_coi_title: "交配COI予測",
    expected_puppy_coi: "予想される子犬のCOI",
    lbl_sire_side: "父方",
    lbl_dam_side: "母方",
    lbl_gen: "世代",
    lbl_contribution: "寄与"
  }}
}};

// Store per-dog sex labels for ja
(function() {{
  var jaExtra = {json.dumps({f"sex_{re.sub(r'[^a-zA-Z0-9]', '_', (d.pet_name or d.registered_name or f'犬{i+1}').lower())}": "オス" if "male" in d.sex.lower() else "メス" for i, d in enumerate(dogs)}, ensure_ascii=False)};
  Object.assign(REPORT_I18N.ja, jaExtra);
  Object.assign(REPORT_I18N.ja, {json.dumps(het_i18n_ja, ensure_ascii=False)});
}})();

function __applyLang(lang) {{
  var dict = REPORT_I18N[lang];
  if (!dict) return;
  document.querySelectorAll('[data-i18n]').forEach(function(el) {{
    var key = el.getAttribute('data-i18n');
    if (dict[key] !== undefined) el.innerHTML = dict[key];
  }});
  document.documentElement.lang = lang;
}}

// Expose for parent iframe access
document.__applyLang = __applyLang;
window.__applyLang = __applyLang;

// Auto-apply saved language
(function() {{
  try {{
    var lang = (window.parent !== window) ? null : localStorage.getItem('appLang');
    if (lang) __applyLang(lang);
  }} catch(e) {{}}
}})();

{auto_activate_js}
</script>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"\n統合HTMLレポート出力: {output_path}")


# ████████████████████████████████████████████████████████████
# PART 4: Excel出力
# ████████████████████████████████████████████████████████████

def generate_excel(dogs: list, pedigrees: list, output_path: str):
    """Excel スプレッドシートを生成"""
    if not HAS_OPENPYXL:
        print("  openpyxl が未インストールのためExcel出力をスキップします。")
        return

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="4A1A7A")
    normal_fill = PatternFill("solid", fgColor="DCFCE7")
    carrier_fill = PatternFill("solid", fgColor="FEF3C7")
    positive_fill = PatternFill("solid", fgColor="FEE2E2")
    trait_fill = PatternFill("solid", fgColor="E0E7FF")
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    def style_cell(ws, row, col, status=None):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if status == "normal":
            cell.fill = normal_fill
        elif status == "carrier":
            cell.fill = carrier_fill
        elif status == "positive":
            cell.fill = positive_fill
        elif status == "trait":
            cell.fill = trait_fill

    # ── Sheet 1: Overview ──
    if dogs:
        ws = wb.active
        ws.title = "概要"
        headers = ["ペット名", "登録名", "犬種", "性別", "生年月日", "マイクロチップ", "ケース番号", "検査日"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        style_header(ws, 1, len(headers))

        for r, dog in enumerate(dogs, 2):
            data = [dog.pet_name, dog.registered_name, dog.breed, dog.sex, dog.dob, dog.microchip, dog.case_number, dog.test_date]
            for c, val in enumerate(data, 1):
                ws.cell(row=r, column=c, value=val)
                style_cell(ws, r, c)

        for col_letter in ['A','B','C','D','E','F','G','H']:
            ws.column_dimensions[col_letter].width = 22

        # ── Per-dog sheets ──
        for dog in dogs:
            name = (dog.pet_name or dog.registered_name or "犬")[:30]
            ws = wb.create_sheet(title=name)

            ws.cell(row=1, column=1, value="ペット名").font = Font(bold=True, name="Arial")
            ws.cell(row=1, column=2, value=dog.pet_name)
            ws.cell(row=2, column=1, value="登録名").font = Font(bold=True, name="Arial")
            ws.cell(row=2, column=2, value=dog.registered_name)
            ws.cell(row=3, column=1, value="犬種").font = Font(bold=True, name="Arial")
            ws.cell(row=3, column=2, value=dog.breed)
            ws.cell(row=4, column=1, value="性別").font = Font(bold=True, name="Arial")
            ws.cell(row=4, column=2, value=dog.sex)
            ws.cell(row=5, column=1, value="生年月日").font = Font(bold=True, name="Arial")
            ws.cell(row=5, column=2, value=dog.dob)
            ws.cell(row=6, column=1, value="マイクロチップ").font = Font(bold=True, name="Arial")
            ws.cell(row=6, column=2, value=dog.microchip)
            ws.cell(row=7, column=1, value="ケース番号").font = Font(bold=True, name="Arial")
            ws.cell(row=7, column=2, value=dog.case_number)

            row = 9
            ws.cell(row=row, column=1, value="【健康検査結果】").font = Font(bold=True, size=12, color="4A1A7A", name="Arial")
            row = 10
            for c, h in enumerate(["カテゴリー", "検査項目", "検査項目(英語)", "遺伝子型", "ステータス", "詳細"], 1):
                ws.cell(row=row, column=c, value=h)
            style_header(ws, row, 6)

            for r_idx, result in enumerate(dog.health_results):
                r = row + 1 + r_idx
                ws.cell(row=r, column=1, value=sanitize_for_excel(result.category))
                ws.cell(row=r, column=2, value=sanitize_for_excel(result.japanese_name or result.test_name))
                ws.cell(row=r, column=3, value=sanitize_for_excel(result.test_name))
                ws.cell(row=r, column=4, value=sanitize_for_excel(result.genotype))
                status_jp = {"normal": "ノーマル", "carrier": "キャリア", "positive": "ポジティブ"}.get(result.status, result.status)
                ws.cell(row=r, column=5, value=status_jp)
                ws.cell(row=r, column=6, value=sanitize_for_excel(result.result_text[:150]))
                for c in range(1, 7):
                    style_cell(ws, r, c, result.status)

            row = row + len(dog.health_results) + 3
            ws.cell(row=row, column=1, value="【毛色・形質検査結果】").font = Font(bold=True, size=12, color="4A1A7A", name="Arial")
            row += 1
            for c, h in enumerate(["検査項目", "検査項目(英語)", "遺伝子型", "詳細"], 1):
                ws.cell(row=row, column=c, value=h)
            style_header(ws, row, 4)

            for r_idx, result in enumerate(dog.trait_results):
                r = row + 1 + r_idx
                ws.cell(row=r, column=1, value=sanitize_for_excel(result.japanese_name or result.test_name))
                ws.cell(row=r, column=2, value=sanitize_for_excel(result.test_name))
                ws.cell(row=r, column=3, value=sanitize_for_excel(result.genotype))
                ws.cell(row=r, column=4, value=sanitize_for_excel(result.result_text[:150]))
                for c in range(1, 5):
                    style_cell(ws, r, c, "trait")

            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 60

        # ── Comparison Sheet ──
        if len(dogs) > 1:
            ws = wb.create_sheet(title="比較表")
            headers = ["検査項目"] + [d.pet_name or d.registered_name for d in dogs]
            for c, h in enumerate(headers, 1):
                ws.cell(row=1, column=c, value=h)
            style_header(ws, 1, len(headers))

            all_tests = {}
            for dog in dogs:
                for r in dog.health_results:
                    key = r.test_name
                    if key not in all_tests:
                        all_tests[key] = r.japanese_name or r.test_name

            row = 2
            for test_key, jp_name in all_tests.items():
                ws.cell(row=row, column=1, value=sanitize_for_excel(jp_name))
                style_cell(ws, row, 1)
                for d_idx, dog in enumerate(dogs):
                    col = d_idx + 2
                    found = False
                    for r in dog.health_results:
                        if r.test_name == test_key:
                            ws.cell(row=row, column=col, value=r.genotype)
                            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                            style_cell(ws, row, col, r.status)
                            found = True
                            break
                    if not found:
                        ws.cell(row=row, column=col, value="—")
                        style_cell(ws, row, col)
                row += 1

            ws.column_dimensions['A'].width = 40
            for i in range(len(dogs)):
                ws.column_dimensions[chr(66 + i)].width = 18

    # ── Pedigree Sheet ──
    if pedigrees:
        if not dogs:
            ws = wb.active
            ws.title = "血統書"
        else:
            ws = wb.create_sheet(title="血統書")

        row = 1
        for ped in pedigrees:
            ws.cell(row=row, column=1, value=f"【{ped.dog_name}】").font = Font(bold=True, size=14, color="4A1A7A", name="Arial")
            row += 1
            info_items = [
                ("犬種", ped.breed), ("登録番号", ped.registration),
                ("性別", ped.sex), ("生年月日", ped.dob),
                ("毛色", ped.color), ("マイクロチップ", ped.microchip),
                ("ブリーダー", ped.breeder), ("オーナー", ped.owner),
            ]
            for label, val in info_items:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True, name="Arial")
                ws.cell(row=row, column=2, value=val)
                row += 1

            row += 1
            for c, h in enumerate(["位置", "犬名", "登録番号", "毛色", "タイトル", "DNA番号"], 1):
                ws.cell(row=row, column=c, value=h)
            style_header(ws, row, 6)

            for label, anc in ped.all_ancestors():
                if anc:
                    row += 1
                    ws.cell(row=row, column=1, value=label)
                    ws.cell(row=row, column=2, value=anc.name)
                    ws.cell(row=row, column=3, value=anc.registration)
                    ws.cell(row=row, column=4, value=anc.color)
                    ws.cell(row=row, column=5, value=anc.titles)
                    ws.cell(row=row, column=6, value=anc.dna_number)
                    for c in range(1, 7):
                        style_cell(ws, row, c)

            # COI
            coi_result = calc_coi_3gen(ped)
            row += 2
            ws.cell(row=row, column=1, value="近親交配係数 (COI)").font = Font(bold=True, size=12, color="4A1A7A", name="Arial")
            row += 1
            ws.cell(row=row, column=1, value="Wright's COI (3世代)")
            ws.cell(row=row, column=2, value=f"{coi_result['coi_pct']:.2f}%")
            if coi_result['coi_pct'] > 6.25:
                ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000", name="Arial")
            else:
                ws.cell(row=row, column=2).font = Font(bold=True, color="22C55E", name="Arial")

            if coi_result['common_ancestors']:
                row += 1
                ws.cell(row=row, column=1, value="共通祖先")
                names = ", ".join([c['name'] for c in coi_result['common_ancestors']])
                ws.cell(row=row, column=2, value=names)

            row += 3

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 18

    wb.save(output_path)
    print(f"Excel レポート出力: {output_path}")


# ████████████████████████████████████████████████████████████
# PART 5: メインエントリポイント
# ████████████████████████████████████████████████████████████

def collect_pdf_files(args: list) -> list:
    """引数からPDFファイルリストを収集"""
    pdf_files = []
    for arg in args:
        if arg in ('--output', '-o', '--pedigree', '-p'):
            continue
        if os.path.isdir(arg):
            pdf_files.extend(glob.glob(os.path.join(arg, "*.pdf")))
            pdf_files.extend(glob.glob(os.path.join(arg, "*.PDF")))
        elif os.path.isfile(arg) and arg.lower().endswith('.pdf'):
            pdf_files.append(arg)
        elif '*' in arg or '?' in arg:
            pdf_files.extend(glob.glob(arg))
    return sorted(set(pdf_files))


def collect_pedigree_args(args: list) -> list:
    """血統書関連の引数を収集"""
    pedigree_sources = []
    i = 0
    while i < len(args):
        if args[i] in ('--pedigree', '-p') and i + 1 < len(args):
            pedigree_sources.append(args[i + 1])
            i += 2
        else:
            i += 1
    return pedigree_sources


def get_output_dir(args: list) -> str:
    """出力ディレクトリを取得"""
    for i, arg in enumerate(args):
        if arg in ('--output', '-o') and i + 1 < len(args):
            return args[i + 1]
    return os.getcwd()


def print_usage():
    print("""
使い方:
  python poodle_genetics.py <コマンド> [オプション]

コマンド:
  orivet <PDFファイル...>              Orivet遺伝子検査PDFを解析
  pedigree <写真 or 犬名>             血統書解析 + COI算出
  all <PDFファイル...> -p <犬名>       遺伝子 + 血統書の統合レポート
  demo                                 デモ実行 (Sevenの全データ)

オプション:
  -o, --output <ディレクトリ>   出力先ディレクトリ
  -p, --pedigree <犬名/写真>    血統書データ (複数指定可)

例:
  python poodle_genetics.py orivet *.pdf -o ./output
  python poodle_genetics.py pedigree seven -o ./output
  python poodle_genetics.py all *.pdf -p seven -o ./output
  python poodle_genetics.py demo
""")


def main():
    print("=" * 60)
    print("  Orivet 遺伝子解析ツール — 全犬種対応")
    print("=" * 60)

    if len(sys.argv) < 2:
        print_usage()
        sys.exit(1)

    command = sys.argv[1].lower()
    remaining_args = sys.argv[2:]

    output_dir = get_output_dir(remaining_args)
    os.makedirs(output_dir, exist_ok=True)

    dogs = []       # Orivet解析結果
    pedigrees = []  # 血統書データ

    if command == "demo":
        # デモモード: PDFがあれば解析、なければサンプルメッセージ
        print("\n[デモモード] Sevenの既知データで実行します。")

        # Try to parse PDFs in uploads
        upload_dir = "/sessions/sleepy-practical-cannon/mnt/uploads"
        if os.path.isdir(upload_dir):
            pdf_files = glob.glob(os.path.join(upload_dir, "*.pdf")) + glob.glob(os.path.join(upload_dir, "*.PDF"))
            if pdf_files and HAS_PDFPLUMBER:
                print(f"\n{len(pdf_files)} 個のPDFファイルを検出。解析中...")
                for pdf_path in pdf_files:
                    dog = parse_pdf(pdf_path)
                    if dog:
                        dogs.append(dog)
                if dogs:
                    print(f"  → {len(dogs)} 頭分のOrivetデータを解析しました。")

        # Add Seven's pedigree
        pedigrees.append(KNOWN_PEDIGREES["seven"])

    elif command == "orivet":
        if not HAS_PDFPLUMBER:
            print("\nエラー: pdfplumber が必要です。\n  pip install pdfplumber")
            sys.exit(1)

        pdf_files = collect_pdf_files(remaining_args)

        if not pdf_files:
            pdf_files = glob.glob("*.pdf") + glob.glob("*.PDF")

        if not pdf_files:
            print("\nエラー: PDFファイルが見つかりません。")
            sys.exit(1)

        print(f"\n{len(pdf_files)} 個のPDFファイルを検出。\n")
        for pdf_path in pdf_files:
            dog = parse_pdf(pdf_path)
            if dog:
                dogs.append(dog)

        if not dogs:
            print("\nエラー: 解析可能なOrivetレポートが見つかりませんでした。")
            sys.exit(1)

    elif command == "pedigree":
        pedigree_sources = remaining_args
        # Remove output flag args
        clean_sources = []
        skip_next = False
        for a in pedigree_sources:
            if skip_next:
                skip_next = False
                continue
            if a in ('--output', '-o'):
                skip_next = True
                continue
            clean_sources.append(a)

        if not clean_sources:
            clean_sources = ["seven"]

        for src in clean_sources:
            if src.lower() in KNOWN_PEDIGREES:
                pedigrees.append(KNOWN_PEDIGREES[src.lower()])
                print(f"  → 既知の血統書データ: {KNOWN_PEDIGREES[src.lower()].dog_name}")
            elif os.path.isfile(src):
                print(f"\n画像を解析中: {src}")
                text = try_ocr(src)
                if text:
                    ped = parse_jkc_pedigree_text(text)
                    if ped and ped.dog_name:
                        pedigrees.append(ped)
                        print(f"  → {ped.dog_name} の血統書を解析しました。")
                    else:
                        print("  → 血統書データの解析に失敗しました。")
            else:
                print(f"  → 不明な引数: {src}")

    elif command == "all":
        # 統合モード
        if HAS_PDFPLUMBER:
            pdf_files = collect_pdf_files(remaining_args)
            if pdf_files:
                print(f"\n{len(pdf_files)} 個のPDFファイルを検出。\n")
                for pdf_path in pdf_files:
                    dog = parse_pdf(pdf_path)
                    if dog:
                        dogs.append(dog)

        # Pedigree sources
        ped_sources = collect_pedigree_args(remaining_args)
        if not ped_sources:
            ped_sources = ["seven"]

        for src in ped_sources:
            if src.lower() in KNOWN_PEDIGREES:
                pedigrees.append(KNOWN_PEDIGREES[src.lower()])
            elif os.path.isfile(src):
                text = try_ocr(src)
                if text:
                    ped = parse_jkc_pedigree_text(text)
                    if ped and ped.dog_name:
                        pedigrees.append(ped)

    else:
        print(f"\n不明なコマンド: {command}")
        print_usage()
        sys.exit(1)

    if not dogs and not pedigrees:
        print("\nエラー: 解析可能なデータがありません。")
        sys.exit(1)

    # ── 出力 ──
    print(f"\n{'='*40}")
    if dogs:
        print(f"Orivet解析: {len(dogs)} 頭")
    if pedigrees:
        print(f"血統書データ: {len(pedigrees)} 頭")
        for ped in pedigrees:
            result = calc_coi_3gen(ped)
            print(f"  {ped.dog_name}: COI = {result['coi_pct']:.2f}%")

    html_path = os.path.join(output_dir, "orivet_report.html")
    xlsx_path = os.path.join(output_dir, "orivet_report.xlsx")

    generate_unified_html(dogs, pedigrees, html_path)
    generate_excel(dogs, pedigrees, xlsx_path)

    print("\n完了! 以下のファイルが生成されました:")
    print(f"  HTML: {html_path}")
    print(f"  Excel: {xlsx_path}")
    print()


if __name__ == "__main__":
    main()
